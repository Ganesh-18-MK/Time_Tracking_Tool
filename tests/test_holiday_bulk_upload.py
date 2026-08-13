"""Bulk holiday upload: one sheet, creates or updates Holiday rows directly
(no employee to match against — see app/holiday_bulk_upload.py's module
docstring). Same in-memory SQLite fixture pattern as test_leave_bulk_upload.py.

Holidays are one shared company-wide list (Ganesh, 2026-08-14 — reverted
the brief 2026-08-12 per-country split), so the sheet is just Holiday
Name + Holiday Date; there's no Country column anymore.
"""
import datetime as dt

import pytest
from openpyxl import Workbook
from sqlalchemy import create_engine, select
from sqlalchemy.orm import sessionmaker

from app import models as m
from app.db import Base
from app.holiday_bulk_upload import (
    build_existing_holidays_workbook,
    build_sample_workbook,
    parse_row,
    process_upload,
    read_upload_rows,
)


class TestParseRow:
    def test_happy_path(self):
        row = {"Holiday Name": "Diwali", "Holiday Date": "2026-11-08"}
        result = parse_row(row)
        assert result["mode"] == "ok"
        assert result["date"] == dt.date(2026, 11, 8)
        assert result["name"] == "Diwali"

    def test_excel_date_cell_passes_through(self):
        row = {"Holiday Name": "Independence Day", "Holiday Date": dt.date(2026, 7, 4)}
        result = parse_row(row)
        assert result["mode"] == "ok"
        assert result["date"] == dt.date(2026, 7, 4)

    def test_missing_date_is_an_error(self):
        row = {"Holiday Name": "X", "Holiday Date": None}
        result = parse_row(row)
        assert result["mode"] == "error"
        assert "Date is required" in result["error"]

    def test_blank_name_is_allowed(self):
        row = {"Holiday Name": "", "Holiday Date": "2026-07-04"}
        result = parse_row(row)
        assert result["mode"] == "ok"
        assert result["name"] == ""

    def test_bad_date_format_is_an_error(self):
        row = {"Holiday Name": "X", "Holiday Date": "not a date"}
        result = parse_row(row)
        assert result["mode"] == "error"


class TestSampleAndExportWorkbooks:
    def test_sample_workbook_rows_parse_cleanly(self):
        wb = build_sample_workbook()
        rows, header_error = read_upload_rows(wb)
        assert header_error is None
        assert len(rows) == 2
        for r in rows:
            result = parse_row(r)
            assert result["mode"] == "ok"

    def test_empty_sheet_is_a_sheet_level_error(self):
        wb = Workbook()
        rows, error = read_upload_rows(wb)
        assert rows == [] and error is not None

    def test_blank_rows_skipped_silently(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["Holiday Name", "Holiday Date"])
        ws.append(["Diwali", dt.date(2026, 11, 8)])
        ws.append([None, None])
        rows, error = read_upload_rows(wb)
        assert error is None
        assert len(rows) == 1


@pytest.fixture()
def db():
    eng = create_engine("sqlite://")
    Base.metadata.create_all(eng)
    s = sessionmaker(bind=eng)()
    yield s
    s.close()


def _wb_from_rows(header, rows):
    wb = Workbook()
    ws = wb.active
    ws.append(header)
    for r in rows:
        ws.append(r)
    return wb


class TestProcessUpload:
    def test_valid_row_creates_holiday(self, db):
        wb = _wb_from_rows(
            ["Holiday Name", "Holiday Date"],
            [["Diwali", dt.date(2026, 11, 8)]],
        )
        result = process_upload(db, wb)
        assert result["header_error"] is None
        assert result["added"] == 1
        assert result["updated"] == 0
        row = db.execute(select(m.Holiday)).scalar_one()
        assert row.name == "Diwali"
        # location is stamped on invisibly for schema reasons only (see
        # Holiday's docstring) — never surfaced or read back by the sheet.
        assert row.location == m.DEFAULT_LOCATION

    def test_two_different_dates_both_created(self, db):
        wb = _wb_from_rows(
            ["Holiday Name", "Holiday Date"],
            [["Christmas", dt.date(2026, 12, 25)], ["New Year's Day", dt.date(2027, 1, 1)]],
        )
        result = process_upload(db, wb)
        assert result["added"] == 2
        rows = list(db.execute(select(m.Holiday)).scalars())
        assert {r.name for r in rows} == {"Christmas", "New Year's Day"}

    def test_re_upload_same_date_updates_name_not_duplicates(self, db):
        db.add(m.Holiday(date=dt.date(2026, 11, 8), name="Old Name", location=m.DEFAULT_LOCATION))
        db.commit()
        wb = _wb_from_rows(
            ["Holiday Name", "Holiday Date"],
            [["Diwali (fixed)", dt.date(2026, 11, 8)]],
        )
        result = process_upload(db, wb)
        assert result["added"] == 0
        assert result["updated"] == 1
        rows = list(db.execute(select(m.Holiday)).scalars())
        assert len(rows) == 1
        assert rows[0].name == "Diwali (fixed)"

    def test_second_row_same_date_updates_rather_than_duplicates(self, db):
        """Two rows in the same sheet upload sharing a date (e.g. leftover
        habit from the old per-country sheet) should land as one holiday,
        not two — dedup key is date alone now."""
        wb = _wb_from_rows(
            ["Holiday Name", "Holiday Date"],
            [["First", dt.date(2026, 12, 25)], ["Second", dt.date(2026, 12, 25)]],
        )
        result = process_upload(db, wb)
        assert result["added"] == 1
        assert result["updated"] == 1
        rows = list(db.execute(select(m.Holiday)).scalars())
        assert len(rows) == 1
        assert rows[0].name == "Second"

    def test_multiple_rows_mixed_valid_and_invalid(self, db):
        wb = _wb_from_rows(
            ["Holiday Name", "Holiday Date"],
            [["Good", dt.date(2026, 7, 4)], ["Bad", None]],
        )
        result = process_upload(db, wb)
        assert result["added"] == 1
        assert len(result["skipped"]) == 1

    def test_too_many_rows_rejected_at_sheet_level(self, db):
        rows = [["X", dt.date(2026, 1, 1)] for _ in range(1001)]
        wb = _wb_from_rows(["Holiday Name", "Holiday Date"], rows)
        result = process_upload(db, wb)
        assert result["header_error"] is not None
        assert "max is 1000" in result["header_error"]

    def test_export_round_trips_into_a_valid_row(self, db):
        db.add(m.Holiday(date=dt.date(2026, 11, 8), name="Diwali", location=m.DEFAULT_LOCATION))
        db.commit()
        wb = build_existing_holidays_workbook(db)
        rows, error = read_upload_rows(wb)
        assert error is None
        assert len(rows) == 1
        result = parse_row(rows[0])
        assert result["mode"] == "ok"
        assert result["name"] == "Diwali"

    def test_existing_rows_from_old_per_country_era_dont_block_matching(self, db):
        """Leftover reality check: a date that already has two rows from the
        brief per-country era (one US, one India) shouldn't crash the
        upload or create a third — the new row just matches whichever one
        the dedup dict happened to key first."""
        db.add(m.Holiday(date=dt.date(2026, 7, 4), name="Independence Day", location="US"))
        db.add(m.Holiday(date=dt.date(2026, 7, 4), name="Independence Day (India copy)", location="India"))
        db.commit()
        wb = _wb_from_rows(
            ["Holiday Name", "Holiday Date"],
            [["Independence Day (renamed)", dt.date(2026, 7, 4)]],
        )
        result = process_upload(db, wb)
        assert result["added"] == 0
        assert result["updated"] == 1
        # still two rows total — the upload updated one of them, didn't add a third
        rows = list(db.execute(select(m.Holiday)).scalars())
        assert len(rows) == 2
