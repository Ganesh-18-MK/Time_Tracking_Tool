"""Roster bulk upload: one file, two behaviors told apart by Employee ID.

Blank Employee ID -> onboard a new hire (Full Name/Department/Designation/
Target/day/Workdays required). Filled Employee ID -> update that existing
employee (nothing required; blank cells are left untouched, not defaulted).

Row-level parsing is pure and tested directly. process_upload() needs a
database for employee-code lookups and the duplicate/collision checks, so it
gets the same in-memory SQLite fixture pattern used in test_engine.py.
"""
import datetime as dt

import pytest
from openpyxl import Workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app import models as m
from app.bulk_upload import (
    build_existing_employees_workbook,
    build_sample_workbook,
    parse_action,
    parse_cell_date,
    parse_role,
    parse_row,
    parse_target_minutes,
    parse_workdays,
    process_upload,
    read_upload_rows,
    workdays_to_letters,
)
from app.db import Base
from app.util import format_employee_code, next_employee_code


class TestParseWorkdays:
    def test_letter_codes_mon_fri(self):
        assert parse_workdays("M,T,W,Th,F") == "0,1,2,3,4"

    def test_blank_is_none_not_a_default(self):
        # unlike the first cut of this feature, blank no longer means
        # Mon-Fri here — the caller (parse_row) decides what blank means
        # depending on new-hire vs update mode.
        assert parse_workdays("") is None
        assert parse_workdays(None) is None

    def test_weekend_letters(self):
        assert parse_workdays("Su,S") == "5,6"

    def test_three_letter_aliases_case_insensitive(self):
        assert parse_workdays("mon,tue,WED") == "0,1,2"

    def test_bad_token_raises_with_token_named(self):
        with pytest.raises(ValueError, match="xx"):
            parse_workdays("M,Xx")

    def test_t_and_th_are_distinct(self):
        assert parse_workdays("T") == "1"
        assert parse_workdays("Th") == "3"


class TestParseTargetMinutes:
    def test_blank_is_none_not_a_default(self):
        assert parse_target_minutes("") is None
        assert parse_target_minutes(None) is None

    def test_plain_int_hours(self):
        assert parse_target_minutes(8) == 480

    def test_decimal_hours(self):
        assert parse_target_minutes("7.5") == 450

    def test_hmm_text(self):
        assert parse_target_minutes("8:00") == 480
        assert parse_target_minutes("4:19") == 259

    def test_non_numeric_raises(self):
        with pytest.raises(ValueError):
            parse_target_minutes("not a number")

    def test_zero_or_negative_raises(self):
        with pytest.raises(ValueError):
            parse_target_minutes(0)


class TestParseRole:
    def test_blank_is_none_not_a_default(self):
        assert parse_role("") is None
        assert parse_role(None) is None

    def test_employee_case_insensitive(self):
        # (is_admin, is_super_admin) — see role_to_flags in app/util.py
        assert parse_role("employee") == (False, False)
        assert parse_role("Employee") == (False, False)

    def test_admin_case_insensitive(self):
        # department-scoped admin tier: is_admin=True, is_super_admin=False
        assert parse_role("admin") == (True, False)
        assert parse_role("Admin") == (True, False)

    def test_anything_else_raises(self):
        with pytest.raises(ValueError):
            parse_role("Manager")


class TestParseAction:
    def test_blank_is_none(self):
        assert parse_action("") is None
        assert parse_action(None) is None

    def test_deactivate_case_insensitive(self):
        assert parse_action("Deactivate") is True
        assert parse_action("deactivate") is True

    def test_offboard_is_an_alias(self):
        assert parse_action("Offboard") is True

    def test_anything_else_raises(self):
        with pytest.raises(ValueError):
            parse_action("Delete")


class TestParseCellDate:
    def test_iso_string(self):
        assert parse_cell_date("2026-08-03") == dt.date(2026, 8, 3)

    def test_slash_format_fallback(self):
        assert parse_cell_date("03/08/2026") == dt.date(2026, 8, 3)

    def test_date_and_datetime_passthrough(self):
        assert parse_cell_date(dt.date(2026, 8, 3)) == dt.date(2026, 8, 3)
        assert parse_cell_date(dt.datetime(2026, 8, 3, 9, 0)) == dt.date(2026, 8, 3)

    def test_blank_is_none(self):
        assert parse_cell_date("") is None
        assert parse_cell_date(None) is None

    def test_garbage_raises(self):
        with pytest.raises(ValueError):
            parse_cell_date("not-a-date")


class TestWorkdaysToLetters:
    def test_round_trips_with_parse_workdays(self):
        assert workdays_to_letters("0,1,2,3,4") == "M,T,W,Th,F"
        assert parse_workdays(workdays_to_letters("0,1,2,3,4")) == "0,1,2,3,4"


NEW_ROW = {
    "Employee ID": "", "Full Name": "Jane Doe", "Department": "Accounts",
    "Designation": "Associate", "Target/day": 8, "Workdays": "M,T,W,Th,F",
    "Joining Date": "2026-08-03", "DOB": "1990-05-14",
    "Email": "jane.doe@example.com", "Phone": "+1 555 0100", "Role": "Employee",
}


class TestParseRowNewHire:
    def test_happy_path(self):
        result = parse_row(NEW_ROW, {})
        assert result["error"] is None
        assert result["mode"] == "new"
        f = result["fields"]
        assert f["name"] == "Jane Doe" and f["daily_target_minutes"] == 480
        assert f["work_days"] == "0,1,2,3,4" and f["is_admin"] is False

    def test_missing_full_name_required(self):
        row = dict(NEW_ROW, **{"Full Name": ""})
        result = parse_row(row, {})
        assert result["mode"] == "error"
        assert "Full Name" in result["error"]

    def test_missing_department_designation_target_workdays_all_required(self):
        row = dict(NEW_ROW, **{"Department": "", "Designation": "", "Target/day": "", "Workdays": ""})
        result = parse_row(row, {})
        assert result["mode"] == "error"
        for field in ["Department", "Target/day", "Workdays"]:
            assert field in result["error"], result["error"]

    def test_optional_fields_can_be_blank(self):
        row = dict(NEW_ROW, **{"Joining Date": "", "DOB": "", "Email": "", "Phone": "", "Role": ""})
        result = parse_row(row, {})
        assert result["error"] is None
        f = result["fields"]
        assert "start_date" not in f and "date_of_birth" not in f
        assert "email" not in f and "phone" not in f
        assert f["is_admin"] is False  # Role still defaults for new hires

    def test_country_code_captured_when_present(self):
        row = dict(NEW_ROW, **{"Country Code": "+91"})
        result = parse_row(row, {})
        assert result["error"] is None
        assert result["fields"]["country_code"] == "+91"

    def test_country_code_optional_and_independent_of_phone(self):
        row = dict(NEW_ROW, **{"Country Code": "", "Phone": "9876543210"})
        result = parse_row(row, {})
        assert result["error"] is None
        f = result["fields"]
        assert "country_code" not in f
        assert f["phone"] == "9876543210"

    def test_bad_email_flagged(self):
        row = dict(NEW_ROW, **{"Email": "not-an-email"})
        result = parse_row(row, {})
        assert result["mode"] == "error"
        assert "mail" in result["error"].lower()

    def test_multiple_errors_all_reported_at_once(self):
        row = {"Employee ID": "", "Full Name": "", "Target/day": "junk", "Workdays": "Zz", "Role": "boss"}
        result = parse_row(row, {})
        assert result["mode"] == "error"
        assert result["error"].count(";") >= 3


class TestParseRowUpdate:
    def test_unknown_employee_id_is_an_error(self):
        row = {"Employee ID": "LOMK999", "Phone": "123"}
        result = parse_row(row, {"LOMK001": 1})
        assert result["mode"] == "error"
        assert "not found" in result["error"]

    def test_known_id_with_only_phone_is_a_minimal_patch(self):
        row = {"Employee ID": "lomk001", "Phone": "+1 555 9999"}  # case-insensitive match
        result = parse_row(row, {"LOMK001": 1})
        assert result["error"] is None
        assert result["mode"] == "update"
        assert result["employee_id"] == 1
        assert result["fields"] == {"phone": "+1 555 9999"}

    def test_nothing_required_on_update(self):
        row = {"Employee ID": "LOMK001"}
        result = parse_row(row, {"LOMK001": 1})
        assert result["error"] is None
        assert result["fields"] == {}

    def test_blank_role_on_update_is_not_included_in_patch(self):
        row = {"Employee ID": "LOMK001", "Phone": "123"}
        result = parse_row(row, {"LOMK001": 1})
        assert "is_admin" not in result["fields"]

    def test_provided_role_on_update_is_included(self):
        row = {"Employee ID": "LOMK001", "Role": "Admin"}
        result = parse_row(row, {"LOMK001": 1})
        assert result["fields"]["is_admin"] is True

    def test_invalid_field_on_update_still_reports_error(self):
        row = {"Employee ID": "LOMK001", "Target/day": "garbage"}
        result = parse_row(row, {"LOMK001": 1})
        assert result["mode"] == "error"

    def test_country_code_alone_is_a_minimal_patch(self):
        row = {"Employee ID": "LOMK001", "Country Code": "+44"}
        result = parse_row(row, {"LOMK001": 1})
        assert result["error"] is None
        assert result["fields"] == {"country_code": "+44"}

    def test_blank_country_code_on_update_not_included_in_patch(self):
        row = {"Employee ID": "LOMK001", "Phone": "123"}
        result = parse_row(row, {"LOMK001": 1})
        assert "country_code" not in result["fields"]


class TestSampleAndExportWorkbooks:
    def test_sample_workbook_new_hire_row_parses_cleanly(self):
        wb = build_sample_workbook()
        rows, header_error = read_upload_rows(wb)
        assert header_error is None
        assert len(rows) == 1
        result = parse_row(rows[0], {})
        assert result["error"] is None
        assert result["mode"] == "new"

    def test_sample_workbook_has_country_code_column(self):
        wb = build_sample_workbook()
        rows, _ = read_upload_rows(wb)
        result = parse_row(rows[0], {})
        assert result["fields"]["country_code"]
        assert result["fields"]["phone"]

    def test_reordered_case_insensitive_headers_still_match(self):
        # Header order/case shouldn't matter for matching; all 5 fields
        # required for a new hire are still supplied here so this isolates
        # the header-matching behavior from the required-fields check
        # (covered separately in TestParseRowNewHire).
        wb = Workbook()
        ws = wb.active
        ws.append(["full name", "EMAIL", "Role", "department", "DESIGNATION", "target/day", "workdays"])
        ws.append(["Bob Smith", "bob@example.com", "Admin", "Tech", "Associate", 8, "M,T,W,Th,F"])
        rows, error = read_upload_rows(wb)
        assert error is None
        result = parse_row(rows[0], {})
        assert result["error"] is None

    def test_missing_columns_are_not_a_sheet_level_error(self):
        # no required-header check anymore — requiredness is row-level and
        # mode-dependent (see parse_row), so a sheet with just ID+Phone for
        # updates is perfectly valid even with no Full Name column at all.
        wb = Workbook()
        ws = wb.active
        ws.append(["Employee ID", "Phone"])
        ws.append(["LOMK001", "123"])
        rows, error = read_upload_rows(wb)
        assert error is None
        assert len(rows) == 1

    def test_empty_sheet_is_a_sheet_level_error(self):
        wb = Workbook()
        rows, error = read_upload_rows(wb)
        assert rows == [] and error is not None

    def test_blank_rows_are_skipped_silently(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["Full Name", "Email"])
        ws.append(["Bob", "bob@example.com"])
        ws.append([None, None])
        rows, error = read_upload_rows(wb)
        assert error is None
        assert len(rows) == 1


class TestExistingEmployeesWorkbook:
    def test_export_round_trips_into_a_valid_update_row(self, db):
        db.add(m.Employee(
            name="Jane Doe", department="Accounts", designation="Associate",
            daily_target_minutes=450, work_days="0,1,2,3,4", email="jane@example.com",
            country_code="+1", phone="555 0100", employee_code="LOMK001", is_admin=False, active=True,
        ))
        db.commit()
        wb = build_existing_employees_workbook(db)
        rows, error = read_upload_rows(wb)
        assert error is None
        assert len(rows) == 1
        result = parse_row(rows[0], {"LOMK001": 1})
        assert result["mode"] == "update"
        assert result["error"] is None

    def test_country_code_and_phone_export_as_separate_columns(self, db):
        db.add(m.Employee(
            name="Jane Doe", employee_code="LOMK001", country_code="+91", phone="9876543210",
        ))
        db.commit()
        wb = build_existing_employees_workbook(db)
        rows, _ = read_upload_rows(wb)
        assert rows[0]["Country Code"] == "+91"
        assert rows[0]["Phone"] == "9876543210"

    def test_deactivated_employees_excluded(self, db):
        db.add(m.Employee(name="Gone", employee_code="LOMK001", active=False))
        db.commit()
        wb = build_existing_employees_workbook(db)
        rows, _ = read_upload_rows(wb)
        assert rows == []


@pytest.fixture()
def db():
    eng = create_engine("sqlite://")
    Base.metadata.create_all(eng)
    s = sessionmaker(bind=eng)()
    yield s
    s.close()


def _wb_from_rows(header, rows):
    wb = Workbook()
    ws = wb.active
    ws.append(header)
    for r in rows:
        ws.append(r)
    return wb


class TestProcessUploadOnboarding:
    def test_valid_new_row_added_with_generated_code(self, db):
        wb = _wb_from_rows(
            ["Full Name", "Department", "Designation", "Target/day", "Workdays"],
            [["Alice Kim", "Accounts", "Associate", 8, "M,T,W,Th,F"]],
        )
        result = process_upload(db, wb)
        assert result["header_error"] is None
        assert result["added"] == 1 and result["updated"] == 0
        emp = db.query(m.Employee).filter_by(name="Alice Kim").one()
        assert emp.employee_code == "LOMK001"

    def test_missing_required_field_skips_with_reason_not_defaulted(self, db):
        wb = _wb_from_rows(["Full Name"], [["Bob"]])
        result = process_upload(db, wb)
        assert result["added"] == 0
        assert len(result["skipped"]) == 1
        assert "required" in result["skipped"][0]["reason"].lower()

    def test_skips_row_matching_existing_email(self, db):
        db.add(m.Employee(name="Existing", email="dup@example.com"))
        db.commit()
        wb = _wb_from_rows(
            ["Full Name", "Department", "Designation", "Target/day", "Workdays", "Email"],
            [["New Name", "Ops", "Lead", 8, "M,T,W,Th,F", "dup@example.com"]],
        )
        result = process_upload(db, wb)
        assert result["added"] == 0
        assert "already exists" in result["skipped"][0]["reason"].lower()

    def test_sequential_codes_continue_from_highest_existing(self, db):
        db.add(m.Employee(name="Old", employee_code="LOMK007"))
        db.commit()
        wb = _wb_from_rows(
            ["Full Name", "Department", "Designation", "Target/day", "Workdays"],
            [["New One", "Ops", "Lead", 8, "M,T,W,Th,F"]],
        )
        process_upload(db, wb)
        new_emp = db.query(m.Employee).filter_by(name="New One").one()
        assert new_emp.employee_code == "LOMK008"

    def test_two_new_rows_in_one_batch_get_distinct_sequential_codes(self, db):
        wb = _wb_from_rows(
            ["Full Name", "Department", "Designation", "Target/day", "Workdays"],
            [
                ["Person A", "Ops", "Lead", 8, "M,T,W,Th,F"],
                ["Person B", "Ops", "Lead", 8, "M,T,W,Th,F"],
            ],
        )
        result = process_upload(db, wb)
        assert result["added"] == 2
        codes = sorted(e.employee_code for e in db.query(m.Employee).all())
        assert codes == ["LOMK001", "LOMK002"]


class TestProcessUploadUpdating:
    def test_update_by_id_only_touches_filled_columns(self, db):
        emp = m.Employee(name="Existing Person", department="Accounts", designation="Associate",
                          daily_target_minutes=480, work_days="0,1,2,3,4", employee_code="LOMK001")
        db.add(emp)
        db.commit()
        wb = _wb_from_rows(["Employee ID", "Phone"], [["LOMK001", "+1 555 9999"]])
        result = process_upload(db, wb)
        assert result["updated"] == 1 and result["added"] == 0
        db.refresh(emp)
        assert emp.phone == "+1 555 9999"
        assert emp.department == "Accounts"  # untouched
        assert emp.daily_target_minutes == 480  # untouched

    def test_unknown_employee_id_reported_not_silently_ignored(self, db):
        wb = _wb_from_rows(["Employee ID", "Phone"], [["LOMK999", "123"]])
        result = process_upload(db, wb)
        assert result["updated"] == 0
        assert len(result["skipped"]) == 1
        assert "not found" in result["skipped"][0]["reason"]

    def test_email_collision_with_another_employee_blocked(self, db):
        a = m.Employee(name="A", email="a@example.com", employee_code="LOMK001")
        b = m.Employee(name="B", email="b@example.com", employee_code="LOMK002")
        db.add_all([a, b])
        db.commit()
        wb = _wb_from_rows(["Employee ID", "Email"], [["LOMK001", "b@example.com"]])
        result = process_upload(db, wb)
        assert result["updated"] == 0
        assert len(result["skipped"]) == 1
        db.refresh(a)
        assert a.email == "a@example.com"

    def test_resubmitting_own_unchanged_email_is_a_harmless_noop(self, db):
        a = m.Employee(name="A", email="a@example.com", employee_code="LOMK001")
        db.add(a)
        db.commit()
        wb = _wb_from_rows(["Employee ID", "Email"], [["LOMK001", "a@example.com"]])
        result = process_upload(db, wb)
        assert result["updated"] == 1
        assert len(result["skipped"]) == 0

    def test_blank_role_on_update_does_not_reset_admin_flag(self, db):
        admin_emp = m.Employee(name="Admin Person", is_admin=True, employee_code="LOMK001")
        db.add(admin_emp)
        db.commit()
        wb = _wb_from_rows(["Employee ID", "Phone"], [["LOMK001", "999"]])
        process_upload(db, wb)
        db.refresh(admin_emp)
        assert admin_emp.is_admin is True

    def test_no_partial_commit_when_nothing_valid(self, db):
        wb = _wb_from_rows(["Full Name"], [["Bad One"]])
        result = process_upload(db, wb)
        assert result["added"] == 0 and result["updated"] == 0
        assert db.query(m.Employee).count() == 0


class TestProcessUploadDeactivate:
    def test_deactivate_sets_active_false_and_counts_separately(self, db):
        emp = m.Employee(name="Leaving Soon", department="Accounts", designation="Associate",
                          daily_target_minutes=480, work_days="0,1,2,3,4", employee_code="LOMK001")
        db.add(emp)
        db.commit()
        wb = _wb_from_rows(["Employee ID", "Action"], [["LOMK001", "Deactivate"]])
        result = process_upload(db, wb)
        assert result["updated"] == 1
        assert result["deactivated"] == 1
        db.refresh(emp)
        assert emp.active is False

    def test_deactivate_preserves_history_fields_untouched(self, db):
        emp = m.Employee(name="Leaving Soon", department="Accounts", designation="Associate",
                          daily_target_minutes=480, work_days="0,1,2,3,4", employee_code="LOMK001")
        db.add(emp)
        db.commit()
        wb = _wb_from_rows(["Employee ID", "Action"], [["LOMK001", "Deactivate"]])
        process_upload(db, wb)
        db.refresh(emp)
        assert emp.name == "Leaving Soon"
        assert emp.department == "Accounts"
        assert emp.designation == "Associate"
        assert emp.daily_target_minutes == 480

    def test_deactivate_combined_with_other_field_update(self, db):
        emp = m.Employee(name="Leaving Soon", department="Accounts", designation="Associate",
                          daily_target_minutes=480, work_days="0,1,2,3,4", employee_code="LOMK001")
        db.add(emp)
        db.commit()
        wb = _wb_from_rows(
            ["Employee ID", "Phone", "Action"],
            [["LOMK001", "+1 555 2222", "Deactivate"]],
        )
        result = process_upload(db, wb)
        assert result["deactivated"] == 1
        db.refresh(emp)
        assert emp.active is False
        assert emp.phone == "+1 555 2222"

    def test_deactivate_on_new_hire_row_is_an_error_not_a_new_deactivated_employee(self, db):
        wb = _wb_from_rows(
            ["Employee ID", "Full Name", "Department", "Designation", "Target/day", "Workdays", "Action"],
            [["", "Nobody Yet", "Accounts", "Associate", 8, "M,T,W,Th,F", "Deactivate"]],
        )
        result = process_upload(db, wb)
        assert result["added"] == 0
        assert result["deactivated"] == 0
        assert len(result["skipped"]) == 1
        assert "Employee ID is required" in result["skipped"][0]["reason"]
        assert db.query(m.Employee).count() == 0

    def test_blank_action_does_not_deactivate(self, db):
        emp = m.Employee(name="Still Here", department="Accounts", designation="Associate",
                          daily_target_minutes=480, work_days="0,1,2,3,4", employee_code="LOMK001")
        db.add(emp)
        db.commit()
        wb = _wb_from_rows(["Employee ID", "Phone"], [["LOMK001", "+1 555 3333"]])
        result = process_upload(db, wb)
        assert result["deactivated"] == 0
        db.refresh(emp)
        assert emp.active is True


class TestProcessUploadMixedSheet:
    def test_new_hire_and_update_in_same_upload(self, db):
        existing = m.Employee(name="Old Timer", department="Ops", designation="Lead",
                               daily_target_minutes=480, work_days="0,1,2,3,4",
                               employee_code="LOMK001")
        db.add(existing)
        db.commit()
        wb = _wb_from_rows(
            ["Employee ID", "Full Name", "Department", "Designation", "Target/day", "Workdays", "Phone"],
            [
                ["", "New Hire", "Accounts", "Associate", 8, "M,T,W,Th,F", ""],
                ["LOMK001", "", "", "", "", "", "+1 555 1111"],
            ],
        )
        result = process_upload(db, wb)
        assert result["added"] == 1 and result["updated"] == 1
        db.refresh(existing)
        assert existing.phone == "+1 555 1111"
        assert existing.name == "Old Timer"  # untouched
        assert db.query(m.Employee).filter_by(name="New Hire").count() == 1


class TestUtilEmployeeCodeHelpers:
    def test_first_code_is_lomk001(self, db):
        assert next_employee_code(db) == "LOMK001"

    def test_next_code_continues_from_highest(self, db):
        db.add(m.Employee(name="X", employee_code="LOMK042"))
        db.commit()
        assert next_employee_code(db) == "LOMK043"

    def test_format_is_zero_padded(self):
        assert format_employee_code(1) == "LOMK001"
        assert format_employee_code(123) == "LOMK123"
