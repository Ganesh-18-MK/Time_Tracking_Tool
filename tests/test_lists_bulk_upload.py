"""Bulk-add Project/Employer and Task Type dropdown values (Projects & Tasks
-> Bulk upload). One column per sheet, add-only — see
app/lists_bulk_upload.py's module docstring for why this is deliberately
simpler than app/bulk_upload.py / app/leave_bulk_upload.py. Same in-memory
SQLite fixture pattern as test_reports.py/test_compensation.py for the
process_upload() tests; read_upload_names()/build_sample_workbook() are
pure enough to test without a db at all.
"""
import pytest
from openpyxl import Workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app import models as m
from app.db import Base
from app.lists_bulk_upload import (
    HEADERS,
    PROJECT_DEPARTMENT_HEADER,
    TASK_PROJECT_HEADER,
    build_existing_workbook,
    build_sample_workbook,
    process_upload,
    read_project_rows,
    read_task_rows,
    read_upload_names,
)


@pytest.fixture()
def db():
    eng = create_engine("sqlite://")
    Base.metadata.create_all(eng)
    s = sessionmaker(bind=eng)()
    yield s
    s.close()


def _sheet(header, *values):
    wb = Workbook()
    ws = wb.active
    ws.append([header])
    for v in values:
        ws.append([v])
    return wb


class TestReadUploadNames:
    def test_happy_path_skips_blanks_and_trims_whitespace(self):
        wb = _sheet("Project / Employer Name", "Acme Corp.", "", None, "  Beta LLC  ")
        names, err = read_upload_names(wb, "project")
        assert err is None
        assert names == [(2, "Acme Corp."), (5, "Beta LLC")]

    def test_header_match_is_case_insensitive(self):
        wb = _sheet("project / employer name", "Gamma Inc.")
        names, err = read_upload_names(wb, "project")
        assert err is None
        assert names == [(2, "Gamma Inc.")]

    def test_wrong_column_header_is_an_error(self):
        wb = _sheet("Some Other Column", "x")
        names, err = read_upload_names(wb, "project")
        assert names == []
        assert err is not None and HEADERS["project"] in err

    def test_empty_sheet_is_an_error(self):
        wb = Workbook()
        names, err = read_upload_names(wb, "task")
        assert names == []
        assert err is not None and "empty" in err.lower()

    def test_task_sheet_uses_its_own_header(self):
        wb = _sheet("Task Name", "File review")
        names, err = read_upload_names(wb, "task")
        assert err is None
        assert names == [(2, "File review")]


def _project_sheet(*rows, with_department=True):
    """rows are (project_name, department) pairs when with_department=True
    (the two-column shape, Ganesh, 2026-08-28 — see read_project_rows()),
    or plain project names when False (the old single-column shape, for
    backward-compat cases)."""
    wb = Workbook()
    ws = wb.active
    if with_department:
        ws.append([HEADERS["project"], PROJECT_DEPARTMENT_HEADER])
        for name, dept in rows:
            ws.append([name, dept])
    else:
        ws.append([HEADERS["project"]])
        for name in rows:
            ws.append([name])
    return wb


class TestReadProjectRows:
    """Department-scoped projects (Ganesh, 2026-08-28) — the Projects
    sheet's own reader, a superset of read_upload_names() above that also
    reads an optional Department column."""

    def test_happy_path_with_department(self):
        wb = _project_sheet(("Acme Corp.", "Sales"), ("Beta LLC", "Support"))
        rows, err = read_project_rows(wb)
        assert err is None
        assert rows == [(2, "Acme Corp.", "Sales"), (3, "Beta LLC", "Support")]

    def test_same_project_multiple_department_rows(self):
        wb = _project_sheet(("Acme Corp.", "Sales"), ("Acme Corp.", "Support"))
        rows, err = read_project_rows(wb)
        assert err is None
        assert rows == [(2, "Acme Corp.", "Sales"), (3, "Acme Corp.", "Support")]

    def test_blank_department_cell_means_unrestricted(self):
        wb = _project_sheet(("Acme Corp.", ""))
        rows, err = read_project_rows(wb)
        assert err is None
        assert rows == [(2, "Acme Corp.", "")]

    def test_single_column_sheet_still_works_department_column_missing(self):
        wb = _project_sheet("Acme Corp.", "Beta LLC", with_department=False)
        rows, err = read_project_rows(wb)
        assert err is None
        assert rows == [(2, "Acme Corp.", ""), (3, "Beta LLC", "")]

    def test_blank_name_row_skipped_silently_even_with_department(self):
        wb = _project_sheet(("", "Sales"), ("Acme Corp.", "Sales"))
        rows, err = read_project_rows(wb)
        assert err is None
        assert rows == [(3, "Acme Corp.", "Sales")]

    def test_missing_project_name_column_is_a_header_error(self):
        wb = Workbook()
        wb.active.append([PROJECT_DEPARTMENT_HEADER])
        rows, err = read_project_rows(wb)
        assert rows == []
        assert err is not None and HEADERS["project"] in err


def _task_sheet(*rows):
    """rows are (task_name, project_name) pairs — the Tasks sheet's real
    two-column shape (Ganesh, 2026-08-27), see read_task_rows()."""
    wb = Workbook()
    ws = wb.active
    ws.append([HEADERS["task"], TASK_PROJECT_HEADER])
    for task_name, project_name in rows:
        ws.append([task_name, project_name])
    return wb


class TestReadTaskRows:
    """Project-scoped tasks (Ganesh, 2026-08-27) — the Tasks sheet's own
    two-column reader, separate from read_upload_names() above (which
    still only handles the Projects sheet's single column)."""

    def test_happy_path(self):
        wb = _task_sheet(("File review", "Acme Corp."), ("PWD JD", "Beta LLC"))
        rows, err = read_task_rows(wb)
        assert err is None
        assert rows == [(2, "File review", "Acme Corp."), (3, "PWD JD", "Beta LLC")]

    def test_same_task_multiple_project_rows(self):
        wb = _task_sheet(("File review", "Acme Corp."), ("File review", "Beta LLC"))
        rows, err = read_task_rows(wb)
        assert err is None
        assert rows == [(2, "File review", "Acme Corp."), (3, "File review", "Beta LLC")]

    def test_fully_blank_row_skipped_silently(self):
        wb = _task_sheet(("File review", "Acme Corp."), ("", ""))
        rows, err = read_task_rows(wb)
        assert err is None
        assert rows == [(2, "File review", "Acme Corp.")]

    def test_missing_project_name_column_is_a_header_error(self):
        wb = Workbook()
        wb.active.append([HEADERS["task"]])
        rows, err = read_task_rows(wb)
        assert rows == []
        assert err is not None and TASK_PROJECT_HEADER in err


class TestProcessUpload:
    def test_new_names_are_added(self, db):
        from sqlalchemy import select
        wb = _sheet("Project / Employer Name", "Acme Corp.", "Beta LLC")
        result = process_upload(db, wb, "project")
        assert result["header_error"] is None
        assert result["added"] == 2
        assert result["skipped"] == []
        names = {p.name for p in db.execute(select(m.Project)).scalars()}
        assert names == {"Acme Corp.", "Beta LLC"}

    def test_name_already_on_the_list_is_skipped_not_duplicated(self, db):
        db.add(m.Project(name="Acme Corp."))
        db.commit()
        wb = _sheet("Project / Employer Name", "Acme Corp.", "New Client Inc.")
        result = process_upload(db, wb, "project")
        assert result["added"] == 1
        assert len(result["skipped"]) == 1
        assert result["skipped"][0]["name"] == "Acme Corp."
        assert "already on the list" in result["skipped"][0]["reason"].lower()

    def test_duplicate_within_the_same_file_is_added_once(self, db):
        wb = _sheet("Project / Employer Name", "New Client Inc.", "New Client Inc.")
        result = process_upload(db, wb, "project")
        assert result["added"] == 1
        assert len(result["skipped"]) == 1
        assert "duplicate" in result["skipped"][0]["reason"].lower()

    def test_header_error_short_circuits_before_touching_the_db(self, db):
        wb = _sheet("Wrong Column", "x")
        result = process_upload(db, wb, "project")
        assert result["header_error"] is not None
        assert result["added"] == 0
        from sqlalchemy import select
        assert list(db.execute(select(m.Project)).scalars()) == []

    def test_no_new_rows_means_no_commit_needed_but_still_reports_correctly(self, db):
        db.add(m.Project(name="Acme Corp."))
        db.commit()
        wb = _sheet("Project / Employer Name", "Acme Corp.")
        result = process_upload(db, wb, "project")
        assert result["added"] == 0
        assert len(result["skipped"]) == 1


class TestProcessProjectUploadDepartments:
    """Department-scoped projects (Ganesh, 2026-08-28) — see
    ProjectDepartment's docstring in app/models.py and this module's own
    docstring for the one-row-per-pair convention."""

    def test_new_project_with_department_creates_both(self, db):
        from sqlalchemy import select
        wb = _project_sheet(("Acme Corp.", "Sales"))
        result = process_upload(db, wb, "project")
        assert result["header_error"] is None
        assert result["added"] == 1
        assert result["department_links_added"] == 1
        assert result["skipped"] == []
        proj = db.execute(select(m.Project)).scalar_one()
        assert proj.name == "Acme Corp."
        link = db.execute(select(m.ProjectDepartment)).scalar_one()
        assert link.project_id == proj.id and link.department == "Sales"

    def test_same_project_linked_to_two_departments_creates_project_once(self, db):
        from sqlalchemy import select
        wb = _project_sheet(("Gamma Inc.", "Sales"), ("Gamma Inc.", "Support"))
        result = process_upload(db, wb, "project")
        assert result["added"] == 1
        assert result["department_links_added"] == 2
        projects = list(db.execute(select(m.Project)).scalars())
        assert len(projects) == 1
        links = list(db.execute(select(m.ProjectDepartment)).scalars())
        assert len(links) == 2

    def test_blank_department_creates_unrestricted_project_no_link(self, db):
        from sqlalchemy import select
        wb = _project_sheet(("Acme Corp.", ""))
        result = process_upload(db, wb, "project")
        assert result["added"] == 1
        assert result["department_links_added"] == 0
        assert list(db.execute(select(m.ProjectDepartment)).scalars()) == []

    def test_department_link_added_to_already_existing_project(self, db):
        from sqlalchemy import select
        db.add(m.Project(name="Acme Corp."))
        db.commit()
        wb = _project_sheet(("Acme Corp.", "Legal"))
        result = process_upload(db, wb, "project")
        assert result["added"] == 0  # project already existed
        assert result["department_links_added"] == 1
        proj = db.execute(select(m.Project)).scalar_one()
        link = db.execute(select(m.ProjectDepartment)).scalar_one()
        assert link.project_id == proj.id and link.department == "Legal"

    def test_already_linked_department_is_skipped_not_duplicated(self, db):
        proj = m.Project(name="Acme Corp.")
        db.add(proj)
        db.commit()
        db.add(m.ProjectDepartment(project_id=proj.id, department="Sales", created_by="test"))
        db.commit()
        wb = _project_sheet(("Acme Corp.", "Sales"))
        result = process_upload(db, wb, "project")
        assert result["added"] == 0
        assert result["department_links_added"] == 0
        assert len(result["skipped"]) == 1
        assert "already linked" in result["skipped"][0]["reason"].lower()

    def test_duplicate_department_pair_within_file_is_linked_once(self, db):
        wb = _project_sheet(("Gamma Inc.", "Sales"), ("Gamma Inc.", "Sales"))
        result = process_upload(db, wb, "project")
        assert result["added"] == 1
        assert result["department_links_added"] == 1
        assert len(result["skipped"]) == 1
        assert "duplicate" in result["skipped"][0]["reason"].lower()

    def test_single_column_upload_still_works_unchanged(self, db):
        wb = _sheet("Project / Employer Name", "Acme Corp.", "Beta LLC")
        result = process_upload(db, wb, "project")
        assert result["added"] == 2
        assert result["department_links_added"] == 0
        assert result["skipped"] == []


class TestProcessTaskUpload:
    """Project-scoped tasks (Ganesh, 2026-08-27) — each Tasks-sheet row is
    now a (task, project) pair; see app/lists_bulk_upload.py's module
    docstring and ProjectTask's docstring in app/models.py."""

    def test_new_task_creates_task_and_links_it(self, db):
        db.add(m.Project(name="Acme Corp."))
        db.commit()
        wb = _task_sheet(("File review", "Acme Corp."))
        result = process_upload(db, wb, "task")
        assert result["header_error"] is None
        assert result["added"] == 1
        from sqlalchemy import select
        task = db.execute(select(m.TaskType)).scalar_one()
        assert task.name == "File review"
        link = db.execute(select(m.ProjectTask)).scalar_one()
        assert link.task_type_id == task.id

    def test_same_task_linked_to_two_projects_creates_task_once(self, db):
        db.add(m.Project(name="Acme Corp."))
        db.add(m.Project(name="Beta LLC"))
        db.commit()
        wb = _task_sheet(("File review", "Acme Corp."), ("File review", "Beta LLC"))
        result = process_upload(db, wb, "task")
        assert result["added"] == 2  # 2 links
        from sqlalchemy import select
        tasks = list(db.execute(select(m.TaskType)).scalars())
        assert len(tasks) == 1  # task created only once
        links = list(db.execute(select(m.ProjectTask)).scalars())
        assert len(links) == 2

    def test_project_not_found_is_skipped_with_reason(self, db):
        wb = _task_sheet(("File review", "Nonexistent Client"))
        result = process_upload(db, wb, "task")
        assert result["added"] == 0
        assert len(result["skipped"]) == 1
        assert "not found" in result["skipped"][0]["reason"].lower()
        from sqlalchemy import select
        assert list(db.execute(select(m.TaskType)).scalars()) == []

    def test_already_linked_pair_is_skipped_not_duplicated(self, db):
        from sqlalchemy import select
        db.add(m.Project(name="Acme Corp."))
        db.commit()
        proj = db.execute(select(m.Project)).scalar_one()
        task = m.TaskType(name="File review")
        db.add(task)
        db.commit()
        db.add(m.ProjectTask(project_id=proj.id, task_type_id=task.id, created_by="test"))
        db.commit()
        wb = _task_sheet(("File review", "Acme Corp."))
        result = process_upload(db, wb, "task")
        assert result["added"] == 0
        assert len(result["skipped"]) == 1
        assert "already linked" in result["skipped"][0]["reason"].lower()

    def test_missing_project_name_cell_is_skipped(self, db):
        db.add(m.Project(name="Acme Corp."))
        db.commit()
        wb = _task_sheet(("File review", ""))
        result = process_upload(db, wb, "task")
        assert result["added"] == 0
        assert len(result["skipped"]) == 1
        assert "missing" in result["skipped"][0]["reason"].lower()


class TestSampleAndExistingWorkbooks:
    def test_sample_workbook_header_matches_kind(self):
        for kind in ("project", "task"):
            wb = build_sample_workbook(kind)
            assert wb.active["A1"].value == HEADERS[kind]

    def test_existing_workbook_lists_only_active_names(self, db):
        db.add(m.Project(name="Active One", active=True))
        db.add(m.Project(name="Retired One", active=False))
        db.commit()
        wb = build_existing_workbook(db, "project")
        ws = wb.active
        values = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
        assert values == ["Active One"]
