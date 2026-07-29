"""Bulk leave-allocation assignment: one small sheet, one job — set an
existing employee's annual Casual/Sick/Vacation entitlement. Never creates
employees (unlike app/bulk_upload.py); every row must resolve to a known
Employee ID. Blank leave-count cells mean "leave unchanged", filled cells
overwrite — same in-memory SQLite fixture pattern as test_bulk_upload.py.
"""
import pytest
from openpyxl import Workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app import models as m
from app.db import Base
from app.leave_bulk_upload import (
    build_existing_allocations_workbook,
    build_sample_workbook,
    parse_leave_count,
    parse_row,
    process_upload,
    read_upload_rows,
)


class TestParseLeaveCount:
    def test_blank_is_none_not_a_default(self):
        assert parse_leave_count("", "Casual Leaves/Year") is None
        assert parse_leave_count(None, "Casual Leaves/Year") is None

    def test_whole_number_parses(self):
        assert parse_leave_count(12, "Casual Leaves/Year") == 12
        assert parse_leave_count("12", "Casual Leaves/Year") == 12

    def test_zero_is_valid(self):
        assert parse_leave_count(0, "Sick Leaves/Year") == 0

    def test_negative_raises(self):
        with pytest.raises(ValueError, match="negative"):
            parse_leave_count(-1, "Sick Leaves/Year")

    def test_half_day_raises(self):
        with pytest.raises(ValueError, match="whole number"):
            parse_leave_count(12.5, "Vacation Leaves/Year")

    def test_non_numeric_raises(self):
        with pytest.raises(ValueError):
            parse_leave_count("lots", "Casual Leaves/Year")


class TestParseRow:
    def test_unknown_employee_id_is_an_error(self):
        row = {"Employee ID": "LOMK999", "Casual Leaves/Year": 12}
        result = parse_row(row, {"LOMK001": 1})
        assert result["mode"] == "error"
        assert "not found" in result["error"]

    def test_blank_employee_id_is_an_error(self):
        row = {"Employee ID": "", "Casual Leaves/Year": 12}
        result = parse_row(row, {"LOMK001": 1})
        assert result["mode"] == "error"
        assert "required" in result["error"].lower()

    def test_case_insensitive_id_match(self):
        row = {"Employee ID": "lomk001", "Casual Leaves/Year": 12}
        result = parse_row(row, {"LOMK001": 1})
        assert result["error"] is None
        assert result["employee_id"] == 1

    def test_happy_path_all_three_fields(self):
        row = {
            "Employee ID": "LOMK001", "Employee Name": "Jane Doe",
            "Casual Leaves/Year": 12, "Sick Leaves/Year": 10, "Vacation Leaves/Year": 15,
        }
        result = parse_row(row, {"LOMK001": 1})
        assert result["error"] is None
        assert result["mode"] == "update"
        assert result["fields"] == {
            "casual_leave_days": 12, "sick_leave_days": 10, "vacation_leave_days": 15,
        }

    def test_partial_fields_is_a_minimal_patch(self):
        row = {"Employee ID": "LOMK001", "Sick Leaves/Year": 10}
        result = parse_row(row, {"LOMK001": 1})
        assert result["error"] is None
        assert result["fields"] == {"sick_leave_days": 10}

    def test_id_with_no_leave_numbers_is_an_error(self):
        row = {"Employee ID": "LOMK001", "Employee Name": "Jane Doe"}
        result = parse_row(row, {"LOMK001": 1})
        assert result["mode"] == "error"
        assert "nothing to apply" in result["error"].lower()

    def test_zero_counts_as_a_provided_value_not_blank(self):
        row = {"Employee ID": "LOMK001", "Casual Leaves/Year": 0}
        result = parse_row(row, {"LOMK001": 1})
        assert result["error"] is None
        assert result["fields"] == {"casual_leave_days": 0}

    def test_invalid_value_reports_error_with_other_valid_fields_still_present(self):
        row = {"Employee ID": "LOMK001", "Casual Leaves/Year": "garbage", "Sick Leaves/Year": 10}
        result = parse_row(row, {"LOMK001": 1})
        assert result["mode"] == "error"
        assert "Casual" in result["error"]


class TestSampleAndExportWorkbooks:
    def test_sample_workbook_row_parses_cleanly(self):
        wb = build_sample_workbook()
        rows, header_error = read_upload_rows(wb)
        assert header_error is None
        assert len(rows) == 1
        result = parse_row(rows[0], {"LOMK001": 1})
        assert result["error"] is None
        assert result["mode"] == "update"

    def test_empty_sheet_is_a_sheet_level_error(self):
        wb = Workbook()
        rows, error = read_upload_rows(wb)
        assert rows == [] and error is not None

    def test_blank_rows_are_skipped_silently(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["Employee ID", "Casual Leaves/Year"])
        ws.append(["LOMK001", 12])
        ws.append([None, None])
        rows, error = read_upload_rows(wb)
        assert error is None
        assert len(rows) == 1

    def test_export_round_trips_into_a_valid_update_row(self, db):
        db.add(m.Employee(
            name="Jane Doe", employee_code="LOMK001",
            casual_leave_days=12, sick_leave_days=10, vacation_leave_days=15,
        ))
        db.commit()
        wb = build_existing_allocations_workbook(db)
        rows, error = read_upload_rows(wb)
        assert error is None
        assert len(rows) == 1
        assert rows[0]["Casual Leaves/Year"] == 12
        result = parse_row(rows[0], {"LOMK001": 1})
        assert result["mode"] == "update"
        assert result["error"] is None

    def test_export_shows_zero_for_unset_allocations(self, db):
        db.add(m.Employee(name="Jane Doe", employee_code="LOMK001"))
        db.commit()
        wb = build_existing_allocations_workbook(db)
        rows, _ = read_upload_rows(wb)
        assert rows[0]["Casual Leaves/Year"] == 0
        assert rows[0]["Sick Leaves/Year"] == 0
        assert rows[0]["Vacation Leaves/Year"] == 0

    def test_deactivated_employees_excluded_from_export(self, db):
        db.add(m.Employee(name="Gone", employee_code="LOMK001", active=False))
        db.commit()
        wb = build_existing_allocations_workbook(db)
        rows, _ = read_upload_rows(wb)
        assert rows == []


@pytest.fixture()
def db():
    eng = create_engine("sqlite://")
    Base.metadata.create_all(eng)
    s = sessionmaker(bind=eng)()
    yield s
    s.close()


def _wb_from_rows(header, rows):
    wb = Workbook()
    ws = wb.active
    ws.append(header)
    for r in rows:
        ws.append(r)
    return wb


class TestProcessUpload:
    def test_valid_row_updates_employee(self, db):
        emp = m.Employee(name="Jane Doe", employee_code="LOMK001")
        db.add(emp)
        db.commit()
        wb = _wb_from_rows(
            ["Employee ID", "Casual Leaves/Year", "Sick Leaves/Year", "Vacation Leaves/Year"],
            [["LOMK001", 12, 10, 15]],
        )
        result = process_upload(db, wb)
        assert result["header_error"] is None
        assert result["updated"] == 1
        db.refresh(emp)
        assert emp.casual_leave_days == 12
        assert emp.sick_leave_days == 10
        assert emp.vacation_leave_days == 15

    def test_blank_cell_leaves_existing_value_untouched(self, db):
        emp = m.Employee(name="Jane Doe", employee_code="LOMK001",
                          casual_leave_days=12, sick_leave_days=10, vacation_leave_days=15)
        db.add(emp)
        db.commit()
        wb = _wb_from_rows(["Employee ID", "Sick Leaves/Year"], [["LOMK001", 20]])
        result = process_upload(db, wb)
        assert result["updated"] == 1
        db.refresh(emp)
        assert emp.sick_leave_days == 20
        assert emp.casual_leave_days == 12  # untouched
        assert emp.vacation_leave_days == 15  # untouched

    def test_unknown_employee_id_reported_not_silently_ignored(self, db):
        wb = _wb_from_rows(["Employee ID", "Casual Leaves/Year"], [["LOMK999", 12]])
        result = process_upload(db, wb)
        assert result["updated"] == 0
        assert len(result["skipped"]) == 1
        assert "not found" in result["skipped"][0]["reason"]

    def test_row_with_no_leave_numbers_skipped_with_reason(self, db):
        db.add(m.Employee(name="Jane Doe", employee_code="LOMK001"))
        db.commit()
        wb = _wb_from_rows(["Employee ID", "Employee Name"], [["LOMK001", "Jane Doe"]])
        result = process_upload(db, wb)
        assert result["updated"] == 0
        assert len(result["skipped"]) == 1

    def test_re_upload_overwrites_previous_values(self, db):
        emp = m.Employee(name="Jane Doe", employee_code="LOMK001", casual_leave_days=12)
        db.add(emp)
        db.commit()
        wb = _wb_from_rows(["Employee ID", "Casual Leaves/Year"], [["LOMK001", 5]])
        process_upload(db, wb)
        db.refresh(emp)
        assert emp.casual_leave_days == 5

    def test_multiple_rows_mixed_valid_and_invalid(self, db):
        db.add(m.Employee(name="Jane Doe", employee_code="LOMK001"))
        db.commit()
        wb = _wb_from_rows(
            ["Employee ID", "Casual Leaves/Year"],
            [["LOMK001", 12], ["LOMK999", 8]],
        )
        result = process_upload(db, wb)
        assert result["updated"] == 1
        assert len(result["skipped"]) == 1

    def test_no_partial_commit_when_nothing_valid(self, db):
        wb = _wb_from_rows(["Employee ID", "Casual Leaves/Year"], [["LOMK999", 12]])
        result = process_upload(db, wb)
        assert result["updated"] == 0

    def test_too_many_rows_rejected_at_sheet_level(self, db):
        rows = [["LOMK001", 12] for _ in range(501)]
        wb = _wb_from_rows(["Employee ID", "Casual Leaves/Year"], rows)
        result = process_upload(db, wb)
        assert result["header_error"] is not None
        assert "max is 500" in result["header_error"]
