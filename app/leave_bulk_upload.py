"""Bulk leave-allocation assignment via Excel (Leave -> Bulk assign leaves).

Separate, smaller sibling of app/bulk_upload.py: that one onboards/edits the
whole employee record; this one only ever touches three numbers — an
employee's annual Casual/Sick/Vacation leave entitlement (in whole days).
Display-only for now (PRD open question 6: "no quotas enforced, totals
displayed") — nothing in the app blocks a leave request or approval against
these numbers, they just show up next to each employee on the Leave page.

Every row must match an existing employee by Employee ID (the same
LOMK001-style employee_code used everywhere else) — this sheet never
creates new employees. A blank Casual/Sick/Vacation cell means "leave this
one alone"; a filled cell overwrites whatever was there before, so
re-uploading the same sheet after fixing one number is safe and expected.

Row-parsing rules are plain functions, testable without a database, the
same way app/bulk_upload.py keeps parsing separate from process_upload().
"""
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import select

from app import models as m

MAX_ROWS = 500

TEMPLATE_HEADERS = [
    "Employee ID", "Employee Name", "Casual Leaves/Year", "Sick Leaves/Year", "Vacation Leaves/Year",
]
COL_WIDTHS = [14, 24, 18, 16, 20]
COL_LETTERS = "ABCDE"


def parse_leave_count(raw, label: str) -> Optional[int]:
    """Blank -> None ("leave unchanged"). Otherwise a non-negative whole
    number of days — half-days aren't supported by this sheet."""
    if raw is None or str(raw).strip() == "":
        return None
    try:
        value = float(raw)
    except (TypeError, ValueError):
        raise ValueError(f"{label} must be a whole number of days, e.g. 12")
    if value < 0:
        raise ValueError(f"{label} can't be negative")
    if value != int(value):
        raise ValueError(f"{label} must be a whole number of days (no half-days on this sheet)")
    return int(value)


def parse_row(raw: dict, code_to_id: Dict[str, int]) -> dict:
    """Returns one of:
      {"mode": "update", "employee_id": int, "fields": {...}, "error": None}
      {"mode": "error", "fields": {}, "error": "..."}

    "fields" is a patch — only keys with a value actually filled in on the
    sheet are present; the caller must only touch those attributes.
    """
    emp_id_raw = str(raw.get("Employee ID") or "").strip()
    if not emp_id_raw:
        return {"mode": "error", "fields": {}, "error": "Employee ID is required — this sheet never creates new employees"}
    target_id = code_to_id.get(emp_id_raw.upper())
    if target_id is None:
        return {"mode": "error", "fields": {}, "error": f"Employee ID '{emp_id_raw}' not found"}

    errors: List[str] = []
    fields: dict = {}

    try:
        casual = parse_leave_count(raw.get("Casual Leaves/Year"), "Casual Leaves/Year")
        if casual is not None:
            fields["casual_leave_days"] = casual
    except ValueError as e:
        errors.append(str(e))

    try:
        sick = parse_leave_count(raw.get("Sick Leaves/Year"), "Sick Leaves/Year")
        if sick is not None:
            fields["sick_leave_days"] = sick
    except ValueError as e:
        errors.append(str(e))

    try:
        vacation = parse_leave_count(raw.get("Vacation Leaves/Year"), "Vacation Leaves/Year")
        if vacation is not None:
            fields["vacation_leave_days"] = vacation
    except ValueError as e:
        errors.append(str(e))

    if errors:
        return {"mode": "error", "fields": {}, "error": "; ".join(errors)}
    if not fields:
        return {"mode": "error", "fields": {}, "error": "Row has an Employee ID but no leave numbers filled in — nothing to apply"}
    return {"mode": "update", "employee_id": target_id, "fields": fields, "error": None}


def read_upload_rows(wb: Workbook) -> Tuple[List[dict], Optional[str]]:
    """Returns (rows, error). error is set (rows empty) only when the sheet
    itself is unusable (no header row at all)."""
    ws = wb.active
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    # See app/bulk_upload.py's read_upload_rows for why the all-blank check
    # is needed on top of the None check (a brand-new openpyxl Workbook()
    # isn't truly empty).
    if header_row is None or all(c is None or str(c).strip() == "" for c in header_row):
        return [], "The sheet is empty — no header row found."
    header_map = {}
    for idx, cell in enumerate(header_row):
        if cell is None:
            continue
        header_map[str(cell).strip().lower()] = idx
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue  # blank row — skip silently, not an error
        rows.append({
            field: (row[header_map[field.lower()]]
                     if field.lower() in header_map and header_map[field.lower()] < len(row)
                     else None)
            for field in TEMPLATE_HEADERS
        })
    return rows, None


def process_upload(db, wb: Workbook) -> dict:
    """Parses + applies an uploaded workbook. Valid rows are applied and
    committed in one transaction; invalid or unresolvable rows are skipped
    and listed with a reason, never silently dropped. Returns
    {"updated": int, "skipped": [{"row": int, "name": str, "reason": str}],
    "header_error": str | None}."""
    rows, header_error = read_upload_rows(wb)
    if header_error:
        return {"updated": 0, "skipped": [], "header_error": header_error}
    if len(rows) > MAX_ROWS:
        return {
            "updated": 0, "skipped": [],
            "header_error": f"Sheet has {len(rows)} data rows — max is {MAX_ROWS} per upload. Split it into batches.",
        }

    existing = list(db.execute(select(m.Employee.id, m.Employee.employee_code)).all())
    code_to_id = {code.upper(): eid for (eid, code) in existing if code}

    to_update: List[Tuple[int, dict]] = []
    skipped = []
    for i, raw in enumerate(rows, start=2):  # row 1 is the header
        display = (str(raw.get("Employee Name") or raw.get("Employee ID") or "").strip()) or "(blank)"
        result = parse_row(raw, code_to_id)
        if result["error"]:
            skipped.append({"row": i, "name": display, "reason": result["error"]})
            continue
        to_update.append((result["employee_id"], result["fields"]))

    updated = 0
    if to_update:
        ids = [eid for eid, _ in to_update]
        emp_by_id = {e.id: e for e in db.execute(select(m.Employee).where(m.Employee.id.in_(ids))).scalars()}
        for emp_id, fields in to_update:
            emp = emp_by_id.get(emp_id)
            if emp is None:
                continue
            for key, value in fields.items():
                setattr(emp, key, value)
            updated += 1

    if updated:
        db.commit()
    return {"updated": updated, "skipped": skipped, "header_error": None}


def build_sample_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Leave allocations"
    ws.append(TEMPLATE_HEADERS)
    for c in ws[1]:
        c.font = Font(bold=True)
    ws.append(["LOMK001", "Jane Doe", 12, 10, 15])
    for col, width in zip(COL_LETTERS, COL_WIDTHS):
        ws.column_dimensions[col].width = width

    info = wb.create_sheet("Instructions")
    info.append(["Column", "Required?", "Format / allowed values"])
    for c in info[1]:
        c.font = Font(bold=True)
    for row in [
        ("Employee ID", "Yes", "Must match an existing employee's ID (LOMK001, ...) — this sheet never creates new employees"),
        ("Employee Name", "No", "For your own reference only — not used to match the row, Employee ID is"),
        ("Casual Leaves/Year", "No", "Whole number of days, e.g. 12. Blank = leave unchanged"),
        ("Sick Leaves/Year", "No", "Whole number of days. Blank = leave unchanged"),
        ("Vacation Leaves/Year", "No", "Whole number of days. Blank = leave unchanged"),
        ("", "", ""),
        ("These are annual entitlements for display next to each employee", "", ""),
        ("on the Leave page — nothing here blocks approving leave past them.", "", ""),
        ("Use \"Download current allocations\" to see what's already set", "", ""),
        ("before editing just the numbers you want to change.", "", ""),
    ]:
        info.append(row)
    info.column_dimensions["A"].width = 46
    info.column_dimensions["B"].width = 12
    info.column_dimensions["C"].width = 70
    return wb


def build_existing_allocations_workbook(db) -> Workbook:
    """Every active employee, one row each, pre-filled with their current
    entitlement (0 where unset) so re-uploading unchanged is a safe no-op."""
    emps = list(
        db.execute(
            select(m.Employee)
            .where(m.Employee.active.is_(True))
            .order_by(m.Employee.department, m.Employee.name)
        ).scalars()
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "Leave allocations"
    ws.append(TEMPLATE_HEADERS)
    for c in ws[1]:
        c.font = Font(bold=True)
    for e in emps:
        ws.append([
            e.employee_code or "", e.name,
            e.casual_leave_days or 0, e.sick_leave_days or 0, e.vacation_leave_days or 0,
        ])
    for col, width in zip(COL_LETTERS, COL_WIDTHS):
        ws.column_dimensions[col].width = width
    return wb
