"""Bulk holiday upload via Excel (Settings & Configurations -> Holiday
Management -> Bulk upload holidays).

Separate, smaller sibling of app/leave_bulk_upload.py, same shape: plain
row-parsing functions testable without a database, and a process_upload()
that applies everything in one transaction. Unlike the leave sheet (which
only ever patches an existing employee), this one creates or updates
Holiday rows directly — there's no employee to match against, just a date.

Holidays are one shared company-wide list (Ganesh, 2026-08-14 — reverted
the brief 2026-08-12 per-country split), so the sheet only needs two
columns: Holiday Name and Holiday Date. Every row this module writes gets
DEFAULT_LOCATION stamped on internally (see Holiday's docstring in
app/models.py for why the location column still exists at all) — that
value is never surfaced to the person filling in the sheet and never read
back out anywhere.
"""
import datetime as dt
from typing import List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import select

from app import models as m
from app.bulk_upload import parse_cell_date

MAX_ROWS = 1000

TEMPLATE_HEADERS = ["Holiday Name", "Holiday Date"]
COL_WIDTHS = [30, 16]
COL_LETTERS = "AB"


def parse_row(raw: dict) -> dict:
    """Returns one of:
      {"mode": "ok", "date": dt.date, "name": str, "error": None}
      {"mode": "error", "error": "..."}
    """
    name = str(raw.get("Holiday Name") or "").strip()

    try:
        date = parse_cell_date(raw.get("Holiday Date"))
    except ValueError as e:
        return {"mode": "error", "error": str(e)}
    if date is None:
        return {"mode": "error", "error": "Holiday Date is required"}

    return {"mode": "ok", "date": date, "name": name, "error": None}


def read_upload_rows(wb: Workbook) -> Tuple[List[dict], Optional[str]]:
    """Returns (rows, error). error is set (rows empty) only when the sheet
    itself is unusable (no header row at all) — same shape as
    leave_bulk_upload.read_upload_rows."""
    ws = wb.active
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if header_row is None or all(c is None or str(c).strip() == "" for c in header_row):
        return [], "The sheet is empty — no header row found."
    header_map = {}
    for idx, cell in enumerate(header_row):
        if cell is None:
            continue
        header_map[str(cell).strip().lower()] = idx
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue  # blank row — skip silently, not an error
        rows.append({
            field: (row[header_map[field.lower()]]
                     if field.lower() in header_map and header_map[field.lower()] < len(row)
                     else None)
            for field in TEMPLATE_HEADERS
        })
    return rows, None


def process_upload(db, wb: Workbook) -> dict:
    """Parses + applies an uploaded workbook. Valid rows are applied and
    committed in one transaction; invalid rows are skipped and listed with
    a reason, never silently dropped. A row matching an existing Holiday
    Date updates that row's name; otherwise a new Holiday is created (with
    DEFAULT_LOCATION stamped on internally — see module docstring). Returns
    {"added": int, "updated": int,
    "skipped": [{"row": int, "name": str, "reason": str}],
    "header_error": str | None}."""
    rows, header_error = read_upload_rows(wb)
    if header_error:
        return {"added": 0, "updated": 0, "skipped": [], "header_error": header_error}
    if len(rows) > MAX_ROWS:
        return {
            "added": 0, "updated": 0, "skipped": [],
            "header_error": f"Sheet has {len(rows)} data rows — max is {MAX_ROWS} per upload. Split it into batches.",
        }

    # Keyed by date alone now — one holiday per calendar date, company-wide.
    # A date with more than one existing row (leftover from the brief
    # per-country era) is matched to whichever comes back first; the rest
    # are harmless duplicates from engine.holidays_set()'s point of view
    # (it just wants the date in the set) and aren't touched by this upload.
    existing = {}
    for h in db.execute(select(m.Holiday)).scalars():
        existing.setdefault(h.date, h)

    added = updated = 0
    skipped = []
    for i, raw in enumerate(rows, start=2):  # row 1 is the header
        display = str(raw.get("Holiday Name") or raw.get("Holiday Date") or "").strip() or "(blank)"
        result = parse_row(raw)
        if result["error"]:
            skipped.append({"row": i, "name": display, "reason": result["error"]})
            continue
        row = existing.get(result["date"])
        if row is None:
            row = m.Holiday(date=result["date"], location=m.DEFAULT_LOCATION, name=result["name"])
            db.add(row)
            existing[result["date"]] = row
            added += 1
        else:
            row.name = result["name"]
            updated += 1

    if added or updated:
        db.commit()
    return {"added": added, "updated": updated, "skipped": skipped, "header_error": None}


def build_sample_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Holidays"
    ws.append(TEMPLATE_HEADERS)
    for c in ws[1]:
        c.font = Font(bold=True)
    ws.append(["Independence Day", dt.date(2026, 7, 4)])
    ws.append(["Diwali", dt.date(2026, 11, 8)])
    for col, width in zip(COL_LETTERS, COL_WIDTHS):
        ws.column_dimensions[col].width = width

    info = wb.create_sheet("Instructions")
    info.append(["Column", "Required?", "Format / allowed values"])
    for c in info[1]:
        c.font = Font(bold=True)
    for row in [
        ("Holiday Name", "No", "Free text, e.g. 'Independence Day' — shown to employees"),
        ("Holiday Date", "Yes", "YYYY-MM-DD, or an Excel date cell"),
        ("", "", ""),
        ("Holidays are one shared list for everyone — no country column.", "", ""),
        ("A row whose Date already exists updates that holiday's name", "", ""),
        ("instead of creating a duplicate — safe to re-upload after", "", ""),
        ("fixing a typo.", "", ""),
    ]:
        info.append(row)
    info.column_dimensions["A"].width = 55
    info.column_dimensions["B"].width = 12
    info.column_dimensions["C"].width = 50
    return wb


def build_existing_holidays_workbook(db) -> Workbook:
    """Every holiday, one row each — download-before-edit companion to the
    sample template, same idea as
    leave_bulk_upload.build_existing_allocations_workbook."""
    holidays = list(db.execute(select(m.Holiday).order_by(m.Holiday.date)).scalars())
    wb = Workbook()
    ws = wb.active
    ws.title = "Holidays"
    ws.append(TEMPLATE_HEADERS)
    for c in ws[1]:
        c.font = Font(bold=True)
    for h in holidays:
        ws.append([h.name, h.date])
    for col, width in zip(COL_LETTERS, COL_WIDTHS):
        ws.column_dimensions[col].width = width
    return wb
