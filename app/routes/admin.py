"""Admin screens (PRD §7): compliance dashboard, person detail, roster,
lists, leave + compensation, config, audit."""
import datetime as dt
import io
import json
from typing import Dict, Optional

from fastapi import APIRouter, Depends, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import RedirectResponse, StreamingResponse
from openpyxl import load_workbook
from sqlalchemy import delete, func, select
from sqlalchemy.orm import Session

from app import bulk_upload, compensation, engine, holiday_bulk_upload, leave_bulk_upload, lists_bulk_upload, models as m, reports
from app.auth import Forbidden, admin_department_scope, led_by, require_admin, require_super_admin
from app.db import get_db
# TK-04 (Ganesh, 2026-08-28) — _client_required_error() is the one rule
# that decides whether a Case Type project needs its Client field filled
# in; imported rather than duplicated so the admin-side Assign-a-task form
# (admin_add_plan/admin_edit_plan below) can't quietly drift from the
# employee-side rule add_plan()/start_task_timer()/add_entry() already
# enforce. No circular import risk — app/routes/employee.py never imports
# from this module.
from app.routes.employee import _client_required_error
from app.validation import task_allowed_for_project
from app.templating import (
    HOLIDAY_MANAGEMENT_ENABLED,
    LEAVE_MANAGEMENT_V2_ENABLED,
    TICKETING_ENABLED,
    flash,
    render,
)
from app.util import (
    FormError,
    ROLE_EMPLOYEE,
    audit,
    capitalize_first,
    fmt_hm,
    next_employee_code,
    overtime_row_flags,
    parse_date_field,
    parse_hours_field,
    parse_int_field,
    parse_ym,
    prev_next_month,
    role_to_flags,
    today_local,
)

router = APIRouter(prefix="/admin")


# --------------------------------------------------------------------------
# Compliance dashboard — the monthly sheet, rebuilt live (the core payoff)
# --------------------------------------------------------------------------
@router.get("")
def dashboard(
    request: Request,
    ym: Optional[str] = None,
    dept: Optional[str] = None,
    exceptions: int = 0,
    admin: m.Employee = Depends(require_admin),
    db: Session = Depends(get_db),
):
    cfg = engine.get_config(db)
    year, month = parse_ym(ym)
    first, last = engine.month_range(year, month)
    today = today_local()

    # live months stay fresh on load (cheap at this scale; nightly job optional)
    if first <= today:
        engine.recompute_all(db, first, min(last, today))

    # Department-scoped admin (team lead): None for a super admin (no
    # restriction). Filtering all_emps here — before any downstream
    # list/count is built from it — scopes the whole page in one place;
    # nothing below needs to know the difference between the two tiers.
    scope = admin_department_scope(admin)

    all_emps = list(
        db.execute(
            select(m.Employee)
            .where(m.Employee.active.is_(True), m.Employee.tracked.is_(True))
            .order_by(m.Employee.department, m.Employee.name)
        ).scalars()
    )
    if scope is not None:
        all_emps = [e for e in all_emps if (e.department or "—") == scope]
        if dept is not None:
            dept = scope  # ignore/override ?dept= tampering — always their own
    all_depts = sorted({e.department or "—" for e in all_emps})
    dept_counts = {}
    for e in all_emps:
        key = e.department or "—"
        dept_counts[key] = dept_counts.get(key, 0) + 1
    total_emps = len(all_emps)
    emps = [e for e in all_emps if (e.department or "—") == dept] if dept else all_emps

    # The detail grid (every employee's day-by-day cells) only renders once
    # a department has actually been picked — dept is None on first load
    # ("/admin" with no query string) vs "" once the "All departments" card
    # is explicitly clicked ("/admin?dept="). Checked here, before dept is
    # coerced to "" below for display/filtering, since that coercion would
    # otherwise erase the distinction the template needs for this branch.
    show_grid = dept is not None

    # live today-snapshot for the landing KPI cards — independent of which
    # month's grid is being browsed below (see engine.today_attendance)
    attendance = engine.today_attendance(db, cfg, today)
    # same snapshot, broken down per department — feeds the "N present
    # today" number on each department card (dept_counts stays the total
    # headcount, shown alongside it).
    dept_present = {}
    for e in attendance["logged"]:
        key = e.department or "—"
        dept_present[key] = dept_present.get(key, 0) + 1

    by_emp = engine.statuses_for_month(db, year, month)
    comp_erases = cfg.get("comp_erases_strike") == "1"
    threshold = engine.cfg_int(cfg, "strike_threshold")
    days = [first + dt.timedelta(days=i) for i in range((last - first).days + 1)]

    # "Needs attention" + "Recent activity" — only rendered on the landing
    # view (not show_grid), so skip the extra queries when they won't be
    # used. Violations are computed org-wide from all_emps (not the
    # dept-filtered `emps`) since this is meant to surface everything that
    # needs a look, regardless of which department card was last clicked.
    pending_leave_rows, open_support_rows, violations, recent_audit, unlock_requests = [], [], [], [], []
    if not show_grid:
        # Leave Management is now Super-Admin-only (Ganesh, 2026-08-28 —
        # narrowed the department-scoped Team Lead's access to 5 specific
        # screens, Leave not among them), so this landing-page preview is
        # skipped entirely for a department-scoped admin the same way
        # Support Inbox/Audit Log/Unlock requests already were below —
        # showing "leave requests need attention" to someone with no
        # route left to act on them would be a dead end, not a signal.
        if scope is None:
            pending_leave_rows = list(
                db.execute(
                    select(m.LeaveRecord)
                    .where(m.LeaveRecord.status == m.LEAVE_REQUESTED)
                    .order_by(m.LeaveRecord.created_at)
                    .limit(5)
                ).scalars()
            )
            open_support_rows = list(
                db.execute(
                    select(m.SupportQuery)
                    .where(m.SupportQuery.status == m.SUPPORT_OPEN)
                    .order_by(m.SupportQuery.created_at)
                    .limit(5)
                ).scalars()
            )
            recent_audit = list(
                db.execute(select(m.AuditLog).order_by(m.AuditLog.at.desc()).limit(8)).scalars()
            )
            # Unlock requests (Ganesh, 2026-08-27) — super-admin-only
            # preview, same reasoning as Support Inbox/Audit Log just
            # above: only a super admin can actually act on one (see
            # unlock_day/reject_unlock_request's require_super_admin), so
            # there's no meaningful department-scoped version to show a
            # Team Lead instead.
            unlock_requests = list(
                db.execute(
                    select(m.UnlockRequest)
                    .where(m.UnlockRequest.status == m.LEAVE_REQUESTED)
                    .order_by(m.UnlockRequest.created_at)
                    .limit(5)
                ).scalars()
            )
        # Ganesh, 2026-08-01: also surface anyone sitting on an open (not
        # yet compensation-linked) shortfall day, not just employees who've
        # already crossed the strike threshold — so an admin can catch and
        # fix a single Partial/Missing day via Person Detail's
        # Compensation links before it ever becomes a violation.
        # strikes_in() already only counts effective_status() days still in
        # STRIKE_STATUSES (a compensated day reads Complete and drops out),
        # so e_strikes doubles as "how many open shortfall days" — no
        # separate query needed. is_violation flags the third tuple element
        # so the template can badge threshold-crossers differently from a
        # lone still-fixable shortfall.
        for e in all_emps:
            e_strikes = engine.strikes_in(by_emp.get(e.id, {}).values(), comp_erases)
            if e_strikes > 0:
                violations.append((e, e_strikes, e_strikes >= threshold))
        violations.sort(key=lambda row: (not row[2], -row[1]))
        violations = violations[:8]

    groups = {}
    for e in emps:
        rows = by_emp.get(e.id, {})
        strikes = engine.strikes_in(rows.values(), comp_erases)
        rec = {
            "emp": e,
            "cells": [rows.get(d) for d in days],
            "strikes": strikes,
            "violation": strikes >= threshold,
        }
        if exceptions and strikes == 0:
            continue
        groups.setdefault(e.department or "—", []).append(rec)

    (py, pm), (ny, nm) = prev_next_month(year, month)
    return render(
        request,
        "admin/dashboard.html",
        {
            "user": admin,
            "year": year,
            "month": month,
            "days": days,
            "groups": groups,
            "all_depts": all_depts,
            "dept_counts": dept_counts,
            "dept_present": dept_present,
            "total_emps": total_emps,
            "show_grid": show_grid,
            "attendance": attendance,
            "pending_leave_rows": pending_leave_rows,
            "open_support_rows": open_support_rows,
            "violations": violations,
            "recent_audit": recent_audit,
            "unlock_requests": unlock_requests,
            "dept": dept or "",
            "exceptions": exceptions,
            "threshold": threshold,
            "comp_erases": comp_erases,
            "today": today,
            "prev_ym": f"{py}-{pm:02d}",
            "next_ym": f"{ny}-{nm:02d}",
            "ym": f"{year}-{month:02d}",
        },
        db=db,
    )


@router.post("/recompute")
def recompute(
    request: Request,
    ym: str = Form(...),
    admin: m.Employee = Depends(require_admin),
    db: Session = Depends(get_db),
):
    year, month = parse_ym(ym)
    first, last = engine.month_range(year, month)
    n = engine.recompute_all(db, first, min(last, today_local()))
    audit(db, admin.name, "recompute_month", "DayStatus", ym, {"rows": n})
    flash(request, f"Recomputed {n} day-statuses for {ym}.", "ok")
    return RedirectResponse(f"/admin?ym={ym}", status_code=303)


def _shortfalls_surpluses(statuses, comp_erases: bool):
    """Which DayStatus rows in a month are shortfall days (candidates to
    link FROM) vs. surplus days (candidates to link AS make-up) for a
    Compensation Link. Pulled out of person() below (Ganesh, 2026-08-21) so
    Overtime Management's own quick-link picker can compute the same thing
    for whichever employee is selected there, without duplicating the
    variance/status logic a second time — see person() and overtime_page()."""
    shortfalls = [
        r for r in statuses
        if (r.variance_minutes or 0) < 0 and r.effective_status(comp_erases) in m.STRIKE_STATUSES
    ]
    surpluses = [r for r in statuses if (r.variance_minutes or 0) > 0]
    return shortfalls, surpluses




# --------------------------------------------------------------------------
# Person detail: full log, ledger, leave history, overrides, comp links
# --------------------------------------------------------------------------
@router.get("/person/{emp_id}")
def person(
    emp_id: int,
    request: Request,
    ym: Optional[str] = None,
    admin: m.Employee = Depends(require_admin),
    db: Session = Depends(get_db),
):
    emp = db.get(m.Employee, emp_id)
    if emp is None:
        return RedirectResponse("/admin/roster", status_code=303)
    scope = admin_department_scope(admin)
    if scope is not None and (emp.department or "—") != scope:
        # Department-scoped admin trying to view someone outside their own
        # team (e.g. by guessing/editing the URL) — same "you don't have
        # access" redirect as any other Forbidden.
        raise Forbidden()
    cfg = engine.get_config(db)
    year, month = parse_ym(ym)
    first, last = engine.month_range(year, month)
    today = today_local()
    if first <= today:
        engine.recompute_employee(db, emp, first, min(last, today), cfg)

    statuses = list(
        db.execute(
            select(m.DayStatus)
            .where(m.DayStatus.employee_id == emp.id, m.DayStatus.date.between(first, last))
            .order_by(m.DayStatus.date)
        ).scalars()
    )
    # Unlock requests (Ganesh, 2026-08-27) — surfaced right next to the
    # existing per-day unlock control below, so an admin has the
    # employee's own note in view before typing their unlock reason.
    pending_unlocks_by_date = {
        r.date: r
        for r in db.execute(
            select(m.UnlockRequest).where(
                m.UnlockRequest.employee_id == emp.id,
                m.UnlockRequest.date.between(first, last),
                m.UnlockRequest.status == m.LEAVE_REQUESTED,
            )
        ).scalars()
    }
    ledger = engine.running_ledger(db, emp, first, min(last, today))
    subs = {
        s.date: s
        for s in db.execute(
            select(m.DaySubmission).where(
                m.DaySubmission.employee_id == emp.id,
                m.DaySubmission.date.between(first, last),
            )
        ).scalars()
    }
    entries = list(
        db.execute(
            select(m.TaskEntry)
            .where(m.TaskEntry.employee_id == emp.id, m.TaskEntry.date.between(first, last))
            .order_by(m.TaskEntry.date.desc(), m.TaskEntry.start_minute)
        ).scalars()
    )
    by_day = {}
    for e in entries:
        by_day.setdefault(e.date, []).append(e)
    # Overtime-colored task log rows (Ganesh, 2026-08-21) — same treatment as
    # Today's live view (see employee.py's _day_context), just driven by each
    # day's already-computed DayStatus.target_minutes instead of a fresh
    # leave/break recompute, since this route already ran recompute_employee
    # above and target_minutes reflects the identical leave/break-adjusted
    # target. Per-day entries here are still ascending by start_minute (the
    # `entries` query orders by date desc, start_minute asc, and grouping
    # preserves that order within each date), so util.overtime_row_flags()'s
    # cumulative-sum assumption holds the same way it does on Today.
    status_by_day = {s.date: s for s in statuses}
    for day, items in by_day.items():
        day_target = status_by_day[day].target_minutes if day in status_by_day else 0
        for e, is_ot in zip(items, overtime_row_flags([x.duration_minutes for x in items], day_target)):
            e.is_overtime = is_ot
    leaves = list(
        db.execute(
            select(m.LeaveRecord)
            .where(m.LeaveRecord.employee_id == emp.id)
            .order_by(m.LeaveRecord.start_date.desc())
        ).scalars()
    )
    links = list(
        db.execute(
            select(m.CompensationLink)
            .where(m.CompensationLink.employee_id == emp.id)
            .order_by(m.CompensationLink.shortfall_date.desc())
        ).scalars()
    )
    breaks = list(
        db.execute(
            select(m.BreakEntry)
            .where(m.BreakEntry.employee_id == emp.id, m.BreakEntry.date.between(first, last))
            .order_by(m.BreakEntry.date.desc(), m.BreakEntry.start_minute)
        ).scalars()
    )
    comp_erases = cfg.get("comp_erases_strike") == "1"
    strikes = engine.strikes_in(statuses, comp_erases)
    shortfalls, surpluses = _shortfalls_surpluses(statuses, comp_erases)
    shortfall_allocated_by_date = engine.shortfall_allocated_minutes_by_date(db, emp.id)
    surplus_used_by_date = engine.surplus_minutes_used_by_date(db, emp.id)
    comp = compensation.monthly_summary(db, emp, year, month, today)
    (py, pm), (ny, nm) = prev_next_month(year, month)
    # TK-04 (Ganesh, 2026-08-28) — "Assigned tasks" card: the Assign form's
    # Project/Task combo data, plus every still-Planned row for this
    # employee (any date, not scoped to the month being viewed — an
    # assignment could be for a future date) so the admin can edit/remove
    # one without hunting through the Task log below.
    plan_project_items, plan_task_items = _plan_combo_items(db)
    assigned_plans = list(
        db.execute(
            select(m.PlannedTask)
            .where(m.PlannedTask.employee_id == emp.id, m.PlannedTask.status == m.PLAN_PLANNED)
            .order_by(m.PlannedTask.date)
        ).scalars()
    )
    return render(
        request,
        "admin/person.html",
        {
            "user": admin,
            "emp": emp,
            "plan_project_items": plan_project_items,
            "plan_task_items": plan_task_items,
            "assigned_plans": assigned_plans,
            "statuses": statuses,
            "pending_unlocks_by_date": pending_unlocks_by_date,
            "ledger": ledger,
            "balance": ledger[-1]["balance"] if ledger else 0,
            "subs": subs,
            "by_day": sorted(by_day.items(), reverse=True),
            "leaves": leaves,
            "breaks": breaks,
            "max_break_minutes": engine.cfg_int(cfg, "max_break_minutes"),
            "links": [
                (lk, [dt.date.fromisoformat(x) for x in json.loads(lk.surplus_dates or "[]")])
                for lk in links
            ],
            "strikes": strikes,
            "threshold": engine.cfg_int(cfg, "strike_threshold"),
            "comp_erases": comp_erases,
            "shortfalls": shortfalls,
            "shortfall_allocated_by_date": shortfall_allocated_by_date,
            "surplus_used_by_date": surplus_used_by_date,
            "surpluses": surpluses,
            "comp": comp,
            "year": year,
            "month": month,
            "ym": f"{year}-{month:02d}",
            "prev_ym": f"{py}-{pm:02d}",
            "next_ym": f"{ny}-{nm:02d}",
            "today": today,
        },
        db=db,
    )


@router.post("/person/{emp_id}/unlock")
def unlock_day(
    emp_id: int,
    request: Request,
    date: str = Form(...),
    reason: str = Form(...),
    ym: str = Form(""),
    admin: m.Employee = Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    try:
        day = parse_date_field(date)
    except FormError as e:
        flash(request, e.message, "err")
        return RedirectResponse(f"/admin/person/{emp_id}?ym={ym}", status_code=303)
    sub = db.execute(
        select(m.DaySubmission).where(
            m.DaySubmission.employee_id == emp_id, m.DaySubmission.date == day
        )
    ).scalar_one_or_none()
    if sub is None:
        flash(request, "No submission found for that day.", "err")
    else:
        sub.locked = False
        sub.unlock_count += 1
        # Unlock requests (Ganesh, 2026-08-27) — auto-resolve any pending
        # request for this exact employee+date the moment the day actually
        # gets unlocked, whether the admin arrived here via the request
        # queue or just unlocked directly without noticing a request
        # existed. One decision (locked or not), not two things that can
        # disagree — see UnlockRequest's docstring in app/models.py.
        req = db.execute(
            select(m.UnlockRequest).where(
                m.UnlockRequest.employee_id == emp_id, m.UnlockRequest.date == day,
                m.UnlockRequest.status == m.LEAVE_REQUESTED,
            )
        ).scalar_one_or_none()
        if req is not None:
            req.status = m.LEAVE_APPROVED
            req.reviewed_by = admin.name
            req.reviewed_at = dt.datetime.utcnow()
            req.review_note = reason
        db.commit()
        # every unlock is logged: who, when, what (PRD §4)
        audit(
            db, admin.name, "unlock_day", "DaySubmission", f"{emp_id}:{date}",
            {"reason": reason, "unlock_count": sub.unlock_count},
        )
        flash(request, f"Unlocked {date} — the employee can now edit and resubmit.", "ok")
    return RedirectResponse(f"/admin/person/{emp_id}?ym={ym}", status_code=303)


def _plan_combo_items(db: Session):
    """Plain {id, name, ...} dicts for the "Assign a task" combo on Person
    Detail (TK-04, Ganesh, 2026-08-28) — same shape combo.js's
    initProjectTaskCombo already expects everywhere else it's used (Today
    page, Lists page). No "assigned to you" star-sorting the way
    employee.py's own _combo_items() does — that's an employee-specific
    convenience about picking your OWN work, not relevant to an admin
    picking on someone else's behalf. Task items still carry project_ids
    so the widget filters the Task list down to whatever's valid for the
    selected Project, same as everywhere else (see ProjectTask's docstring
    in app/models.py) — task_allowed_for_project() is still the real
    server-side gate in admin_add_plan/admin_edit_plan below, this is only
    the UI convenience."""
    projects = list(
        db.execute(
            select(m.Project).where(m.Project.active.is_(True), m.Project.status == m.LIST_APPROVED).order_by(m.Project.name)
        ).scalars()
    )
    tasks = list(
        db.execute(
            select(m.TaskType).where(m.TaskType.active.is_(True), m.TaskType.status == m.LIST_APPROVED).order_by(m.TaskType.name)
        ).scalars()
    )
    task_project_links: Dict[int, list] = {}
    for pid, tid in db.execute(select(m.ProjectTask.project_id, m.ProjectTask.task_type_id)).all():
        task_project_links.setdefault(tid, []).append(pid)
    project_items = [{"id": p.id, "name": p.name, "is_case_type": bool(p.is_case_type)} for p in projects]
    task_items = [{"id": t.id, "name": t.name, "project_ids": task_project_links.get(t.id)} for t in tasks]
    return project_items, task_items


def _plan_redirect(emp_id: int, ym: str, return_to: str) -> str:
    """Where admin_add_plan/admin_edit_plan/admin_delete_plan send the admin
    back to. Defaults to Person Detail (the original TK-04 flow, 2026-08-28)
    — Assign Work's own "Assign a new task" form (Ganesh, 2026-08-29: "i want
    this assign tasks feature to be in Assignments") sets return_to=assignments
    so assigning from there doesn't bounce the admin to Person Detail, and
    keeps the same employee selected so the result shows immediately. Same
    precedent as _complink_redirect() above for Overtime Management."""
    if return_to == "assignments":
        return f"/admin/assignments?employee_id={emp_id}"
    return f"/admin/person/{emp_id}?ym={ym}"


@router.post("/person/{emp_id}/plan/add")
def admin_add_plan(
    emp_id: int,
    request: Request,
    project_id: int = Form(...),
    task_type_id: int = Form(...),
    date: str = Form(...),
    details: str = Form(""),
    client: str = Form(""),
    ym: str = Form(""),
    return_to: str = Form(""),
    admin: m.Employee = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """TK-04 — "Admin creates a project/task in an employee log." Admin/
    team lead picks a Project, Task, date and an optional note; it's added
    as a PLAN_PLANNED row that sits in the employee's log until they Start
    it — same Start/Pause/Resume/Stop controls their own self-planned rows
    already use, no employee-side route changes needed (see PlannedTask's
    docstring). Gated the same require_admin/admin_department_scope tier
    Person Detail itself already uses — "assign Projects/Tasks to team
    members" is already one of the 5 capabilities a department admin
    keeps (see the department-scoped-admin CLAUDE.md bullet), and this is
    that same capability applied to a specific unit of work rather than a
    standing ProjectAssignment/TaskAssignment row.

    Unlike add_plan()'s own note requirement ("Say what you plan to do"),
    the note here is genuinely optional per TK-04's acceptance criteria —
    left blank is fine.

    task_allowed_for_project() (data integrity — which tasks exist under
    which project) IS enforced, same as add_plan(). Department scoping
    (validation.project_allowed_for_department) is deliberately NOT
    enforced here — same acting_admin precedent validate_entry() already
    uses elsewhere in this app: an admin deliberately assigning work on
    someone's behalf isn't the employee whose department that scoping
    exists to restrict."""
    led = admin_department_scope(admin)
    emp = db.get(m.Employee, emp_id)
    if emp is None or (led is not None and (emp.department or "—") != led):
        raise Forbidden()
    dest = _plan_redirect(emp_id, ym, return_to)
    try:
        d = parse_date_field(date)
    except FormError as e:
        flash(request, e.message, "err")
        return RedirectResponse(dest, status_code=303)

    project = db.get(m.Project, project_id)
    task = db.get(m.TaskType, task_type_id)
    if project is None or not project.active or task is None or not task.active:
        flash(request, "Choose a Project and Task.", "err")
        return RedirectResponse(dest, status_code=303)
    if not task_allowed_for_project(db, project_id, task_type_id):
        flash(request, f"'{task.name}' isn't set up for '{project.name}' — link them under Lists first.", "err")
        return RedirectResponse(dest, status_code=303)
    client_err = _client_required_error(project, client)
    if client_err:
        flash(request, client_err, "err")
        return RedirectResponse(dest, status_code=303)
    sub = db.execute(
        select(m.DaySubmission).where(m.DaySubmission.employee_id == emp_id, m.DaySubmission.date == d)
    ).scalar_one_or_none()
    if sub is not None and sub.locked:
        flash(request, f"{d.strftime('%m/%d/%Y')} is already submitted and locked.", "err")
        return RedirectResponse(dest, status_code=303)

    plan = m.PlannedTask(
        employee_id=emp_id, date=d, project_id=project_id, task_type_id=task_type_id,
        details=capitalize_first(details.strip()), client=client.strip(), status=m.PLAN_PLANNED,
        created_by_employee_id=admin.id,
    )
    db.add(plan)
    db.commit()
    audit(
        db, admin.name, "assign_plan", "PlannedTask", str(plan.id),
        {"employee_id": emp_id, "date": d.isoformat(), "project": project.name, "task": task.name},
    )
    flash(request, f"Assigned to {emp.name}'s log for {d.strftime('%m/%d/%Y')}.", "ok")
    return RedirectResponse(dest, status_code=303)


@router.post("/plan/{plan_id}/edit")
def admin_edit_plan(
    plan_id: int,
    request: Request,
    project_id: int = Form(...),
    task_type_id: int = Form(...),
    date: str = Form(...),
    details: str = Form(""),
    client: str = Form(""),
    ym: str = Form(""),
    return_to: str = Form(""),
    admin: m.Employee = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """TK-04 — "Admin can edit ... the entry while status is Planned."
    Unlike the employee's own edit_plan() (plan text only), the admin can
    change Project/Task/Date/Note/Client too — a wrong pick when assigning
    shouldn't require delete-and-re-add. Resets assigned_notified_at back
    to None so the employee sees the assignment banner again for the
    updated version — same "notify again on a meaningful change" instinct
    _pending_edit_notices()'s own edited_at/employee_notified_at pair
    already has for suggestion renames, just simpler here since there's
    only ever one notice per plan rather than a rename history."""
    plan = db.get(m.PlannedTask, plan_id)
    if plan is None:
        raise HTTPException(404)
    led = admin_department_scope(admin)
    emp = db.get(m.Employee, plan.employee_id)
    if emp is None or (led is not None and (emp.department or "—") != led):
        raise Forbidden()
    dest = _plan_redirect(plan.employee_id, ym, return_to)
    if plan.status != m.PLAN_PLANNED:
        flash(request, "Only a not-yet-started plan can be edited.", "err")
        return RedirectResponse(dest, status_code=303)
    try:
        d = parse_date_field(date)
    except FormError as e:
        flash(request, e.message, "err")
        return RedirectResponse(dest, status_code=303)
    project = db.get(m.Project, project_id)
    task = db.get(m.TaskType, task_type_id)
    if project is None or not project.active or task is None or not task.active:
        flash(request, "Choose a Project and Task.", "err")
        return RedirectResponse(dest, status_code=303)
    if not task_allowed_for_project(db, project_id, task_type_id):
        flash(request, f"'{task.name}' isn't set up for '{project.name}' — link them under Lists first.", "err")
        return RedirectResponse(dest, status_code=303)
    client_err = _client_required_error(project, client)
    if client_err:
        flash(request, client_err, "err")
        return RedirectResponse(dest, status_code=303)

    plan.project_id = project_id
    plan.task_type_id = task_type_id
    plan.date = d
    plan.details = capitalize_first(details.strip())
    plan.client = client.strip()
    if plan.created_by_employee_id is not None and plan.created_by_employee_id != plan.employee_id:
        plan.assigned_notified_at = None
    db.commit()
    audit(db, admin.name, "edit_assigned_plan", "PlannedTask", str(plan.id), {})
    flash(request, "Updated.", "ok")
    return RedirectResponse(dest, status_code=303)


@router.post("/plan/{plan_id}/delete")
def admin_delete_plan(
    plan_id: int,
    request: Request,
    ym: str = Form(""),
    return_to: str = Form(""),
    admin: m.Employee = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """TK-04 — "Admin can ... remove the entry while status is Planned."
    Works on any still-Planned row in the admin's scope, not just ones
    they personally created — same general oversight precedent Person
    Detail's other write actions (Override, Compensation links) already
    have over an employee's own data."""
    plan = db.get(m.PlannedTask, plan_id)
    if plan is None:
        raise HTTPException(404)
    led = admin_department_scope(admin)
    emp = db.get(m.Employee, plan.employee_id)
    if emp is None or (led is not None and (emp.department or "—") != led):
        raise Forbidden()
    employee_id = plan.employee_id
    dest = _plan_redirect(employee_id, ym, return_to)
    if plan.status != m.PLAN_PLANNED:
        flash(request, "Only a not-yet-started plan can be removed.", "err")
        return RedirectResponse(dest, status_code=303)
    db.delete(plan)
    db.commit()
    audit(db, admin.name, "delete_assigned_plan", "PlannedTask", str(plan_id), {})
    flash(request, "Removed.", "ok")
    return RedirectResponse(dest, status_code=303)


@router.post("/unlock-request/{req_id}/reject")
def reject_unlock_request(
    req_id: int,
    request: Request,
    review_note: str = Form(""),
    ym: str = Form(""),
    admin: m.Employee = Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    """Decline an unlock request without unlocking the day (Ganesh,
    2026-08-27) — e.g. the day's fine as it is, or the correction should
    happen a different way. No DaySubmission change at all; the request
    just stops showing up as pending."""
    req = db.get(m.UnlockRequest, req_id)
    if req is None:
        return RedirectResponse("/admin", status_code=303)
    emp_id = req.employee_id
    if req.status == m.LEAVE_REQUESTED:
        req.status = m.LEAVE_REJECTED
        req.reviewed_by = admin.name
        req.reviewed_at = dt.datetime.utcnow()
        req.review_note = review_note.strip()
        db.commit()
        audit(db, admin.name, "reject_unlock_request", "UnlockRequest", f"{emp_id}:{req.date}", {"note": review_note.strip()})
        flash(request, "Unlock request declined.", "ok")
    return RedirectResponse(f"/admin/person/{emp_id}?ym={ym}", status_code=303)


@router.post("/person/{emp_id}/override")
def override_day(
    emp_id: int,
    request: Request,
    date: str = Form(...),
    status: str = Form(""),
    reason: str = Form(""),
    ym: str = Form(""),
    admin: m.Employee = Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    emp = db.get(m.Employee, emp_id)
    try:
        day = parse_date_field(date)
    except FormError as e:
        flash(request, e.message, "err")
        return RedirectResponse(f"/admin/person/{emp_id}?ym={ym}", status_code=303)
    row = db.execute(
        select(m.DayStatus).where(
            m.DayStatus.employee_id == emp_id, m.DayStatus.date == day
        )
    ).scalar_one_or_none()
    if row is None:
        # materialize a base row so the override has something to sit on
        engine.recompute_employee(db, emp, day, day)
        row = db.execute(
            select(m.DayStatus).where(
                m.DayStatus.employee_id == emp_id, m.DayStatus.date == day
            )
        ).scalar_one_or_none()
        if row is None:
            row = m.DayStatus(
                employee_id=emp_id, date=day, status=m.MISSING,
                actual_minutes=0, target_minutes=emp.daily_target_minutes,
                variance_minutes=None, source="computed",
            )
            db.add(row)
            db.commit()
    before = row.override_status or row.status
    if status == "":  # clear override
        row.override_status = None
        row.override_reason = ""
        row.override_by = ""
        row.override_at = None
        db.commit()
        audit(db, admin.name, "clear_override", "DayStatus", f"{emp_id}:{date}", {"was": before})
        flash(request, f"Override cleared for {date}.", "ok")
    else:
        if not reason.strip():
            flash(request, "Override requires a reason (it changes the strike count).", "err")
            return RedirectResponse(f"/admin/person/{emp_id}?ym={ym}", status_code=303)
        row.override_status = status
        row.override_reason = reason.strip()
        row.override_by = admin.name
        row.override_at = dt.datetime.utcnow()
        db.commit()
        audit(
            db, admin.name, "override_status", "DayStatus", f"{emp_id}:{date}",
            {"before": before, "after": status, "reason": reason.strip()},
        )
        flash(request, f"Status for {date} overridden to {status}.", "ok")
    return RedirectResponse(f"/admin/person/{emp_id}?ym={ym}", status_code=303)


def _complink_redirect(emp_id: int, ym: str, return_to: str) -> str:
    """Where add_complink sends the admin back to. Defaults to Person Detail
    (the original, still-used flow) — Overtime Management's quick-link form
    (Ganesh, 2026-08-21) sets return_to=overtime so creating a link from
    there doesn't bounce the admin to a page they didn't come from; it also
    keeps the same employee selected so the result shows immediately."""
    if return_to == "overtime":
        return f"/admin/overtime?ym={ym}&employee_id={emp_id}"
    return f"/admin/person/{emp_id}?ym={ym}"


@router.post("/person/{emp_id}/complink")
def add_complink(
    emp_id: int,
    request: Request,
    shortfall_date: str = Form(...),
    surplus_dates: list = Form([]),  # checkboxes, same convention as assignments_save's project_ids
    note: str = Form(""),
    ym: str = Form(""),
    return_to: str = Form(""),
    admin: m.Employee = Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    dest = _complink_redirect(emp_id, ym, return_to)
    try:
        shortfall = parse_date_field(shortfall_date, "Shortfall date")
        surplus = sorted({dt.date.fromisoformat(x.strip()).isoformat()
                          for x in surplus_dates if x.strip()})
    except (FormError, ValueError) as e:
        flash(request, e.message if isinstance(e, FormError)
              else "Surplus dates must be valid ISO dates.", "err")
        return RedirectResponse(dest, status_code=303)
    if not surplus:
        flash(request, "Pick at least one surplus day.", "err")
        return RedirectResponse(dest, status_code=303)

    # Partial allocation (Ganesh, 2026-08-25) — replaces the old whole-day
    # "surplus day already linked to another shortfall" clash rejection.
    # A surplus day can now back more than one shortfall over time (just
    # not more minutes than it actually has), and a shortfall can be
    # covered by more than one link (e.g. link what's available now, link
    # the rest later once there's more overtime to draw on). See
    # engine.evaluate_link()/shortfall_allocated_minutes()/
    # surplus_minutes_used_by_date() for the read side of this same model.
    short_row = db.execute(
        select(m.DayStatus).where(m.DayStatus.employee_id == emp_id, m.DayStatus.date == shortfall)
    ).scalar_one_or_none()
    full_deficit = -(short_row.variance_minutes or 0) if short_row is not None and (short_row.variance_minutes or 0) < 0 else 0
    deficit_remaining = full_deficit - engine.shortfall_allocated_minutes(db, emp_id, shortfall)
    if full_deficit <= 0:
        flash(request, f"{shortfall_date} isn't a shortfall day — nothing to link.", "err")
        return RedirectResponse(dest, status_code=303)
    if deficit_remaining <= 0:
        flash(request, f"{shortfall_date} is already fully compensated — nothing left to link.", "err")
        return RedirectResponse(dest, status_code=303)

    # Ticked days are consumed oldest-first (surplus is already date-sorted
    # above) until the remaining deficit is covered; a day with nothing left
    # (or ticked past the point the deficit's already met) is simply left
    # untouched — skipped, not blocked or errored on — so it stays 100%
    # available for a future link rather than being wasted or rejected. See
    # engine.allocate_surplus_minutes() — shared with the employee
    # self-service match-request flow (app/routes/employee.py) so the two
    # can't drift apart.
    allocation = engine.allocate_surplus_minutes(db, emp_id, shortfall, surplus)
    deficit_remaining -= sum(allocation.values())

    if not allocation:
        flash(request, "None of the selected surplus day(s) have any hours left to link — pick a different day.", "err")
        return RedirectResponse(dest, status_code=303)

    link = m.CompensationLink(
        employee_id=emp_id,
        shortfall_date=shortfall,
        surplus_dates=json.dumps(sorted(allocation.keys())),
        surplus_minutes=json.dumps(allocation),
        note=note.strip(),
        linked_by=admin.name,
    )
    db.add(link)
    db.commit()
    engine.evaluate_link(db, link)
    audit(
        db, admin.name, "compensation_link", "CompensationLink", link.id,
        {"shortfall": shortfall_date, "allocation": allocation, "fully": link.fully_compensated,
         "note": note.strip()},
    )
    # deficit_remaining here is the state AFTER this new link (it started
    # from shortfall_allocated_minutes(), which already accounted for any
    # earlier links on this same day) — that's the day's true aggregate
    # outcome, which is what short_row.compensated now reflects too. Don't
    # use link.fully_compensated for this message: it's this ONE link's own
    # coverage, correctly False for a second link that finishes off a day
    # an earlier partial link started, even though the day itself is done.
    if deficit_remaining <= 0:
        detail = " — shortfall fully covered, day now reads Complete."
    else:
        detail = f" — {fmt_hm(deficit_remaining)} still short; link more surplus day(s) to finish covering it."
    flash(request, "Compensation link created" + detail, "ok")
    return RedirectResponse(dest, status_code=303)


@router.post("/complink/{link_id}/delete")
def delete_complink(
    link_id: int,
    request: Request,
    ym: str = Form(""),
    admin: m.Employee = Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    link = db.get(m.CompensationLink, link_id)
    if link is not None:
        emp_id = link.employee_id
        row = db.execute(
            select(m.DayStatus).where(
                m.DayStatus.employee_id == emp_id, m.DayStatus.date == link.shortfall_date
            )
        ).scalar_one_or_none()
        if row is not None:
            row.compensated = False
        db.delete(link)
        db.commit()
        audit(db, admin.name, "delete_compensation_link", "CompensationLink", link_id,
              {"shortfall": link.shortfall_date.isoformat()})
        return RedirectResponse(f"/admin/person/{emp_id}?ym={ym}", status_code=303)
    return RedirectResponse("/admin", status_code=303)


# Requirement 9 (Overtime-for-Missed-Hours, employee-requested match,
# Ganesh, 2026-08-21) — the employee submits the match via
# POST /leave/match-request (app/routes/employee.py), which creates a
# CompensationLink with status=LEAVE_REQUESTED, requested_by_employee=True.
# These two routes are the admin decision on that request: approve (which
# re-checks engine.compensation_window_ok for every surplus date, then runs
# the same engine.evaluate_link the admin-direct add_complink path already
# uses) or reject (no engine side effects — the day's compliance status is
# untouched since it was never marked compensated).
@router.post("/complink/{link_id}/approve")
def approve_complink(
    link_id: int,
    request: Request,
    review_note: str = Form(""),
    admin: m.Employee = Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    if not LEAVE_MANAGEMENT_V2_ENABLED:
        raise HTTPException(status_code=404)
    link = db.get(m.CompensationLink, link_id)
    if link is None:
        flash(request, "Match request not found.", "err")
        return RedirectResponse("/admin/leave", status_code=303)
    surplus_dates = [dt.date.fromisoformat(x) for x in json.loads(link.surplus_dates or "[]")]
    bad = [s.isoformat() for s in surplus_dates if not engine.compensation_window_ok(link.shortfall_date, s)]
    if bad:
        flash(
            request,
            f"Can't approve — surplus day(s) fall outside the compensation window: {', '.join(bad)}",
            "err",
        )
        return RedirectResponse("/admin/leave", status_code=303)
    link.status = m.LEAVE_APPROVED
    link.reviewed_by = admin.name
    link.reviewed_at = dt.datetime.utcnow()
    link.review_note = review_note.strip()
    db.commit()
    engine.evaluate_link(db, link)
    audit(
        db, admin.name, "approve_compensation_match", "CompensationLink", link.id,
        {"shortfall": link.shortfall_date.isoformat(), "surplus": [s.isoformat() for s in surplus_dates],
         "fully": link.fully_compensated, "review_note": review_note.strip()},
    )
    # Same reasoning as add_complink()'s flash message (Ganesh, 2026-08-25):
    # the day's TRUE aggregate state (across every link that targets it, not
    # just this one) is what actually matters here, in case this request
    # happens to be the one that finishes off a day an earlier link already
    # partially covered.
    short_row = db.execute(
        select(m.DayStatus).where(m.DayStatus.employee_id == link.employee_id, m.DayStatus.date == link.shortfall_date)
    ).scalar_one_or_none()
    day_deficit = 0
    if short_row is not None and (short_row.variance_minutes or 0) < 0:
        day_deficit = -(short_row.variance_minutes or 0)
    day_remaining = day_deficit - engine.shortfall_allocated_minutes(db, link.employee_id, link.shortfall_date)
    flash(
        request,
        "Match approved"
        + (" — shortfall fully covered, day now reads Complete." if day_deficit > 0 and day_remaining <= 0
           else f" — {fmt_hm(day_remaining)} still short; another link can finish covering it." if day_remaining > 0
           else " — not yet fully covered (kept as-is)."),
        "ok",
    )
    return RedirectResponse("/admin/leave", status_code=303)


@router.post("/complink/{link_id}/reject")
def reject_complink(
    link_id: int,
    request: Request,
    review_note: str = Form(""),
    admin: m.Employee = Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    if not LEAVE_MANAGEMENT_V2_ENABLED:
        raise HTTPException(status_code=404)
    link = db.get(m.CompensationLink, link_id)
    if link is None:
        flash(request, "Match request not found.", "err")
        return RedirectResponse("/admin/leave", status_code=303)
    link.status = m.LEAVE_REJECTED
    link.reviewed_by = admin.name
    link.reviewed_at = dt.datetime.utcnow()
    link.review_note = review_note.strip()
    db.commit()
    audit(db, admin.name, "reject_compensation_match", "CompensationLink", link.id,
          {"shortfall": link.shortfall_date.isoformat(), "review_note": review_note.strip()})
    flash(request, "Match request declined.", "ok")
    return RedirectResponse("/admin/leave", status_code=303)


# --------------------------------------------------------------------------
# Roster
# --------------------------------------------------------------------------
@router.get("/roster")
def roster(
    request: Request,
    show: str = "active",
    admin: m.Employee = Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    q = select(m.Employee).order_by(m.Employee.active.desc(), m.Employee.department, m.Employee.name)
    emps = list(db.execute(q).scalars())
    # Two distinct lists, not "active" vs. "everyone mixed together" (Ganesh,
    # 2026-08-25 — "when i offboard the employee the offboarded employees
    # should be under deactivated list, options should be active-Employees
    # and Ex-Employees"): offboarding someone now visibly moves them into
    # their own Ex-Employees view instead of just quietly sinking to the
    # bottom of one combined table. `show="all"` (the old value) is no
    # longer offered from the toggle above, but still falls through to
    # "everyone, active first" here rather than erroring, in case a bookmark
    # or old link still points at it.
    if show == "active":
        emps = [e for e in emps if e.active]
    elif show == "ex":
        emps = [e for e in emps if not e.active]

    # department pill row above the table — counts reflect whatever's
    # actually listed (respects the active/all toggle above)
    dept_counts: dict = {}
    for e in emps:
        key = e.department or "—"
        dept_counts[key] = dept_counts.get(key, 0) + 1
    all_depts = sorted(dept_counts)
    # dropdown always offers active people to report to, regardless of the
    # active/all toggle above (which only controls the table listing).
    # is_admin-only (Ganesh's manager, 2026-08-03): a "Reports to" pick is
    # now also that employee's Team Lead for Overtime Requests (see
    # app/auth.py led_by()) — someone with no admin access couldn't act on
    # a request routed to them anyway, so they're not offered as a choice.
    lead_choices = [e for e in emps if e.active and e.is_admin]

    return render(
        request, "admin/roster.html",
        {
            "user": admin, "emps": emps, "show": show, "dept_counts": dept_counts,
            "all_depts": all_depts, "lead_choices": lead_choices,
            "locations": m.LOCATIONS, "default_location": m.DEFAULT_LOCATION,
        },
        db=db,
    )


def _emp_from_form(
    db: Session, emp: m.Employee, name, email, department, designation, target_hours,
    work_days, start_date, active, tracked, role, dob="", phone="", country_code="",
    reports_to_id="", is_developer=False, location=None,
    is_on_pip=None, probation_days="",
):
    emp.name = name.strip()
    # Work location / country (Ganesh, 2026-08-12) — admin-set here as an
    # alternative to the employee's own Profile self-service dropdown, same
    # dual-editable pattern as department/designation. Only touched when the
    # caller actually passes one (both roster forms always submit a value,
    # since it's a required <select>, but keeping this guarded means a
    # future caller of _emp_from_form doesn't have to supply it to avoid
    # silently resetting an existing employee back to the default).
    if location in m.LOCATIONS:
        emp.location = location
    emp.country_code = country_code.strip() or None
    emp.phone = phone.strip() or None
    # None (not "") so multiple blank emails never collide on the unique
    # constraint; signup later matches an employee by this exact field.
    new_email = email.strip() or None
    if new_email is not None:
        clash = db.execute(
            select(m.Employee).where(
                m.Employee.email == new_email, m.Employee.id != (emp.id or -1)
            )
        ).scalar_one_or_none()
        if clash is not None:
            raise FormError(f"Email '{new_email}' is already used by {clash.name}.")
    emp.email = new_email
    emp.department = department.strip()
    emp.designation = designation.strip()
    target_minutes = parse_hours_field(target_hours or "8", "Daily target hours")
    if target_minutes <= 0:
        raise FormError("Daily target hours must be greater than zero.")
    emp.daily_target_minutes = target_minutes
    try:
        emp.work_days = (
            ",".join(str(d) for d in sorted({int(x) for x in work_days})) if work_days else "0,1,2,3,4"
        )
    except (ValueError, TypeError):
        raise FormError("Work days must be whole numbers 0-6 (Monday=0).")
    emp.start_date = parse_date_field(start_date, "Start date") if start_date else None
    emp.date_of_birth = parse_date_field(dob, "Date of birth") if dob else None
    emp.active = bool(active)
    emp.tracked = bool(tracked)
    emp.is_admin, emp.is_super_admin = role_to_flags(role)
    # Ticketing System (Ganesh, 2026-08-06) — Developer is a third,
    # independent axis, not tied to role at all (see Employee.is_developer's
    # docstring): a plain employee can be a developer just as much as an
    # admin can, so this isn't part of role_to_flags. While TICKETING_ENABLED
    # is off the checkbox is hidden from both roster forms (see
    # app/templating.py), so the form never submits this field at all —
    # leave the stored value untouched rather than force it to False on
    # every save, which would silently wipe it the instant the flag flips
    # back on.
    if TICKETING_ENABLED:
        emp.is_developer = bool(is_developer)
    # blank/"0" -> no reporting lead set; never allow someone to report to
    # themselves, or to someone who isn't an admin (silently ignored rather
    # than a hard error — a stray/crafted selection shouldn't block saving
    # the rest of the form). The is_admin check guards the same invariant
    # the roster/roster_edit dropdowns already enforce by only listing
    # admins as choices (Ganesh's manager, 2026-08-03: "Reports to" is now
    # also Team Lead authority for Overtime Requests, see app/auth.py
    # led_by() — a non-admin could never act on anything routed to them).
    rid = int(reports_to_id) if str(reports_to_id).strip().isdigit() else None
    if rid and rid != emp.id:
        lead = db.get(m.Employee, rid)
        rid = rid if lead is not None and lead.is_admin else None
    emp.reports_to_id = rid or None
    # Leave Management V2 (Ganesh, 2026-08-21) — same "guarded by the flag"
    # convention as is_developer/TICKETING_ENABLED just above: while
    # LEAVE_MANAGEMENT_V2_ENABLED is off, roster_edit.html doesn't render
    # these fields at all, so is_on_pip stays None here and nothing is
    # touched — leave whatever was already stored untouched rather than
    # silently reset it every save.
    if is_on_pip is not None:
        emp.is_on_pip = bool(is_on_pip)
        probation_days = (probation_days or "").strip()
        if probation_days == "":
            emp.probation_days = None  # explicit blank -> "use the company default"
        else:
            try:
                pd = int(probation_days)
            except ValueError:
                raise FormError("Probation days must be a whole number (blank = use the company default).")
            if pd < 0:
                raise FormError("Probation days can't be negative.")
            emp.probation_days = pd


@router.post("/roster/add")
def roster_add(
    request: Request,
    name: str = Form(...),
    email: str = Form(""),
    department: str = Form(""),
    designation: str = Form(""),
    target_hours: str = Form("8"),
    work_days: list = Form(["0", "1", "2", "3", "4"]),
    start_date: str = Form(""),
    dob: str = Form(""),
    country_code: str = Form(""),
    phone: str = Form(""),
    active: str = Form("1"),
    tracked: str = Form("1"),
    role: str = Form(ROLE_EMPLOYEE),
    reports_to_id: str = Form(""),
    is_developer: str = Form(""),
    location: str = Form(m.DEFAULT_LOCATION),
    admin: m.Employee = Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    emp = m.Employee(name=name.strip())
    try:
        _emp_from_form(db, emp, name, email, department, designation, target_hours,
                       work_days, start_date, active == "1", tracked == "1", role,
                       dob, phone, country_code, reports_to_id, is_developer == "1",
                       # Held back from the 2026-08-13 deploy: Country is hidden from
                       # this form, so `location` is just Form()'s unsubmitted-field
                       # default, not an admin choice — None leaves the new
                       # Employee row on its own model-level default (still India).
                       location if HOLIDAY_MANAGEMENT_ENABLED else None)
    except FormError as e:
        flash(request, e.message, "err")
        return RedirectResponse("/admin/roster", status_code=303)
    emp.employee_code = next_employee_code(db)
    db.add(emp)
    db.commit()
    audit(db, admin.name, "roster_add", "Employee", emp.id, {"name": emp.name})
    flash(request, f"Added {emp.name}.", "ok")
    return RedirectResponse("/admin/roster", status_code=303)


@router.post("/roster/{emp_id}/reset-password")
def roster_reset_password(
    emp_id: int,
    request: Request,
    admin: m.Employee = Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    """Clears the stored hash so the employee can re-run /signup. Stand-in
    for a self-service reset flow until there's an email service to send
    reset links from (AUTH_MODE=password only has meaning if there's a
    password to clear)."""
    emp = db.get(m.Employee, emp_id)
    if emp is None:
        return RedirectResponse("/admin/roster", status_code=303)
    emp.password_hash = None
    db.commit()
    audit(db, admin.name, "reset_password", "Employee", emp.id, {"name": emp.name})
    flash(request, f"Password cleared for {emp.name} — they can run Sign up again.", "ok")
    return RedirectResponse(f"/admin/roster/{emp_id}/edit", status_code=303)


@router.get("/roster/{emp_id}/edit")
def roster_edit_page(
    emp_id: int,
    request: Request,
    admin: m.Employee = Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    emp = db.get(m.Employee, emp_id)
    if emp is None:
        return RedirectResponse("/admin/roster", status_code=303)
    # is_admin-only — see the matching comment in roster_page's lead_choices.
    lead_choices = list(
        db.execute(
            select(m.Employee)
            .where(m.Employee.active.is_(True), m.Employee.is_admin.is_(True), m.Employee.id != emp_id)
            .order_by(m.Employee.name)
        ).scalars()
    )
    cfg = engine.get_config(db)
    return render(
        request, "admin/roster_edit.html",
        {
            "user": admin, "emp": emp, "lead_choices": lead_choices, "locations": m.LOCATIONS,
            "leave_probation_days_default": engine.cfg_int(cfg, "probation_days_default"),
        },
        db=db,
    )


@router.post("/roster/{emp_id}/edit")
def roster_edit(
    emp_id: int,
    request: Request,
    name: str = Form(...),
    email: str = Form(""),
    department: str = Form(""),
    designation: str = Form(""),
    target_hours: str = Form("8"),
    work_days: list = Form([]),
    start_date: str = Form(""),
    dob: str = Form(""),
    country_code: str = Form(""),
    phone: str = Form(""),
    active: str = Form(""),
    tracked: str = Form(""),
    role: str = Form(ROLE_EMPLOYEE),
    reports_to_id: str = Form(""),
    is_developer: str = Form(""),
    location: str = Form(m.DEFAULT_LOCATION),
    is_on_pip: str = Form(""),
    probation_days: str = Form(""),
    admin: m.Employee = Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    emp = db.get(m.Employee, emp_id)
    if emp is None:
        return RedirectResponse("/admin/roster", status_code=303)
    before = {
        "name": emp.name, "department": emp.department, "designation": emp.designation,
        "target": emp.daily_target_minutes, "work_days": emp.work_days,
        "active": emp.active, "tracked": emp.tracked, "is_admin": emp.is_admin,
        "is_super_admin": emp.is_super_admin, "is_developer": emp.is_developer,
        "location": emp.location, "is_on_pip": emp.is_on_pip, "probation_days": emp.probation_days,
    }
    try:
        _emp_from_form(db, emp, name, email, department, designation, target_hours,
                       work_days, start_date, active == "1", tracked == "1", role,
                       dob, phone, country_code, reports_to_id, is_developer == "1",
                       # Held back from the 2026-08-13 deploy: Country is hidden from
                       # this form, so `location` is just Form()'s unsubmitted-field
                       # default, not an admin choice — None leaves emp.location
                       # exactly as it was (see _emp_from_form's own guard); without
                       # this, every roster edit would silently reset location back
                       # to India regardless of what it was set to.
                       location if HOLIDAY_MANAGEMENT_ENABLED else None,
                       # Same guard, same reasoning, for Leave Management V2's
                       # PIP/probation fields — is_on_pip=None (not "0"/"1")
                       # while the flag is off means _emp_from_form leaves
                       # both untouched instead of resetting them.
                       (is_on_pip == "1") if LEAVE_MANAGEMENT_V2_ENABLED else None,
                       probation_days)
    except FormError as e:
        flash(request, e.message, "err")
        return RedirectResponse("/admin/roster", status_code=303)
    db.commit()
    audit(db, admin.name, "roster_edit", "Employee", emp.id, {"before": before})
    flash(request, f"Saved {emp.name}." + ("" if emp.active else " (deactivated — history kept, dropped from compliance runs)"), "ok")
    return RedirectResponse("/admin/roster", status_code=303)


# --------------------------------------------------------------------------
# Bulk upload (Roster -> Bulk upload) — parsing rules live in app/bulk_upload.py
# --------------------------------------------------------------------------
@router.get("/roster/bulk-upload")
def roster_bulk_upload_page(
    request: Request,
    admin: m.Employee = Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    return render(request, "admin/roster_bulk_upload.html", {"user": admin, "result": None}, db=db)


@router.get("/roster/bulk-upload/sample.xlsx")
def roster_bulk_upload_sample(admin: m.Employee = Depends(require_super_admin)):
    buf = io.BytesIO()
    bulk_upload.build_sample_workbook().save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="employee_upload_template.xlsx"'},
    )


@router.get("/roster/bulk-upload/existing.xlsx")
def roster_bulk_upload_existing(
    admin: m.Employee = Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    buf = io.BytesIO()
    bulk_upload.build_existing_employees_workbook(db).save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="existing_employees.xlsx"'},
    )


@router.post("/roster/bulk-upload")
def roster_bulk_upload(
    request: Request,
    file: UploadFile = File(...),
    admin: m.Employee = Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xlsm")):
        flash(request, "Please upload an .xlsx file — use the sample template.", "err")
        return RedirectResponse("/admin/roster/bulk-upload", status_code=303)
    try:
        wb = load_workbook(io.BytesIO(file.file.read()), data_only=True)
    except Exception:
        flash(request, "Couldn't read that file — is it a valid, unprotected .xlsx?", "err")
        return RedirectResponse("/admin/roster/bulk-upload", status_code=303)

    result = bulk_upload.process_upload(db, wb)
    if result["header_error"]:
        flash(request, result["header_error"], "err")
        return RedirectResponse("/admin/roster/bulk-upload", status_code=303)
    if result["added"] or result["updated"]:
        audit(db, admin.name, "roster_bulk_upload", "Employee", "",
              {"added": result["added"], "updated": result["updated"],
               "deactivated": result["deactivated"], "skipped": len(result["skipped"])})
    summary = f"{result['added']} employee(s) added, {result['updated']} updated"
    summary += f" ({result['deactivated']} deactivated)." if result["deactivated"] else "."
    if result["skipped"]:
        summary += f" {len(result['skipped'])} row(s) skipped — see details below."
    flash(request, summary, "ok" if (result["added"] or result["updated"]) else "err")
    return render(request, "admin/roster_bulk_upload.html", {"user": admin, "result": result}, db=db)


# --------------------------------------------------------------------------
# Lists (Project/Employer + Task dropdowns)
# --------------------------------------------------------------------------
@router.get("/lists")
def lists_page(
    request: Request,
    admin: m.Employee = Depends(require_admin),
    db: Session = Depends(get_db),
):
    projects = list(db.execute(select(m.Project).order_by(m.Project.active.desc(), m.Project.name)).scalars())
    tasks = list(db.execute(select(m.TaskType).order_by(m.TaskType.active.desc(), m.TaskType.name)).scalars())
    # Project-scoped tasks (Ganesh, 2026-08-27) — task_project_ids feeds
    # both the read-only "Projects" column and the pre-checked state of
    # each task's "Manage projects" panel below. A task_type_id absent
    # here (empty list) has no links yet, i.e. still unrestricted — see
    # ProjectTask's docstring in app/models.py.
    task_project_ids: Dict[int, list] = {t.id: [] for t in tasks}
    for tid, pid in db.execute(select(m.ProjectTask.task_type_id, m.ProjectTask.project_id)).all():
        task_project_ids.setdefault(tid, []).append(pid)
    # Department-scoped projects (Ganesh, 2026-08-28) — project_department_
    # names feeds both the read-only department badges and the pre-checked
    # state of each project's "Manage departments" panel below. A
    # project.id absent here (empty list) has no links yet, i.e. still
    # unrestricted — see ProjectDepartment's docstring in app/models.py.
    # all_departments is every department name currently in use by an
    # active/tracked employee (reports.departments_list) — this app has no
    # canonical department table, so that's the same source of truth the
    # Dashboard/Roster/Reports department pickers already use.
    project_department_names: Dict[int, list] = {p.id: [] for p in projects}
    for pid, dept in db.execute(select(m.ProjectDepartment.project_id, m.ProjectDepartment.department)).all():
        project_department_names.setdefault(pid, []).append(dept)
    all_departments = [d for d in reports.departments_list(db) if d != "—"]

    # Department -> Project -> Task tree (Ganesh, 2026-08-29, matching a
    # pasted mockup — "just the tree UI", confirmed via AskUserQuestion:
    # no new fields/entities, built entirely from what's already above.
    # A project with zero ProjectDepartment links is unrestricted (see that
    # model's docstring) — shown once under "shared_projects" rather than
    # copied into every real department group below, both because that's
    # what "unrestricted" actually means and because with ~300 real
    # Project/Employer rows and (so far) 0 department links in the live
    # data, copying every unscoped project into every department group
    # would make each group identical and the tree pointless. A project
    # linked to MORE THAN ONE department (allowed — see ProjectDepartment's
    # docstring) is shown once, under its alphabetically-first linked
    # department, to avoid rendering (and duplicate-DOM-id-ing) the same
    # project's edit panel more than once; its "Manage departments" panel
    # still lists every department it's actually linked to, unaffected by
    # which one it happens to be filed under here.
    shared_projects = [p for p in projects if not project_department_names.get(p.id)]
    dept_groups = []
    for d in all_departments:
        dept_projects = [
            p for p in projects
            if project_department_names.get(p.id) and d == sorted(project_department_names[p.id])[0]
        ]
        dept_groups.append({"name": d, "projects": dept_projects})

    # project_task_ids: project_id -> [task_type_id, ...] EXPLICITLY linked
    # to it (the inverse of task_project_ids above) — what a project's tree
    # node shows inline. A task with NO links at all (unrestricted, usable
    # under every project — the large majority right now, see ProjectTask's
    # docstring) is deliberately NOT duplicated under all ~300 projects;
    # it only ever shows once, in the "All Tasks" list below the tree,
    # which is also still the one place to narrow/rename/deactivate any
    # task regardless of whether it happens to be narrowed to a project.
    project_task_ids: Dict[int, list] = {p.id: [] for p in projects}
    for tid, pids in task_project_ids.items():
        for pid in pids:
            project_task_ids.setdefault(pid, []).append(tid)
    tasks_by_id = {t.id: t for t in tasks}
    unrestricted_task_count = sum(1 for t in tasks if not task_project_ids.get(t.id))
    # Plain "name, name, name" per project (not a Jinja custom filter) —
    # feeds the tree's client-side search box (data-tasks="...") so typing
    # a task name expands and shows the right project without JS needing
    # to know how tasks_by_id is shaped.
    project_task_names: Dict[int, str] = {
        pid: ", ".join(tasks_by_id[tid].name for tid in tids if tid in tasks_by_id)
        for pid, tids in project_task_ids.items()
    }

    return render(
        request, "admin/lists.html",
        {
            "user": admin, "projects": projects, "tasks": tasks, "task_project_ids": task_project_ids,
            "project_department_names": project_department_names, "all_departments": all_departments,
            "shared_projects": shared_projects, "dept_groups": dept_groups,
            "project_task_ids": project_task_ids, "tasks_by_id": tasks_by_id,
            "unrestricted_task_count": unrestricted_task_count, "project_task_names": project_task_names,
            # plain {id, name} dicts for the Add-task Project combo picker's
            # script tag — tojson (app/templating.py) refuses raw ORM rows.
            "active_projects_json": [{"id": p.id, "name": p.name} for p in projects if p.active],
        },
        db=db,
    )


@router.post("/lists/add")
def lists_add(
    request: Request,
    kind: str = Form(...),
    name: str = Form(...),
    project_id: int = Form(0),
    is_case_type: str = Form(""),
    admin: m.Employee = Depends(require_admin),
    db: Session = Depends(get_db),
):
    model = m.Project if kind == "project" else m.TaskType
    name = name.strip()
    if not name:
        return RedirectResponse("/admin/lists", status_code=303)
    # Project-scoped tasks (Ganesh, 2026-08-27) — adding a task here now
    # requires picking the (first) project it belongs to, same as a
    # suggested task now requires (see suggest_list_item() in
    # app/routes/employee.py) and the Tasks bulk-upload sheet's new
    # Project Name column. To link the same task name to MORE projects
    # later, use that task's "Manage projects" panel (lists_task_projects
    # below) rather than re-adding it here — the name is still unique.
    project = None
    if kind == "task":
        project = db.get(m.Project, project_id) if project_id else None
        if project is None:
            flash(request, "Choose which Project this task belongs to.", "err")
            return RedirectResponse("/admin/lists", status_code=303)
    exists = db.execute(select(model).where(model.name == name)).scalar_one_or_none()
    if exists is None:
        item = model(name=name)
        if kind == "project":
            # Case Type (Ganesh, 2026-08-28) — settable at creation by
            # whoever can add a project (both admin tiers, same as the
            # rest of Add); see Project.is_case_type's docstring.
            item.is_case_type = is_case_type == "on"
        db.add(item)
        if kind == "task":
            db.flush()  # need item.id for the ProjectTask link below
            db.add(m.ProjectTask(project_id=project.id, task_type_id=item.id, created_by=admin.name))
        db.commit()
        audit(db, admin.name, f"add_{kind}", kind, name, {"project": project.name if project else None})
    else:
        flash(request, f"'{name}' already exists.", "err")
    return RedirectResponse("/admin/lists", status_code=303)


@router.post("/lists/task/{task_id}/projects")
def lists_task_projects(
    task_id: int,
    request: Request,
    project_ids: list = Form([]),
    admin: m.Employee = Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    """Replace-all save of one task's linked projects (Ganesh, 2026-08-27)
    — same "replace-all: simplest correct way to sync a set of checkboxes"
    pattern assignments_save() already uses below. Saving with every box
    unchecked returns the task to unrestricted (usable under every
    project) rather than usable under none — an empty ProjectTask set
    means "no restriction", not "restricted to nothing" (see that model's
    docstring), so there's no separate "clear" action needed."""
    task = db.get(m.TaskType, task_id)
    if task is None:
        return RedirectResponse("/admin/lists", status_code=303)
    db.execute(delete(m.ProjectTask).where(m.ProjectTask.task_type_id == task_id))
    added = 0
    for pid in project_ids:
        db.add(m.ProjectTask(project_id=int(pid), task_type_id=task_id, created_by=admin.name))
        added += 1
    db.commit()
    audit(db, admin.name, "task_projects_save", "TaskType", task.name, {"project_count": added})
    flash(
        request,
        f"'{task.name}' is now {'usable under every project' if added == 0 else f'linked to {added} project(s)'}.",
        "ok",
    )
    return RedirectResponse("/admin/lists", status_code=303)


@router.post("/lists/project/{project_id}/departments")
def lists_project_departments(
    project_id: int,
    request: Request,
    departments: list = Form([]),
    other_department: str = Form(""),
    admin: m.Employee = Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    """Replace-all save of one project's linked departments (Ganesh,
    2026-08-28 — see ProjectDepartment's docstring in app/models.py).
    Same "replace-all: simplest correct way to sync a set of checkboxes"
    pattern lists_task_projects() above and assignments_save() below both
    use. Saving with every box unchecked (and no free-text addition)
    returns the project to unrestricted (usable by every department)
    rather than usable by none — an empty ProjectDepartment set means "no
    restriction", not "restricted to nobody" (see that model's
    docstring), so there's no separate "clear" action needed.

    other_department (Ganesh, 2026-08-28) — department has no canonical
    list anywhere in this app (see reports.departments_list()'s own
    docstring-adjacent reasoning), so the checkbox list is only ever
    "departments currently in use by an active employee" — this free-text
    box covers a department that doesn't have any active employees typed
    into Employee.department yet, without needing a separate admin screen
    to pre-declare department names before they can be used here."""
    project = db.get(m.Project, project_id)
    if project is None:
        return RedirectResponse("/admin/lists", status_code=303)
    names = {d.strip() for d in departments if d and d.strip()}
    extra = other_department.strip()
    if extra:
        names.add(extra)
    db.execute(delete(m.ProjectDepartment).where(m.ProjectDepartment.project_id == project_id))
    added = 0
    for dept in names:
        db.add(m.ProjectDepartment(project_id=project_id, department=dept, created_by=admin.name))
        added += 1
    db.commit()
    audit(db, admin.name, "project_departments_save", "Project", project.name, {"department_count": added})
    flash(
        request,
        f"'{project.name}' is now {'available to every department' if added == 0 else f'restricted to {added} department(s)'}.",
        "ok",
    )
    return RedirectResponse("/admin/lists", status_code=303)


# --------------------------------------------------------------------------
# Bulk upload (Projects & Tasks -> Bulk upload) — one column per sheet,
# add-only; parsing rules live in app/lists_bulk_upload.py
# --------------------------------------------------------------------------
@router.get("/lists/bulk-upload")
def lists_bulk_upload_page(
    request: Request,
    admin: m.Employee = Depends(require_admin),
    db: Session = Depends(get_db),
):
    return render(request, "admin/lists_bulk_upload.html", {"user": admin, "result": None, "result_kind": None}, db=db)


@router.get("/lists/bulk-upload/sample.xlsx")
def lists_bulk_upload_sample(
    kind: str,
    admin: m.Employee = Depends(require_admin),
    db: Session = Depends(get_db),
):
    buf = io.BytesIO()
    # db (Ganesh, 2026-08-27) — the Tasks sample template's "Projects"
    # reference sheet needs real project names, not placeholders; see
    # lists_bulk_upload.build_sample_workbook()'s docstring.
    lists_bulk_upload.build_sample_workbook(kind, db).save(buf)
    buf.seek(0)
    filename = "projects_upload_template.xlsx" if kind == "project" else "tasks_upload_template.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/lists/bulk-upload/existing.xlsx")
def lists_bulk_upload_existing(
    kind: str,
    admin: m.Employee = Depends(require_admin),
    db: Session = Depends(get_db),
):
    buf = io.BytesIO()
    lists_bulk_upload.build_existing_workbook(db, kind).save(buf)
    buf.seek(0)
    filename = "existing_projects.xlsx" if kind == "project" else "existing_tasks.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/lists/bulk-upload")
def lists_bulk_upload_post(
    request: Request,
    kind: str = Form(...),
    file: UploadFile = File(...),
    admin: m.Employee = Depends(require_admin),
    db: Session = Depends(get_db),
):
    if kind not in ("project", "task"):
        flash(request, "Unknown list — use the Projects or Tasks upload form.", "err")
        return RedirectResponse("/admin/lists/bulk-upload", status_code=303)
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xlsm")):
        flash(request, "Please upload an .xlsx file — use the sample template.", "err")
        return RedirectResponse("/admin/lists/bulk-upload", status_code=303)
    try:
        wb = load_workbook(io.BytesIO(file.file.read()), data_only=True)
    except Exception:
        flash(request, "Couldn't read that file — is it a valid, unprotected .xlsx?", "err")
        return RedirectResponse("/admin/lists/bulk-upload", status_code=303)

    result = lists_bulk_upload.process_upload(db, wb, kind)
    if result["header_error"]:
        flash(request, result["header_error"], "err")
        return RedirectResponse("/admin/lists/bulk-upload", status_code=303)
    # Department-scoped projects (Ganesh, 2026-08-28) — a Projects upload
    # can also add department links (to brand-new or already-existing
    # projects) without necessarily creating any new project rows, so
    # anything_added has to count both, not just result["added"] alone —
    # otherwise a file that only added department scoping would audit as
    # a no-op and flash red even though it changed something real.
    dept_links_added = result.get("department_links_added", 0)
    anything_added = result["added"] or dept_links_added
    if anything_added:
        audit(db, admin.name, f"lists_bulk_upload_{kind}", kind, "",
              {"added": result["added"], "department_links_added": dept_links_added, "skipped": len(result["skipped"])})
    label = "project(s)" if kind == "project" else "task(s)"
    summary = f"{result['added']} {label} added."
    if dept_links_added:
        summary += f" {dept_links_added} department link(s) added."
    if result["skipped"]:
        summary += f" {len(result['skipped'])} row(s) skipped — see details below."
    flash(request, summary, "ok" if anything_added else "err")
    return render(
        request, "admin/lists_bulk_upload.html",
        {"user": admin, "result": result, "result_kind": kind}, db=db,
    )


@router.post("/lists/{kind}/{item_id}/toggle")
def lists_toggle(
    kind: str,
    item_id: int,
    request: Request,
    admin: m.Employee = Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    model = m.Project if kind == "project" else m.TaskType
    item = db.get(model, item_id)
    if item is not None:
        # deactivating hides from new entries without breaking old rows (PRD §7)
        item.active = not item.active
        db.commit()
        audit(db, admin.name, f"toggle_{kind}", kind, item.name, {"active": item.active})
    return RedirectResponse("/admin/lists", status_code=303)


@router.post("/lists/project/{project_id}/case-type/toggle")
def lists_project_case_type_toggle(
    project_id: int,
    request: Request,
    admin: m.Employee = Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    """Flip Project.is_case_type on an existing project (Ganesh,
    2026-08-28 — see that column's docstring). Super-Admin-only, same
    tier as lists_toggle/lists_rename above — a department admin can mark
    Case Type at creation time (see lists_add's is_case_type form field,
    reachable by both tiers), but changing it on a LIVE project later is
    the same "corrects existing data, not just adds new" category as
    Rename/Deactivate, so it stays narrow on purpose."""
    project = db.get(m.Project, project_id)
    if project is not None:
        project.is_case_type = not project.is_case_type
        db.commit()
        audit(db, admin.name, "toggle_case_type", "project", project.name, {"is_case_type": project.is_case_type})
    return RedirectResponse("/admin/lists", status_code=303)


@router.post("/lists/{kind}/{item_id}/rename")
def lists_rename(
    kind: str,
    item_id: int,
    request: Request,
    name: str = Form(...),
    admin: m.Employee = Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    """Rename a live Project/Task from the Lists page (Norine, via Teams,
    2026-08-27: "can we add something to edit a project name and task
    name. maybe only by SuperAdmin in case there's a mistake" — e.g. she
    deactivated a project rather than fix it because there was no way to
    correct a typo'd name). SuperAdmin-only, same gating every other route
    on this page already uses (require_super_admin), not require_admin —
    matches what she asked for exactly.

    Deliberately a separate route from suggestion_edit() above, not a
    reuse of it — that one is scoped ONLY to still-pending suggestions
    (see its own docstring) and 403s outside a department-scoped admin's
    team; this one works on ANY Project/TaskType regardless of status
    (approved, pending, even rejected) or who created it, since a plain
    admin-added value (no suggestion involved at all) has no "pending"
    state to be scoped to in the first place. It does reuse the same
    original_name/edited_by/edited_at/employee_notified_at columns those
    suggestion routes already added to both models, though — renaming
    something here that happens to be an approved former employee
    suggestion still surfaces the "an admin rewrote your suggestion"
    banner via _pending_edit_notices() (app/routes/employee.py), for free,
    since that function only ever filters on created_by_employee_id +
    edited_at + employee_notified_at, never on status.

    Renaming is safe to do on a live value: every relationship that
    matters (TaskEntry.project_id/task_type_id, ProjectTask, leave/
    compensation records, etc.) is keyed by id, never by name — the name
    is purely a display string everywhere else in the app, so nothing
    downstream needs to be touched or re-pointed."""
    model = m.Project if kind == "project" else m.TaskType
    item = db.get(model, item_id)
    if item is None:
        return RedirectResponse("/admin/lists", status_code=303)
    new_name = name.strip()
    if not new_name:
        flash(request, "Enter a name.", "err")
        return RedirectResponse("/admin/lists", status_code=303)
    existing = db.execute(
        select(model).where(func.lower(model.name) == new_name.lower(), model.id != item.id)
    ).scalar_one_or_none()
    if existing is not None:
        flash(request, f"'{existing.name}' already exists — pick a different name.", "err")
        return RedirectResponse("/admin/lists", status_code=303)
    old_name = item.name
    if old_name == new_name:
        return RedirectResponse("/admin/lists", status_code=303)
    if item.original_name is None:
        item.original_name = old_name  # only ever captured once, same convention suggestion_edit() uses
    item.name = new_name
    item.edited_by = admin.name
    item.edited_at = dt.datetime.utcnow()
    item.employee_notified_at = None
    db.commit()
    audit(db, admin.name, f"rename_{kind}", kind, item.name, {"from": old_name, "to": new_name})
    flash(request, f"Renamed '{old_name}' to '{new_name}'.", "ok")
    return RedirectResponse("/admin/lists", status_code=303)


# --------------------------------------------------------------------------
# Suggestions (Ganesh, 2026-08-01) — employee/lead-suggested Projects/Tasks
# awaiting review. Deliberately require_admin, not require_super_admin,
# same tier as Leave Requests: "Team lead approval gatekeeping" was the
# explicit ask, so a department-scoped admin needs to be able to act on
# their own team's suggestions without needing org-wide access. Scoped by
# the SUBMITTER's department (a suggestion has no department of its own).
# --------------------------------------------------------------------------
def _pending_suggestions(db: Session, admin: m.Employee):
    scope = admin_department_scope(admin)
    projects = list(
        db.execute(
            select(m.Project).where(m.Project.status == m.LIST_PENDING).order_by(m.Project.created_at)
        ).scalars()
    )
    tasks = list(
        db.execute(
            select(m.TaskType).where(m.TaskType.status == m.LIST_PENDING).order_by(m.TaskType.created_at)
        ).scalars()
    )
    if scope is not None:
        projects = [p for p in projects if p.created_by and (p.created_by.department or "—") == scope]
        tasks = [t for t in tasks if t.created_by and (t.created_by.department or "—") == scope]
    return projects, tasks


def _approved_suggestions(db: Session, admin: m.Employee):
    """Suggestions that already went through this same queue and were
    approved (Ganesh, 2026-08-21 — the page previously only ever showed
    pending rows, so once approved a suggestion effectively disappeared
    from view). created_by_employee_id is not null is what distinguishes
    an employee suggestion from a Project/TaskType an admin added directly
    via Lists (those default straight to LIST_APPROVED with no submitter —
    see Project/TaskType docstrings) — without that filter this would just
    be "every active project/task," which is already what Lists shows."""
    scope = admin_department_scope(admin)
    projects = list(
        db.execute(
            select(m.Project)
            .where(m.Project.status == m.LIST_APPROVED, m.Project.created_by_employee_id.isnot(None))
            .order_by(m.Project.reviewed_at.desc())
        ).scalars()
    )
    tasks = list(
        db.execute(
            select(m.TaskType)
            .where(m.TaskType.status == m.LIST_APPROVED, m.TaskType.created_by_employee_id.isnot(None))
            .order_by(m.TaskType.reviewed_at.desc())
        ).scalars()
    )
    if scope is not None:
        projects = [p for p in projects if p.created_by and (p.created_by.department or "—") == scope]
        tasks = [t for t in tasks if t.created_by and (t.created_by.department or "—") == scope]
    return projects[:100], tasks[:100]


@router.get("/suggestions")
def suggestions_page(
    request: Request,
    admin: m.Employee = Depends(require_admin),
    db: Session = Depends(get_db),
):
    projects, tasks = _pending_suggestions(db, admin)
    approved_projects, approved_tasks = _approved_suggestions(db, admin)
    # Project-scoped tasks (Ganesh, 2026-08-27) — every suggested task now
    # carries a ProjectTask link set at suggestion time (see
    # suggest_list_item() in app/routes/employee.py), so an admin
    # reviewing it can see which project it was meant for instead of just
    # a bare name. One task suggestion -> one project by construction
    # (suggest_list_item only ever creates one link), but read as a list
    # for display consistency with the Lists page's own "N project(s)".
    all_task_ids = [t.id for t in tasks] + [t.id for t in approved_tasks]
    task_suggestion_projects: Dict[int, list] = {tid: [] for tid in all_task_ids}
    if all_task_ids:
        for tid, pname in db.execute(
            select(m.ProjectTask.task_type_id, m.Project.name)
            .join(m.Project, m.Project.id == m.ProjectTask.project_id)
            .where(m.ProjectTask.task_type_id.in_(all_task_ids))
        ).all():
            task_suggestion_projects.setdefault(tid, []).append(pname)
    return render(
        request, "admin/suggestions.html",
        {
            "user": admin, "projects": projects, "tasks": tasks,
            "approved_projects": approved_projects, "approved_tasks": approved_tasks,
            "task_suggestion_projects": task_suggestion_projects,
        }, db=db,
    )


def _suggestion_or_forbidden(db: Session, admin: m.Employee, kind: str, item_id: int):
    model = m.Project if kind == "project" else m.TaskType
    item = db.get(model, item_id)
    if item is None:
        return None
    scope = admin_department_scope(admin)
    if scope is not None:
        submitter_dept = (item.created_by.department or "—") if item.created_by else None
        if submitter_dept != scope:
            # Not just a UI filter — a department-scoped admin can't approve/
            # reject someone else's team's suggestion even via a direct POST.
            raise Forbidden()
    return item


@router.post("/suggestions/{kind}/{item_id}/approve")
def suggestion_approve(
    kind: str,
    item_id: int,
    request: Request,
    admin: m.Employee = Depends(require_admin),
    db: Session = Depends(get_db),
):
    item = _suggestion_or_forbidden(db, admin, kind, item_id)
    if item is None:
        return RedirectResponse("/admin/suggestions", status_code=303)
    item.status = m.LIST_APPROVED
    item.reviewed_by = admin.name
    item.reviewed_at = dt.datetime.utcnow()
    db.commit()
    audit(db, admin.name, f"suggestion_approve_{kind}", kind, item.name, {})
    flash(request, f"Approved '{item.name}' — now visible to everyone.", "ok")
    return RedirectResponse("/admin/suggestions", status_code=303)


@router.post("/suggestions/{kind}/{item_id}/edit")
def suggestion_edit(
    kind: str,
    item_id: int,
    request: Request,
    name: str = Form(...),
    admin: m.Employee = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Admin rewrites a still-pending suggestion's name before approving/
    rejecting it (Ganesh, 2026-08-21) — e.g. fixing a typo, matching how
    an existing client's name is normally written, or just tightening the
    wording. Deliberately doesn't run the name through
    normalize_title_case() the way the employee's own original submission
    did (see app/routes/employee.py suggest_list_item()) — an admin
    typing a specific name is trusted to have typed it the way they want
    it, not auto-corrected out from under them.

    Only ever touches a still-PENDING row (same scope as Approve/Reject —
    _suggestion_or_forbidden() below also 403s a department-scoped admin
    editing outside their own team). Editing an already-approved/rejected
    suggestion isn't offered here; the Suggestions page itself only ever
    lists pending ones, so there's nothing to click Edit on once a
    decision's been made.

    employee_notified_at is reset to NULL on every save (even a second or
    third edit of the same row) so each rewrite gets its own fresh banner
    on the employee's Today page — see app/routes/employee.py's
    _pending_edit_notices()."""
    item = _suggestion_or_forbidden(db, admin, kind, item_id)
    if item is None:
        return RedirectResponse("/admin/suggestions", status_code=303)
    if item.status != m.LIST_PENDING:
        flash(request, "That suggestion has already been decided — nothing to edit.", "err")
        return RedirectResponse("/admin/suggestions", status_code=303)
    new_name = name.strip()
    if not new_name:
        flash(request, "Enter a name.", "err")
        return RedirectResponse("/admin/suggestions", status_code=303)
    model = m.Project if kind == "project" else m.TaskType
    existing = db.execute(
        select(model).where(func.lower(model.name) == new_name.lower(), model.id != item.id)
    ).scalar_one_or_none()
    if existing is not None:
        flash(request, f"'{existing.name}' already exists — pick a different name.", "err")
        return RedirectResponse("/admin/suggestions", status_code=303)
    old_name = item.name
    if item.original_name is None:
        item.original_name = old_name  # only ever captured once — see docstring above
    item.name = new_name
    item.edited_by = admin.name
    item.edited_at = dt.datetime.utcnow()
    item.employee_notified_at = None  # re-arm the Today-page banner for this edit
    db.commit()
    audit(db, admin.name, f"suggestion_edit_{kind}", kind, item.name, {"from": old_name, "to": new_name})
    flash(request, f"'{old_name}' renamed to '{new_name}'. {item.created_by.name if item.created_by else 'The submitter'} will see this on their Today page.", "ok")
    return RedirectResponse("/admin/suggestions", status_code=303)


@router.post("/suggestions/{kind}/{item_id}/reject")
def suggestion_reject(
    kind: str,
    item_id: int,
    request: Request,
    review_note: str = Form(""),
    admin: m.Employee = Depends(require_admin),
    db: Session = Depends(get_db),
):
    item = _suggestion_or_forbidden(db, admin, kind, item_id)
    if item is None:
        return RedirectResponse("/admin/suggestions", status_code=303)
    item.status = m.LIST_REJECTED
    # also deactivate, belt-and-suspenders: LIST_REJECTED already isn't
    # LIST_APPROVED so validate_entry/_visible_projects_and_tasks in
    # app/routes/employee.py already reject/hide it either way, but active
    # is what every OTHER dropdown-visibility check in the app keys off —
    # keeps a rejected row consistent with how a deactivated one behaves
    # everywhere else, not just in the suggestion flow.
    item.active = False
    item.reviewed_by = admin.name
    item.reviewed_at = dt.datetime.utcnow()
    item.review_note = review_note.strip()
    db.commit()
    audit(db, admin.name, f"suggestion_reject_{kind}", kind, item.name, {"reason": review_note.strip()})
    flash(request, f"Rejected '{item.name}'.", "ok")
    return RedirectResponse("/admin/suggestions", status_code=303)


# --------------------------------------------------------------------------
# Project/Task assignment (Ganesh, 2026-08-01) — a team lead's "this
# employee works on this project/task" marker. Deliberately advisory, NOT
# enforced: app/validation.py never rejects a TaskEntry against an
# unassigned project — this only changes what's shown first/highlighted on
# the Today entry form (see app/routes/employee.py
# _visible_projects_and_tasks). Same require_admin + department-scope tier
# as Leave Requests/Suggestions — a team lead manages their own team only.
# --------------------------------------------------------------------------
@router.get("/assignments")
def assignments_page(
    request: Request,
    employee_id: Optional[int] = None,
    admin: m.Employee = Depends(require_admin),
    db: Session = Depends(get_db),
):
    scope = admin_department_scope(admin)
    emps = list(
        db.execute(select(m.Employee).where(m.Employee.active.is_(True)).order_by(m.Employee.name)).scalars()
    )
    if scope is not None:
        emps = [e for e in emps if (e.department or "—") == scope]

    selected = next((e for e in emps if e.id == employee_id), None) if employee_id is not None else None

    # Department -> Employee picker tree (Ganesh, 2026-08-29: "instead it
    # should show departments once admin clicked on department it should
    # show employee") — same collapsible-tree convention the Projects &
    # Tasks page already established (.tree/.tree-dept CSS, reused as-is),
    # just one level deep here since an employee is a leaf, not something
    # that itself expands further. `emps` is already scope-filtered above,
    # so a department-scoped admin only ever sees their own one group.
    emp_dept_groups: Dict[str, list] = {}
    for e in emps:
        emp_dept_groups.setdefault(e.department or "—", []).append(e)
    emp_dept_groups = [
        {"name": d, "employees": emp_dept_groups[d]} for d in sorted(emp_dept_groups.keys())
    ]

    projects = list(
        db.execute(
            select(m.Project)
            .where(m.Project.active.is_(True), m.Project.status == m.LIST_APPROVED)
            .order_by(m.Project.name)
        ).scalars()
    )
    tasks = list(
        db.execute(
            select(m.TaskType)
            .where(m.TaskType.active.is_(True), m.TaskType.status == m.LIST_APPROVED)
            .order_by(m.TaskType.name)
        ).scalars()
    )

    assigned_project_ids, assigned_task_ids = set(), set()
    plan_project_items, plan_task_items, assigned_plans = [], [], []
    if selected is not None:
        assigned_project_ids = {
            row[0] for row in db.execute(
                select(m.ProjectAssignment.project_id).where(m.ProjectAssignment.employee_id == selected.id)
            ).all()
        }
        assigned_task_ids = {
            row[0] for row in db.execute(
                select(m.TaskAssignment.task_type_id).where(m.TaskAssignment.employee_id == selected.id)
            ).all()
        }
        # "Assign a new task" card (Ganesh, 2026-08-29: "i want this assign
        # tasks feature to be in Assignments") — same TK-04 PlannedTask
        # assignment Person Detail already has (app/models.py's PlannedTask
        # docstring), just reachable from here too now. Same
        # _plan_combo_items()/query shape person() uses above, scoped to
        # whichever employee is selected in this page's own picker rather
        # than a URL path segment.
        plan_project_items, plan_task_items = _plan_combo_items(db)
        assigned_plans = list(
            db.execute(
                select(m.PlannedTask)
                .where(m.PlannedTask.employee_id == selected.id, m.PlannedTask.status == m.PLAN_PLANNED)
                .order_by(m.PlannedTask.date)
            ).scalars()
        )

    return render(
        request, "admin/assignments.html",
        {
            "user": admin, "emps": emps, "emp_dept_groups": emp_dept_groups, "selected": selected,
            "projects": projects, "tasks": tasks,
            "assigned_project_ids": assigned_project_ids, "assigned_task_ids": assigned_task_ids,
            "plan_project_items": plan_project_items, "plan_task_items": plan_task_items,
            "assigned_plans": assigned_plans, "today": today_local(),
        },
        db=db,
    )


@router.post("/assignments/{employee_id}")
def assignments_save(
    employee_id: int,
    request: Request,
    project_ids: list = Form([]),
    task_type_ids: list = Form([]),
    admin: m.Employee = Depends(require_admin),
    db: Session = Depends(get_db),
):
    emp = db.get(m.Employee, employee_id)
    if emp is None:
        return RedirectResponse("/admin/assignments", status_code=303)
    scope = admin_department_scope(admin)
    if scope is not None and (emp.department or "—") != scope:
        raise Forbidden()

    # replace-all: simplest correct way to sync a set of checkboxes without
    # tracking individual add/remove diffs
    db.execute(delete(m.ProjectAssignment).where(m.ProjectAssignment.employee_id == employee_id))
    db.execute(delete(m.TaskAssignment).where(m.TaskAssignment.employee_id == employee_id))
    for pid in project_ids:
        db.add(m.ProjectAssignment(employee_id=employee_id, project_id=int(pid), assigned_by=admin.name))
    for tid in task_type_ids:
        db.add(m.TaskAssignment(employee_id=employee_id, task_type_id=int(tid), assigned_by=admin.name))
    db.commit()
    audit(db, admin.name, "assignments_save", "Employee", employee_id,
          {"projects": len(project_ids), "tasks": len(task_type_ids)})
    flash(request, f"Saved assignments for {emp.name}.", "ok")
    return RedirectResponse(f"/admin/assignments?employee_id={employee_id}", status_code=303)


# --------------------------------------------------------------------------
# Leave entry (admin-entered in POC — open question 5)
# --------------------------------------------------------------------------
@router.get("/leave")
def leave_page(
    request: Request,
    admin: m.Employee = Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    # Department-scoped admin (team lead) — Leave Requests is one of the
    # three screens they keep (see Employee.is_super_admin docstring), but
    # only for their own team: the employee picker, pending queue and
    # approved history are all filtered to it. LeaveRecord has no
    # department column of its own, so this filters by employee_id
    # membership in the scoped roster instead.
    scope = admin_department_scope(admin)
    emps = list(
        db.execute(
            select(m.Employee).where(m.Employee.active.is_(True)).order_by(m.Employee.name)
        ).scalars()
    )
    if scope is not None:
        emps = [e for e in emps if (e.department or "—") == scope]
    scoped_ids = {e.id for e in emps} if scope is not None else None

    pending = list(
        db.execute(
            select(m.LeaveRecord)
            .where(m.LeaveRecord.status == m.LEAVE_REQUESTED)
            .order_by(m.LeaveRecord.created_at)
        ).scalars()
    )
    approved_q = (
        select(m.LeaveRecord)
        .where(m.LeaveRecord.status == m.LEAVE_APPROVED)
        .order_by(m.LeaveRecord.created_at.desc())
    )
    if scoped_ids is None:
        approved_q = approved_q.limit(100)
    approved = list(db.execute(approved_q).scalars())
    if scoped_ids is not None:
        # filter *before* trimming to 100 — otherwise a small team's older
        # approved leave could get pushed out by other departments' more
        # recent rows before the scoping filter ever sees them
        pending = [lv for lv in pending if lv.employee_id in scoped_ids]
        approved = [lv for lv in approved if lv.employee_id in scoped_ids][:100]
    if LEAVE_MANAGEMENT_V2_ENABLED:
        cfg = engine.get_config(db)
        today = today_local()
        balances_v2 = {e.id: engine.leave_balance_v2(db, e, today, cfg) for e in emps}
        # Overtime ↔ Missed Hours match requests moved to Overtime
        # Management (Ganesh, 2026-08-22 — see overtime_page()'s
        # pending_matches) since they're overtime decisions, not leave
        # ones; leave.html no longer renders that card.
        return render(
            request, "admin/leave.html",
            {
                "user": admin, "emps": emps, "pending": pending, "approved": approved,
                "leave_types": m.LEAVE_TYPES_V2, "balances_v2": balances_v2,
            },
            db=db,
        )
    balances = {e.id: engine.leave_balance(db, e) for e in emps}
    return render(
        request, "admin/leave.html",
        {
            "user": admin, "emps": emps, "pending": pending, "approved": approved,
            "leave_types": m.LEAVE_TYPES, "balances": balances, "balance_year": today_local().year,
        },
        db=db,
    )


# --------------------------------------------------------------------------
# Bulk leave-allocation assignment (Leave -> Bulk assign leaves) — parsing
# rules live in app/leave_bulk_upload.py
# --------------------------------------------------------------------------
@router.get("/leave/bulk-upload")
def leave_bulk_upload_page(
    request: Request,
    admin: m.Employee = Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    return render(request, "admin/leave_bulk_upload.html", {"user": admin, "result": None}, db=db)


@router.get("/leave/bulk-upload/sample.xlsx")
def leave_bulk_upload_sample(admin: m.Employee = Depends(require_super_admin)):
    buf = io.BytesIO()
    leave_bulk_upload.build_sample_workbook().save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="leave_allocation_template.xlsx"'},
    )


@router.get("/leave/bulk-upload/existing.xlsx")
def leave_bulk_upload_existing(
    admin: m.Employee = Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    buf = io.BytesIO()
    leave_bulk_upload.build_existing_allocations_workbook(db).save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="existing_leave_allocations.xlsx"'},
    )


@router.post("/leave/bulk-upload")
def leave_bulk_upload_post(
    request: Request,
    file: UploadFile = File(...),
    admin: m.Employee = Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xlsm")):
        flash(request, "Please upload an .xlsx file — use the sample template.", "err")
        return RedirectResponse("/admin/leave/bulk-upload", status_code=303)
    try:
        wb = load_workbook(io.BytesIO(file.file.read()), data_only=True)
    except Exception:
        flash(request, "Couldn't read that file — is it a valid, unprotected .xlsx?", "err")
        return RedirectResponse("/admin/leave/bulk-upload", status_code=303)

    result = leave_bulk_upload.process_upload(db, wb)
    if result["header_error"]:
        flash(request, result["header_error"], "err")
        return RedirectResponse("/admin/leave/bulk-upload", status_code=303)
    if result["updated"]:
        audit(db, admin.name, "leave_bulk_upload", "Employee", "",
              {"updated": result["updated"], "skipped": len(result["skipped"])})
    summary = f"{result['updated']} employee(s) had leave allocations updated."
    if result["skipped"]:
        summary += f" {len(result['skipped'])} row(s) skipped — see details below."
    flash(request, summary, "ok" if result["updated"] else "err")
    return render(request, "admin/leave_bulk_upload.html", {"user": admin, "result": result}, db=db)


@router.post("/leave/{leave_id}/approve")
def leave_approve(
    leave_id: int,
    request: Request,
    review_note: str = Form(""),
    approved_hours: str = Form(""),  # V2 only — requirement 6, partial approval
    admin: m.Employee = Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    lv = db.get(m.LeaveRecord, leave_id)
    if lv is None:
        return RedirectResponse("/admin/leave", status_code=303)
    emp = db.get(m.Employee, lv.employee_id)
    scope = admin_department_scope(admin)
    if scope is not None and emp is not None and (emp.department or "—") != scope:
        # Not just a UI filter — block a department-scoped admin from
        # approving/rejecting leave for someone outside their team even if
        # they craft the POST directly.
        raise Forbidden()
    review_note = review_note.strip()
    approved_minutes = None
    if LEAVE_MANAGEMENT_V2_ENABLED:
        requested_minutes = lv.minutes_per_day if lv.minutes_per_day is not None else (
            emp.daily_target_minutes if emp is not None else None
        )
        if approved_hours.strip():
            try:
                approved_minutes = int(round(float(approved_hours) * 60))
            except ValueError:
                flash(request, "Hours approved must be a number.", "err")
                return RedirectResponse("/admin/leave", status_code=303)
            if approved_minutes < 0:
                flash(request, "Hours approved can't be negative.", "err")
                return RedirectResponse("/admin/leave", status_code=303)
        else:
            approved_minutes = requested_minutes
        # Requirement 6: a note explaining why is strongly encouraged
        # whenever the approved amount is less than what was requested —
        # required, not just encouraged, so the "why partial" reason is
        # never silently missing from the record.
        if requested_minutes is not None and approved_minutes is not None and approved_minutes < requested_minutes and not review_note:
            flash(request, "Approving fewer hours than requested needs a note explaining why.", "err")
            return RedirectResponse("/admin/leave", status_code=303)
        lv.approved_minutes_per_day = approved_minutes
    lv.status = m.LEAVE_APPROVED
    lv.reviewed_by = admin.name
    lv.reviewed_at = dt.datetime.utcnow()
    lv.review_note = review_note
    db.commit()
    audit(db, admin.name, "leave_approve", "LeaveRecord", lv.id,
          {"employee": emp.name if emp else lv.employee_id, "range": f"{lv.start_date}..{lv.end_date}",
           "approved_minutes_per_day": approved_minutes})
    if emp is not None:
        engine.recompute_employee(db, emp, lv.start_date, min(lv.end_date, today_local()))
    flash(request, f"Approved leave for {emp.name if emp else lv.employee_id}.", "ok")
    return RedirectResponse("/admin/leave", status_code=303)


@router.post("/leave/{leave_id}/reject")
def leave_reject(
    leave_id: int,
    request: Request,
    review_note: str = Form(""),
    admin: m.Employee = Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    lv = db.get(m.LeaveRecord, leave_id)
    if lv is None:
        return RedirectResponse("/admin/leave", status_code=303)
    emp = db.get(m.Employee, lv.employee_id)
    scope = admin_department_scope(admin)
    if scope is not None and emp is not None and (emp.department or "—") != scope:
        raise Forbidden()
    lv.status = m.LEAVE_REJECTED
    lv.reviewed_by = admin.name
    lv.reviewed_at = dt.datetime.utcnow()
    lv.review_note = review_note.strip()
    db.commit()
    audit(db, admin.name, "leave_reject", "LeaveRecord", lv.id,
          {"employee": emp.name if emp else lv.employee_id, "reason": review_note.strip()})
    if emp is not None:
        engine.recompute_employee(db, emp, lv.start_date, min(lv.end_date, today_local()))
    flash(request, f"Rejected leave request for {emp.name if emp else lv.employee_id}.", "ok")
    return RedirectResponse("/admin/leave", status_code=303)


@router.post("/leave/add")
def leave_add(
    request: Request,
    employee_id: int = Form(...),
    start_date: str = Form(...),
    end_date: str = Form(""),
    type: str = Form("Other"),
    hours: str = Form(""),
    note: str = Form(""),
    relation: str = Form(""),  # V2 only
    admin: m.Employee = Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    emp = db.get(m.Employee, employee_id)
    if emp is None:
        return RedirectResponse("/admin/leave", status_code=303)
    scope = admin_department_scope(admin)
    if scope is not None and (emp.department or "—") != scope:
        raise Forbidden()
    try:
        start = parse_date_field(start_date, "Start date")
        end = parse_date_field(end_date, "End date") if end_date else start
    except FormError as e:
        flash(request, e.message, "err")
        return RedirectResponse("/admin/leave", status_code=303)
    if end < start:
        flash(request, "End date is before start date.", "err")
        return RedirectResponse("/admin/leave", status_code=303)
    minutes = None  # full day = person's daily target (PRD §5)
    if hours.strip():
        try:
            minutes = int(round(float(hours) * 60))
        except ValueError:
            flash(request, "Hours must be a number (leave blank for full day).", "err")
            return RedirectResponse("/admin/leave", status_code=303)

    relation = relation.strip()
    if LEAVE_MANAGEMENT_V2_ENABLED:
        if type not in m.LEAVE_TYPES_V2:
            flash(request, "Choose a valid leave type.", "err")
            return RedirectResponse("/admin/leave", status_code=303)
        if type == m.LEAVE_SPECIAL_PAID:
            flash(request, "Special Paid Time is granted via 'Grant Special Paid Time', not recorded here.", "err")
            return RedirectResponse("/admin/leave", status_code=303)
        if type != m.LEAVE_BEREAVEMENT:
            relation = ""
    else:
        relation = ""
        if type == "Other" and not note.strip():
            flash(request, "'Other' leave needs a note (PRD §5).", "err")
            return RedirectResponse("/admin/leave", status_code=303)

    lv = m.LeaveRecord(
        employee_id=employee_id, start_date=start, end_date=end,
        type=type, minutes_per_day=minutes, note=note.strip(), entered_by=admin.name,
        relation=relation or None,
        status=m.LEAVE_APPROVED,  # admin direct-entry is already-approved by definition
        approved_minutes_per_day=minutes,  # already-approved: what was entered IS what's approved
    )
    db.add(lv)
    db.commit()
    audit(db, admin.name, "leave_add", "LeaveRecord", lv.id,
          {"employee": emp.name, "range": f"{start}..{end}", "type": type, "minutes": minutes})
    engine.recompute_employee(db, emp, start, min(end, today_local()))
    flash(request, f"Leave recorded for {emp.name}.", "ok")
    return RedirectResponse("/admin/leave", status_code=303)


@router.post("/leave/{leave_id}/delete")
def leave_delete(
    leave_id: int,
    request: Request,
    admin: m.Employee = Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    lv = db.get(m.LeaveRecord, leave_id)
    if lv is not None:
        emp = db.get(m.Employee, lv.employee_id)
        scope = admin_department_scope(admin)
        if scope is not None and emp is not None and (emp.department or "—") != scope:
            raise Forbidden()
        db.delete(lv)
        db.commit()
        audit(db, admin.name, "leave_delete", "LeaveRecord", leave_id,
              {"employee": emp.name if emp else lv.employee_id})
        if emp is not None:
            engine.recompute_employee(db, emp, lv.start_date, min(lv.end_date, today_local()))
    return RedirectResponse("/admin/leave", status_code=303)


# --------------------------------------------------------------------------
# Special Paid Time — management grant (requirement 8, Leave Management V2,
# 2026-08-21). A ledger (SpecialPaidGrant), not an approval queue: granting
# the hours here IS the approval, same one-action-reason-required shape
# Compensation Links already use (see docs/LEAVE_MANAGEMENT_PLAN.md §3).
# SuperAdmin-only, same tier as Roster/bulk-upload/Settings — this hands
# out real paid hours, not something a department-scoped admin should be
# able to do for their own team unchecked.
# --------------------------------------------------------------------------
@router.get("/leave/special-paid")
def special_paid_page(
    request: Request,
    admin: m.Employee = Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    if not LEAVE_MANAGEMENT_V2_ENABLED:
        raise HTTPException(status_code=404)
    emps = list(
        db.execute(
            select(m.Employee).where(m.Employee.active.is_(True)).order_by(m.Employee.name)
        ).scalars()
    )
    grants = list(
        db.execute(
            select(m.SpecialPaidGrant).order_by(m.SpecialPaidGrant.granted_at.desc()).limit(200)
        ).scalars()
    )
    return render(
        request, "admin/leave_special_paid.html",
        {"user": admin, "emps": emps, "grants": grants}, db=db,
    )


@router.post("/leave/special-paid")
def special_paid_grant(
    request: Request,
    employee_id: int = Form(...),
    hours: str = Form(...),
    reason: str = Form(...),
    admin: m.Employee = Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    if not LEAVE_MANAGEMENT_V2_ENABLED:
        raise HTTPException(status_code=404)
    emp = db.get(m.Employee, employee_id)
    if emp is None:
        flash(request, "Choose a valid employee.", "err")
        return RedirectResponse("/admin/leave/special-paid", status_code=303)
    if not reason.strip():
        flash(request, "A reason is required.", "err")
        return RedirectResponse("/admin/leave/special-paid", status_code=303)
    try:
        minutes = int(round(float(hours) * 60))
    except ValueError:
        flash(request, "Hours must be a number.", "err")
        return RedirectResponse("/admin/leave/special-paid", status_code=303)
    if minutes <= 0:
        flash(request, "Hours must be greater than zero.", "err")
        return RedirectResponse("/admin/leave/special-paid", status_code=303)
    grant = m.SpecialPaidGrant(
        employee_id=employee_id, minutes=minutes, reason=reason.strip(), granted_by=admin.name,
    )
    db.add(grant)
    db.commit()
    audit(db, admin.name, "special_paid_grant", "SpecialPaidGrant", grant.id,
          {"employee": emp.name, "minutes": minutes, "reason": reason.strip()})
    flash(request, f"Granted {emp.name} {hours} hour(s) of Special Paid Time.", "ok")
    return RedirectResponse("/admin/leave/special-paid", status_code=303)


# --------------------------------------------------------------------------
# Overtime Requests (Ganesh's manager, 2026-08-03) — same submit -> queue ->
# act shape as Leave Requests just above, but scoped per-person via
# app.auth.led_by() (Team Lead = whoever an employee's reports_to is, IF
# that person is_admin) instead of by department. Approving/rejecting
# doesn't touch engine.py/DayStatus/strikes at all — see OvertimeApproval's
# model docstring — so nothing here needs an engine.recompute_employee call
# the way Leave's approve/reject/add/delete do above.
# --------------------------------------------------------------------------
@router.get("/overtime")
def overtime_page(
    request: Request,
    ym: Optional[str] = None,
    employee_id: Optional[int] = None,
    admin: m.Employee = Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    scope = led_by(admin, db)  # None => Super Admin, sees/can act on everyone
    emps = list(
        db.execute(
            select(m.Employee).where(m.Employee.active.is_(True)).order_by(m.Employee.name)
        ).scalars()
    )
    if scope is not None:
        emps = [e for e in emps if e.id in scope]

    pending = list(
        db.execute(
            select(m.OvertimeApproval)
            .where(m.OvertimeApproval.status == m.OT_REQUESTED)
            .order_by(m.OvertimeApproval.created_at)
        ).scalars()
    )
    approved_q = (
        select(m.OvertimeApproval)
        .where(m.OvertimeApproval.status == m.OT_APPROVED)
        .order_by(m.OvertimeApproval.created_at.desc())
    )
    if scope is None:
        approved_q = approved_q.limit(100)
    approved = list(db.execute(approved_q).scalars())
    if scope is not None:
        # filter *before* trimming to 100 — same reasoning as Leave's
        # leave_page above: a small team's older approvals could otherwise
        # get pushed out by other leads' more recent rows first.
        pending = [ot for ot in pending if ot.employee_id in scope]
        approved = [ot for ot in approved if ot.employee_id in scope][:100]

    # Compensation links (Ganesh, 2026-08-21) — a quick org/team-scoped view
    # of the same CompensationLink rows Person Detail already shows one
    # employee at a time (see person() above), so a lead doesn't have to
    # open each report separately to see who's been linked. The "pick an
    # employee, see their shortfall/surplus days, link them" picker below
    # (link_target/link_shortfalls/link_surpluses) reuses the exact same
    # POST /admin/person/{id}/complink endpoint Person Detail's own form
    # posts to, just with return_to=overtime so it redirects back here
    # instead (see _complink_redirect() below) — so creating a link no
    # longer requires leaving Overtime Management. This list itself stays
    # read-only, though: removing a link is still Person Detail-only, since
    # that's also where its knock-on effect on that day's status is shown.
    cfg = engine.get_config(db)
    links_q = select(m.CompensationLink).order_by(m.CompensationLink.shortfall_date.desc())
    if scope is None:
        links_q = links_q.limit(150)
    comp_link_rows = list(db.execute(links_q).scalars())
    if scope is not None:
        comp_link_rows = [lk for lk in comp_link_rows if lk.employee_id in scope][:150]
    # same (link, parsed-surplus-dates) shape person() builds above, so the
    # template can reuse identical `{% for s in sdates %}{{ s|mdy }}` markup
    comp_links = [
        (lk, [dt.date.fromisoformat(x) for x in json.loads(lk.surplus_dates or "[]")])
        for lk in comp_link_rows
    ]

    # Overtime actually worked vs. approved, for one month at a time
    # (Ganesh, 2026-08-21) — reuses the exact same aggregation Reports ->
    # Attendance already computes (reports.attendance_report's "summary"
    # mode: overtime_minutes from completed Punch In/Out vs. approved_
    # overtime_minutes, the portion covered by an OT_APPROVED date range
    # above). This is a read-only at-a-glance table; Reports -> Attendance
    # is still the place to filter by department/employee/date range.
    year, month = parse_ym(ym)
    first, last = engine.month_range(year, month)
    today = today_local()
    # "Who worked overtime" (Ganesh, 2026-08-25: "overtime from punch in
    # punched out time is not requere... we are considering only based on
    # task log times") — switched from reports.attendance_report()'s
    # Punch-In/Out-based overtime figure to reports.task_log_overtime_
    # report(), which sums each employee's positive DayStatus.variance_
    # minutes instead (the same "surplus" figure the Compensation Links
    # picker above already uses) so the two sections on this page agree on
    # what "overtime" means. Already scoped via `emps` (led_by()), so no
    # separate scope filter needed here unlike the old attendance_report()
    # path.
    overtime_rows = reports.task_log_overtime_report(db, first, min(last, today), [e.id for e in emps])
    (py, pm), (ny, nm) = prev_next_month(year, month)

    # Employee picker for the Compensation links quick-link flow above
    # (Ganesh, 2026-08-21) — only offers employees already in `emps` (this
    # admin's own led_by() scope, or everyone for a Super Admin), same as
    # the rest of this page. Shortfall/surplus candidates are this same
    # employee's for whichever month is currently being viewed (`ym`),
    # computed the identical way person() computes them — see
    # _shortfalls_surpluses() above.
    link_target = next((e for e in emps if e.id == employee_id), None) if employee_id else None
    # Overtime ↔ Missed Hours match requests (Ganesh, 2026-08-22 — moved
    # here from Leave Management, since these are overtime decisions, not
    # leave ones; the underlying data/routes are unchanged). Super-admin
    # only: approve_complink/reject_complink are both require_super_admin
    # (unlike the OvertimeApproval pending/approved above, which a
    # department-scoped Team Lead — `scope` non-None here — can act on via
    # led_by()), so a Lead would just get a 403 clicking Approve/Reject on
    # one of these. Not shown to them at all rather than shown-but-broken.
    pending_matches = []
    if LEAVE_MANAGEMENT_V2_ENABLED and scope is None:
        pending_matches = list(
            db.execute(
                select(m.CompensationLink)
                .where(
                    m.CompensationLink.status == m.LEAVE_REQUESTED,
                    m.CompensationLink.requested_by_employee.is_(True),
                )
                .order_by(m.CompensationLink.id)
            ).scalars()
        )
        pending_matches = [
            (c, [dt.date.fromisoformat(x) for x in json.loads(c.surplus_dates or "[]")])
            for c in pending_matches
        ]

    link_shortfalls, link_surpluses = [], []
    link_shortfall_allocated_by_date: Dict[dt.date, int] = {}
    link_surplus_used_by_date: Dict[dt.date, int] = {}
    if link_target is not None:
        if first <= today:
            engine.recompute_employee(db, link_target, first, min(last, today), cfg)
        target_statuses = list(
            db.execute(
                select(m.DayStatus)
                .where(m.DayStatus.employee_id == link_target.id, m.DayStatus.date.between(first, last))
            ).scalars()
        )
        link_shortfalls, link_surpluses = _shortfalls_surpluses(target_statuses, cfg.get("comp_erases_strike") == "1")
        link_shortfall_allocated_by_date = engine.shortfall_allocated_minutes_by_date(db, link_target.id)
        link_surplus_used_by_date = engine.surplus_minutes_used_by_date(db, link_target.id)
    # Whether at least one shortfall day still has any deficit left to link
    # — used to disable the Link button below the same way `link_shortfalls`
    # alone used to, but now correctly ignores a day that's shown-but-
    # disabled in the dropdown because it's already fully linked (partial
    # allocation, Ganesh, 2026-08-25 — see link_shortfall_allocated_by_date).
    link_has_unlinked_shortfall = any(
        -(r.variance_minutes or 0) > link_shortfall_allocated_by_date.get(r.date, 0)
        for r in link_shortfalls
    )

    return render(
        request, "admin/overtime.html",
        {
            "user": admin, "emps": emps, "pending": pending, "approved": approved,
            "comp_links": comp_links, "overtime_rows": overtime_rows, "pending_matches": pending_matches,
            "link_target": link_target, "link_shortfalls": link_shortfalls, "link_surpluses": link_surpluses,
            "link_shortfall_allocated_by_date": link_shortfall_allocated_by_date,
            "link_surplus_used_by_date": link_surplus_used_by_date,
            "link_has_unlinked_shortfall": link_has_unlinked_shortfall,
            "year": year, "month": month, "ym": f"{year}-{month:02d}",
            "prev_ym": f"{py}-{pm:02d}", "next_ym": f"{ny}-{nm:02d}",
        },
        db=db,
    )


def _ot_forbidden_unless_scoped(admin: m.Employee, ot: m.OvertimeApproval, db: Session) -> None:
    scope = led_by(admin, db)
    if scope is not None and ot.employee_id not in scope:
        # Not just a UI filter — block a Lead from approving/rejecting
        # overtime for someone who isn't their direct report even if they
        # craft the POST directly, same defense-in-depth as Leave's checks.
        raise Forbidden()


@router.post("/overtime/{ot_id}/approve")
def overtime_approve(
    ot_id: int,
    request: Request,
    review_note: str = Form(""),
    admin: m.Employee = Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    ot = db.get(m.OvertimeApproval, ot_id)
    if ot is None:
        return RedirectResponse("/admin/overtime", status_code=303)
    _ot_forbidden_unless_scoped(admin, ot, db)
    emp = db.get(m.Employee, ot.employee_id)
    ot.status = m.OT_APPROVED
    ot.reviewed_by = admin.name
    ot.reviewed_at = dt.datetime.utcnow()
    ot.review_note = review_note.strip()
    db.commit()
    audit(db, admin.name, "overtime_approve", "OvertimeApproval", ot.id,
          {"employee": emp.name if emp else ot.employee_id, "range": f"{ot.start_date}..{ot.end_date}"})
    flash(request, f"Approved overtime for {emp.name if emp else ot.employee_id}.", "ok")
    return RedirectResponse("/admin/overtime", status_code=303)


@router.post("/overtime/{ot_id}/reject")
def overtime_reject(
    ot_id: int,
    request: Request,
    review_note: str = Form(""),
    admin: m.Employee = Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    ot = db.get(m.OvertimeApproval, ot_id)
    if ot is None:
        return RedirectResponse("/admin/overtime", status_code=303)
    _ot_forbidden_unless_scoped(admin, ot, db)
    emp = db.get(m.Employee, ot.employee_id)
    ot.status = m.OT_REJECTED
    ot.reviewed_by = admin.name
    ot.reviewed_at = dt.datetime.utcnow()
    ot.review_note = review_note.strip()
    db.commit()
    audit(db, admin.name, "overtime_reject", "OvertimeApproval", ot.id,
          {"employee": emp.name if emp else ot.employee_id, "reason": review_note.strip()})
    flash(request, f"Rejected overtime request for {emp.name if emp else ot.employee_id}.", "ok")
    return RedirectResponse("/admin/overtime", status_code=303)


@router.post("/overtime/grant")
def overtime_grant(
    request: Request,
    employee_id: int = Form(...),
    start_date: str = Form(...),
    end_date: str = Form(""),
    note: str = Form(""),
    admin: m.Employee = Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    """A Lead/Admin approving overtime *without* a preceding employee
    request — covers both granting it proactively ahead of a known busy
    period, and reviewing after the fact (start/end can be in the past;
    nothing here cares which direction time runs, see model docstring)."""
    emp = db.get(m.Employee, employee_id)
    if emp is None:
        return RedirectResponse("/admin/overtime", status_code=303)
    scope = led_by(admin, db)
    if scope is not None and employee_id not in scope:
        raise Forbidden()
    try:
        start = parse_date_field(start_date, "Start date")
        end = parse_date_field(end_date, "End date") if end_date else start
    except FormError as e:
        flash(request, e.message, "err")
        return RedirectResponse("/admin/overtime", status_code=303)
    if end < start:
        flash(request, "End date is before start date.", "err")
        return RedirectResponse("/admin/overtime", status_code=303)
    ot = m.OvertimeApproval(
        employee_id=employee_id, start_date=start, end_date=end, note=note.strip(),
        requested_by=admin.name,
        status=m.OT_APPROVED,  # lead/admin direct-entry is already-approved by definition
        reviewed_by=admin.name, reviewed_at=dt.datetime.utcnow(),
    )
    db.add(ot)
    db.commit()
    audit(db, admin.name, "overtime_grant", "OvertimeApproval", ot.id,
          {"employee": emp.name, "range": f"{start}..{end}"})
    flash(request, f"Overtime approved for {emp.name}.", "ok")
    return RedirectResponse("/admin/overtime", status_code=303)


@router.post("/overtime/{ot_id}/delete")
def overtime_delete(
    ot_id: int,
    request: Request,
    admin: m.Employee = Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    ot = db.get(m.OvertimeApproval, ot_id)
    if ot is not None:
        _ot_forbidden_unless_scoped(admin, ot, db)
        emp = db.get(m.Employee, ot.employee_id)
        db.delete(ot)
        db.commit()
        audit(db, admin.name, "overtime_delete", "OvertimeApproval", ot_id,
              {"employee": emp.name if emp else ot.employee_id})
    return RedirectResponse("/admin/overtime", status_code=303)


# --------------------------------------------------------------------------
# Config (PRD §10 — every open question is a dial here) + holidays
# --------------------------------------------------------------------------
@router.get("/config")
def config_page(
    request: Request,
    admin: m.Employee = Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    cfg = engine.get_config(db)
    return render(request, "admin/config.html", {"user": admin, "cfg": cfg}, db=db)


@router.post("/config")
def config_save(
    request: Request,
    tolerance_hours: str = Form("1.0"),
    strike_threshold: str = Form("5"),
    max_row_hours: str = Form("4"),
    backdate_working_days: str = Form("1"),
    gap_flag_minutes: str = Form("15"),
    min_details_chars: str = Form("5"),
    max_break_minutes: str = Form("30"),
    comp_erases_strike: str = Form(""),
    live_start_date: str = Form(""),
    admin: m.Employee = Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    before = engine.get_config(db)
    try:
        values = {
            "tolerance_minutes": str(parse_hours_field(tolerance_hours, "Tolerance")),
            "strike_threshold": str(parse_int_field(strike_threshold, "Strike threshold")),
            "max_row_minutes": str(parse_hours_field(max_row_hours, "Max row length")),
            "backdate_working_days": str(parse_int_field(backdate_working_days, "Backdate window")),
            "gap_flag_minutes": str(parse_int_field(gap_flag_minutes, "Gap flag minutes")),
            "min_details_chars": str(parse_int_field(min_details_chars, "Minimum details length")),
            "max_break_minutes": str(parse_int_field(max_break_minutes, "Break allowance")),
            "comp_erases_strike": "1" if comp_erases_strike == "1" else "0",
            "live_start_date": live_start_date.strip(),
        }
    except FormError as e:
        flash(request, e.message, "err")
        return RedirectResponse("/admin/config", status_code=303)
    for k, v in values.items():
        row = db.get(m.Config, k)
        if row is None:
            db.add(m.Config(key=k, value=v))
        else:
            row.value = v
    db.commit()
    audit(db, admin.name, "config_change", "Config", "",
          {"before": {k: before.get(k) for k in values}, "after": values})
    flash(request, "Config saved. Recompute affected months to apply retroactively.", "ok")
    return RedirectResponse("/admin/config", status_code=303)


# --------------------------------------------------------------------------
# Holiday Management (Ganesh, 2026-08-12, moved out of the general Config
# page into its own screen; reverted to one shared company-wide list on
# 2026-08-14 after briefly splitting by country — see Holiday's docstring
# in app/models.py for why the location column still exists on the model
# even though nothing here reads or collects it anymore). Bulk-upload shape
# in app/holiday_bulk_upload.py, same idea as leave_bulk_upload.py.
# --------------------------------------------------------------------------
@router.get("/holidays")
def holidays_page(
    request: Request,
    admin: m.Employee = Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    if not HOLIDAY_MANAGEMENT_ENABLED:
        raise HTTPException(status_code=404)
    holidays = list(db.execute(select(m.Holiday).order_by(m.Holiday.date)).scalars())
    return render(
        request, "admin/holidays.html", {"user": admin, "holidays": holidays}, db=db,
    )


@router.post("/holidays/add")
def holiday_add(
    request: Request,
    date: str = Form(...),
    name: str = Form(""),
    admin: m.Employee = Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    if not HOLIDAY_MANAGEMENT_ENABLED:
        raise HTTPException(status_code=404)
    try:
        d = parse_date_field(date)
    except FormError as e:
        flash(request, e.message, "err")
        return RedirectResponse("/admin/holidays", status_code=303)
    exists = db.execute(select(m.Holiday).where(m.Holiday.date == d)).scalar_one_or_none()
    if exists is None:
        db.add(m.Holiday(date=d, name=name.strip(), location=m.DEFAULT_LOCATION))
        db.commit()
        audit(db, admin.name, "holiday_add", "Holiday", date, {"name": name.strip()})
    else:
        flash(request, "There's already a holiday on that date.", "err")
    return RedirectResponse("/admin/holidays", status_code=303)


@router.post("/holidays/{holiday_id}/delete")
def holiday_delete(
    holiday_id: int,
    request: Request,
    admin: m.Employee = Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    if not HOLIDAY_MANAGEMENT_ENABLED:
        raise HTTPException(status_code=404)
    h = db.get(m.Holiday, holiday_id)
    if h is not None:
        db.delete(h)
        db.commit()
        audit(db, admin.name, "holiday_delete", "Holiday", h.date.isoformat(), {"name": h.name})
    return RedirectResponse("/admin/holidays", status_code=303)


@router.get("/holidays/bulk-upload")
def holiday_bulk_upload_page(
    request: Request,
    admin: m.Employee = Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    if not HOLIDAY_MANAGEMENT_ENABLED:
        raise HTTPException(status_code=404)
    return render(request, "admin/holiday_bulk_upload.html", {"user": admin, "result": None}, db=db)


@router.get("/holidays/bulk-upload/sample.xlsx")
def holiday_bulk_upload_sample(admin: m.Employee = Depends(require_super_admin)):
    if not HOLIDAY_MANAGEMENT_ENABLED:
        raise HTTPException(status_code=404)
    buf = io.BytesIO()
    holiday_bulk_upload.build_sample_workbook().save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="holiday_template.xlsx"'},
    )


@router.get("/holidays/bulk-upload/existing.xlsx")
def holiday_bulk_upload_existing(
    admin: m.Employee = Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    if not HOLIDAY_MANAGEMENT_ENABLED:
        raise HTTPException(status_code=404)
    buf = io.BytesIO()
    holiday_bulk_upload.build_existing_holidays_workbook(db).save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="existing_holidays.xlsx"'},
    )


@router.post("/holidays/bulk-upload")
def holiday_bulk_upload_post(
    request: Request,
    file: UploadFile = File(...),
    admin: m.Employee = Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    if not HOLIDAY_MANAGEMENT_ENABLED:
        raise HTTPException(status_code=404)
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xlsm")):
        flash(request, "Please upload an .xlsx file — use the sample template.", "err")
        return RedirectResponse("/admin/holidays/bulk-upload", status_code=303)
    try:
        wb = load_workbook(io.BytesIO(file.file.read()), data_only=True)
    except Exception:
        flash(request, "Couldn't read that file — is it a valid, unprotected .xlsx?", "err")
        return RedirectResponse("/admin/holidays/bulk-upload", status_code=303)

    result = holiday_bulk_upload.process_upload(db, wb)
    if result["header_error"]:
        flash(request, result["header_error"], "err")
        return RedirectResponse("/admin/holidays/bulk-upload", status_code=303)
    if result["added"] or result["updated"]:
        audit(db, admin.name, "holiday_bulk_upload", "Holiday", "",
              {"added": result["added"], "updated": result["updated"], "skipped": len(result["skipped"])})
    summary = f"{result['added']} holiday(s) added, {result['updated']} updated."
    if result["skipped"]:
        summary += f" {len(result['skipped'])} row(s) skipped — see details below."
    flash(request, summary, "ok" if (result["added"] or result["updated"]) else "err")
    return render(request, "admin/holiday_bulk_upload.html", {"user": admin, "result": result}, db=db)


# --------------------------------------------------------------------------
# Support inbox (employee-submitted questions from /support)
# --------------------------------------------------------------------------
@router.get("/support")
def support_inbox(
    request: Request,
    status: str = "open",
    admin: m.Employee = Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    q = select(m.SupportQuery).order_by(m.SupportQuery.created_at.desc())
    rows = list(db.execute(q).scalars())
    if status == "open":
        rows = [r for r in rows if r.status == m.SUPPORT_OPEN]
    return render(
        request, "admin/support.html",
        {"user": admin, "rows": rows, "status": status}, db=db,
    )


@router.post("/support/{query_id}/resolve")
def support_resolve(
    query_id: int,
    request: Request,
    reply: str = Form(""),
    admin: m.Employee = Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    q = db.get(m.SupportQuery, query_id)
    if q is None:
        return RedirectResponse("/admin/support", status_code=303)
    q.status = m.SUPPORT_RESOLVED
    q.admin_reply = reply.strip()
    q.resolved_by = admin.name
    q.resolved_at = dt.datetime.utcnow()
    db.commit()
    audit(db, admin.name, "support_resolved", "SupportQuery", q.id, {"reply": reply.strip()[:200]})
    flash(request, "Marked resolved.", "ok")
    return RedirectResponse("/admin/support", status_code=303)


# --------------------------------------------------------------------------
# Audit log
# --------------------------------------------------------------------------
@router.get("/audit")
def audit_page(
    request: Request,
    admin: m.Employee = Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    rows = list(
        db.execute(select(m.AuditLog).order_by(m.AuditLog.at.desc()).limit(300)).scalars()
    )
    return render(request, "admin/audit.html", {"user": admin, "rows": rows}, db=db)
