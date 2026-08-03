"""Exports (PRD §7.7): the bridge to the third-party pilot and the safety
net if the POC dies. Dashboard export deliberately mirrors the legacy sheet
layout (Name / DEPARTMENT / DESIGNATION / day columns / STRIKES FOR MONTH)."""
import csv
import datetime as dt
import io
from typing import Optional

from fastapi import APIRouter, Depends
from fastapi.responses import RedirectResponse, StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import select
from sqlalchemy.orm import Session

from app import engine, models as m
from app.auth import Forbidden, admin_department_scope, require_admin, require_super_admin
from app.db import get_db
from app.util import STATUS_LABELS, fmt_date, fmt_hm, parse_ym

router = APIRouter(prefix="/export")

FILLS = {
    m.COMPLETE: PatternFill("solid", start_color="C6EFCE"),
    m.PARTIAL: PatternFill("solid", start_color="FFEB9C"),
    m.MISSING: PatternFill("solid", start_color="FFC7CE"),
    m.LEAVE: PatternFill("solid", start_color="BDD7EE"),
    m.HOLIDAY: PatternFill("solid", start_color="D9D9D9"),
}


def _xlsx_response(wb: Workbook, filename: str) -> StreamingResponse:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/dashboard.xlsx")
def dashboard_xlsx(
    ym: Optional[str] = None,
    admin: m.Employee = Depends(require_admin),
    db: Session = Depends(get_db),
):
    cfg = engine.get_config(db)
    comp_erases = cfg.get("comp_erases_strike") == "1"
    year, month = parse_ym(ym)
    first, last = engine.month_range(year, month)
    days = [first + dt.timedelta(days=i) for i in range((last - first).days + 1)]

    wb = Workbook()
    ws = wb.active
    ws.title = f"{first.strftime('%B')}_{year}"
    header = ["Name", "DEPARTMENT", "DESIGNATION"] + [d.day for d in days] + ["STRIKES FOR MONTH"]
    ws.append(header)
    for c in ws[1]:
        c.font = Font(bold=True)

    emps = list(
        db.execute(
            select(m.Employee)
            .where(m.Employee.active.is_(True), m.Employee.tracked.is_(True))
            .order_by(m.Employee.department, m.Employee.name)
        ).scalars()
    )
    scope = admin_department_scope(admin)
    if scope is not None:
        # Department-scoped admin — this export mirrors the Dashboard grid,
        # which is already restricted to their own department; the XLSX
        # must match, not quietly hand back every department.
        emps = [e for e in emps if (e.department or "—") == scope]
    by_emp = engine.statuses_for_month(db, year, month)
    last_dept = None
    for e in emps:
        rows = by_emp.get(e.id, {})
        strikes = engine.strikes_in(rows.values(), comp_erases)
        cells = []
        for d in days:
            r = rows.get(d)
            if r is None:
                cells.append("")
                continue
            eff = r.effective_status(comp_erases)
            label = STATUS_LABELS.get(eff, eff)
            if eff in (m.WEEKEND, m.HOLIDAY) and (r.actual_minutes or 0) > 0:
                label = fmt_hm(r.actual_minutes)  # weekend/holiday extra hours, like the sheet
            if r.override_status:
                label = f"{label}*"  # overridden — see audit log
            cells.append(label)
        dept = e.department if e.department != last_dept else ""
        ws.append([e.name, dept, e.designation] + cells + [strikes])
        last_dept = e.department
        row_idx = ws.max_row
        for j, d in enumerate(days):
            r = rows.get(d)
            if r is None:
                continue
            fill = FILLS.get(r.effective_status(comp_erases))
            if fill is not None:
                ws.cell(row=row_idx, column=4 + j).fill = fill
    ws.freeze_panes = "D2"
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    return _xlsx_response(wb, f"compliance_{year}-{month:02d}.xlsx")


@router.get("/person/{emp_id}.xlsx")
def person_xlsx(
    emp_id: int,
    ym: Optional[str] = None,
    admin: m.Employee = Depends(require_admin),
    db: Session = Depends(get_db),
):
    emp = db.get(m.Employee, emp_id)
    if emp is None:
        return RedirectResponse("/admin/roster", status_code=303)
    scope = admin_department_scope(admin)
    if scope is not None and (emp.department or "—") != scope:
        raise Forbidden()
    year, month = parse_ym(ym)
    first, last = engine.month_range(year, month)

    wb = Workbook()
    ws = wb.active
    ws.title = "Ledger"
    ws.append(["Date", "Status", "Actual", "Target", "Variance", "Running balance",
               "Compensated", "Override", "Source"])
    for c in ws[1]:
        c.font = Font(bold=True)
    for item in engine.running_ledger(db, emp, first, min(last, dt.date.today())):
        r = item["row"]
        ws.append([
            fmt_date(r.date), r.effective_status(), fmt_hm(r.actual_minutes),
            fmt_hm(r.target_minutes),
            fmt_hm(r.variance_minutes) if r.variance_minutes is not None else "unknown",
            fmt_hm(item["balance"]),
            "yes" if r.compensated else "", r.override_status or "", r.source,
        ])

    ws2 = wb.create_sheet("Entries")
    ws2.append(["Date", "Project/Employer", "Task", "Details", "Start", "End", "Duration"])
    for c in ws2[1]:
        c.font = Font(bold=True)
    entries = db.execute(
        select(m.TaskEntry)
        .where(m.TaskEntry.employee_id == emp_id, m.TaskEntry.date.between(first, last))
        .order_by(m.TaskEntry.date, m.TaskEntry.start_minute)
    ).scalars()
    from app.util import fmt_time

    for e in entries:
        ws2.append([
            fmt_date(e.date), e.project.name, e.task_type.name, e.details,
            fmt_time(e.start_minute), fmt_time(e.end_minute), fmt_hm(e.duration_minutes),
        ])

    ws3 = wb.create_sheet("Leave")
    ws3.append(["From", "To", "Type", "Hours/day", "Note", "Entered by"])
    for c in ws3[1]:
        c.font = Font(bold=True)
    for lv in db.execute(
        select(m.LeaveRecord).where(m.LeaveRecord.employee_id == emp_id)
        .order_by(m.LeaveRecord.start_date)
    ).scalars():
        ws3.append([
            fmt_date(lv.start_date), fmt_date(lv.end_date), lv.type,
            "full day" if lv.minutes_per_day is None else fmt_hm(lv.minutes_per_day),
            lv.note, lv.entered_by,
        ])
    safe = "".join(ch for ch in emp.name if ch.isalnum() or ch in " _-").strip().replace(" ", "_")
    return _xlsx_response(wb, f"{safe}_{year}-{month:02d}.xlsx")


@router.get("/entries.csv")
def entries_csv(
    start: str,
    end: str,
    admin: m.Employee = Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    # Not linked from any Dashboard/Leave/Reports screen a department-scoped
    # admin can see — an org-wide, department-unfiltered raw dump, so it
    # stays super-admin-only rather than trying to retrofit scoping onto it.
    s, e = dt.date.fromisoformat(start), dt.date.fromisoformat(end)
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["employee", "department", "date", "project", "task", "details",
                "start", "end", "duration_minutes", "imported"])
    rows = db.execute(
        select(m.TaskEntry).where(m.TaskEntry.date.between(s, e))
        .order_by(m.TaskEntry.employee_id, m.TaskEntry.date, m.TaskEntry.start_minute)
    ).scalars()
    from app.util import fmt_time

    for t in rows:
        w.writerow([
            t.employee.name, t.employee.department, fmt_date(t.date), t.project.name,
            t.task_type.name, t.details, fmt_time(t.start_minute), fmt_time(t.end_minute),
            t.duration_minutes, "1" if t.imported else "",
        ])
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="entries_{start}_{end}.csv"'},
    )
