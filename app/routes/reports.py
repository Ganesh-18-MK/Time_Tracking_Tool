"""Admin -> Reports: Attendance Reports and Strike Reports.

Both pages share the same cascading filter bar (Department -> Employee ->
Date range) and the same drill-down rule (see app/reports.py). This file is
the thin controller layer — HTML pages plus matching .xlsx exports — per
the layout convention in CLAUDE.md; all the actual aggregation lives in
app/reports.py.
"""
import datetime as dt
from typing import Dict, List, Optional

from fastapi import APIRouter, Depends, Query, Request
from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy.orm import Session

from app import models as m, reports
from app.auth import admin_department_scope, require_admin, require_developer_or_admin
from app.db import get_db
from app.templating import render
from app.util import fmt_date, fmt_hm, fmt_time, month_label, xlsx_response

router = APIRouter(prefix="/admin/reports")


def _header(ws, cols):
    ws.append(cols)
    # ws.max_row (not a hardcoded row 1) so this still bolds the right row
    # when a sheet has summary rows written before the actual header —
    # see reports_time_xlsx's filter-summary block above.
    for c in ws[ws.max_row]:
        c.font = Font(bold=True)


def _parse_date(value: str) -> Optional[dt.date]:
    try:
        return dt.date.fromisoformat(value) if value else None
    except ValueError:
        return None


def _scoped_dept(admin: m.Employee, dept: str) -> str:
    """Force the department filter for a department-scoped admin (team
    lead) — Reports is one of their three allowed screens (see
    Employee.is_super_admin docstring), but only for their own team.
    Ignores/overrides any ?dept= they might pass (including the "All
    Departments" empty string). A no-op for a super admin."""
    scope = admin_department_scope(admin)
    return scope if scope is not None else (dept or "")


def _filter_ctx(db: Session, admin: m.Employee, dept: str, emp: int, rng: str, start: str, end: str) -> dict:
    dept = _scoped_dept(admin, dept)
    start_date, end_date = reports.resolve_date_range(rng, _parse_date(start), _parse_date(end))
    departments = reports.departments_list(db)
    if admin_department_scope(admin) is not None:
        departments = [d for d in departments if d == dept]
    return {
        "dept": dept, "emp": emp, "range": rng, "start": start or "", "end": end or "",
        "resolved_start": start_date, "resolved_end": end_date,
        "departments": departments,
        "employees": reports.employees_list(db, dept or None),
        "range_presets": reports.RANGE_PRESETS,
    }


@router.get("")
def reports_landing(
    request: Request,
    admin: m.Employee = Depends(require_admin),
    db: Session = Depends(get_db),
):
    return render(request, "admin/reports_landing.html", {"user": admin}, db=db)


# ---- Attendance Reports -----------------------------------------------------
@router.get("/time")
def reports_time(
    request: Request,
    dept: str = "",
    emp: List[int] = Query([]),
    project: List[int] = Query([]),
    task: List[int] = Query([]),
    rng: str = Query("90d", alias="range"),
    start: str = "",
    end: str = "",
    admin: m.Employee = Depends(require_admin),
    db: Session = Depends(get_db),
):
    ctx = _filter_ctx(db, admin, dept, 0, rng, start, end)
    result = reports.time_by_activity_report(
        db, ctx["resolved_start"], ctx["resolved_end"], department=ctx["dept"] or None,
        employee_ids=emp or None, project_ids=project or None, task_type_ids=task or None,
    )
    project_result = reports.time_by_project_report(
        db, ctx["resolved_start"], ctx["resolved_end"], department=ctx["dept"] or None,
        employee_ids=emp or None, project_ids=project or None, task_type_ids=task or None,
    )
    filters_summary = reports.time_filters_summary(db, ctx["dept"], emp, project, task)
    kpis = reports.time_kpis(result, project_result)
    ctx.pop("emp", None)
    return render(
        request, "admin/reports_time.html",
        {
            "user": admin, "result": result, "project_result": project_result,
            "emp": emp, "project": project, "task": task,
            "projects": reports.projects_list(db), "tasks": reports.task_types_list(db),
            "filters_summary": filters_summary, "kpis": kpis,
            **ctx,
        },
        db=db,
    )


@router.get("/time.xlsx")
def reports_time_xlsx(
    dept: str = "",
    emp: List[int] = Query([]),
    project: List[int] = Query([]),
    task: List[int] = Query([]),
    rng: str = Query("90d", alias="range"),
    start: str = "",
    end: str = "",
    admin: m.Employee = Depends(require_admin),
    db: Session = Depends(get_db),
):
    dept = _scoped_dept(admin, dept)
    start_date, end_date = reports.resolve_date_range(rng, _parse_date(start), _parse_date(end))
    result = reports.time_by_activity_report(
        db, start_date, end_date, department=dept or None,
        employee_ids=emp or None, project_ids=project or None, task_type_ids=task or None,
    )
    filters_summary = reports.time_filters_summary(db, dept, emp, project, task)
    wb = Workbook()
    ws = wb.active
    ws.title = "Time by employee"
    # Filter summary rows first (Ganesh's manager, 2026-08-06) — so the
    # file is self-describing whenever it's reopened later, without needing
    # to go back to the report screen to remember what was selected.
    ws.append(["Date range", f"{fmt_date(start_date)} to {fmt_date(end_date)}"])
    ws.append(["Department", filters_summary["department"]])
    ws.append(["Employees", filters_summary["employees"]])
    ws.append(["Projects", filters_summary["projects"]])
    ws.append(["Tasks", filters_summary["tasks"]])
    for row in ws["A1:A5"]:
        row[0].font = Font(bold=True)
    ws.append([])
    month_cols = [month_label(y, mo) for y, mo in result["months"]]
    _header(ws, ["Name", "Department"] + month_cols + ["Total"])
    for r in result["rows"]:
        ws.append(
            [r["employee"].name, r["department"]]
            + [round(r["by_month"][ym] / 60, 2) for ym in result["months"]]
            + [round(r["total"] / 60, 2)]
        )
    return xlsx_response(wb, f"time_by_activity_{start_date}_{end_date}.xlsx")


@router.get("/attendance")
def reports_attendance(
    request: Request,
    dept: str = "",
    emp: int = 0,
    rng: str = Query(reports.DEFAULT_RANGE, alias="range"),
    start: str = "",
    end: str = "",
    admin: m.Employee = Depends(require_admin),
    db: Session = Depends(get_db),
):
    ctx = _filter_ctx(db, admin, dept, emp, rng, start, end)
    result = reports.attendance_report(
        db, ctx["resolved_start"], ctx["resolved_end"], department=ctx["dept"] or None, employee_id=emp or None
    )
    kpis = reports.attendance_kpis(result)
    return render(request, "admin/reports_attendance.html", {"user": admin, "result": result, "kpis": kpis, **ctx}, db=db)


@router.get("/attendance.xlsx")
def reports_attendance_xlsx(
    dept: str = "",
    emp: int = 0,
    rng: str = Query(reports.DEFAULT_RANGE, alias="range"),
    start: str = "",
    end: str = "",
    admin: m.Employee = Depends(require_admin),
    db: Session = Depends(get_db),
):
    dept = _scoped_dept(admin, dept)
    start_date, end_date = reports.resolve_date_range(rng, _parse_date(start), _parse_date(end))
    result = reports.attendance_report(db, start_date, end_date, department=dept or None, employee_id=emp or None)
    wb = Workbook()
    ws = wb.active
    if result["mode"] == "daily":
        ws.title = "Daily detail"
        _header(ws, ["Date", "Status", "Overtime (min)", "Approved overtime (min)"])
        for r in result["rows"]:
            ws.append([fmt_date(r["date"]), reports.STATUS_LABELS.get(r["status"], r["status"]),
                       r["overtime"], r["approved_overtime"]])
    else:
        ws.title = "Summary"
        _header(ws, ["Name", "Department"] + [reports.STATUS_LABELS[s] for s in reports.STATUS_ORDER]
                + ["Attendance %", "Overtime (min)", "Approved overtime (min)"])
        for r in result["rows"]:
            ws.append(
                [r["employee"].name, r["department"]]
                + [r["counts"][s] for s in reports.STATUS_ORDER]
                + [r["attendance_pct"] if r["attendance_pct"] is not None else "",
                   r["overtime_minutes"], r["approved_overtime_minutes"]]
            )
    return xlsx_response(wb, f"attendance_{start_date}_{end_date}.xlsx")


# ---- Strike Reports ----------------------------------------------------------
@router.get("/strikes")
def reports_strikes(
    request: Request,
    dept: str = "",
    emp: int = 0,
    rng: str = Query(reports.DEFAULT_RANGE, alias="range"),
    start: str = "",
    end: str = "",
    admin: m.Employee = Depends(require_admin),
    db: Session = Depends(get_db),
):
    ctx = _filter_ctx(db, admin, dept, emp, rng, start, end)
    result = reports.strikes_report(
        db, ctx["resolved_start"], ctx["resolved_end"], department=ctx["dept"] or None, employee_id=emp or None
    )
    kpis = reports.strikes_kpis(result)
    return render(request, "admin/reports_strikes.html", {"user": admin, "result": result, "kpis": kpis, **ctx}, db=db)


@router.get("/strikes.xlsx")
def reports_strikes_xlsx(
    dept: str = "",
    emp: int = 0,
    rng: str = Query(reports.DEFAULT_RANGE, alias="range"),
    start: str = "",
    end: str = "",
    admin: m.Employee = Depends(require_admin),
    db: Session = Depends(get_db),
):
    dept = _scoped_dept(admin, dept)
    start_date, end_date = reports.resolve_date_range(rng, _parse_date(start), _parse_date(end))
    result = reports.strikes_report(db, start_date, end_date, department=dept or None, employee_id=emp or None)
    wb = Workbook()
    ws = wb.active
    if result["mode"] == "daily":
        ws.title = "Strike days"
        _header(ws, ["Date", "Status"])
        for r in result["rows"]:
            ws.append([fmt_date(r["date"]), reports.STATUS_LABELS.get(r["status"], r["status"])])
    else:
        ws.title = "Summary"
        _header(ws, ["Name", "Department", "Strikes"])
        for r in result["rows"]:
            ws.append([r["employee"].name, r["department"], r["strikes"]])
    return xlsx_response(wb, f"strikes_{start_date}_{end_date}.xlsx")


# ---- Developer Usage Report (Ganesh, 2026-08-21) ------------------------------
# Deliberately gated by require_developer_or_admin, not require_admin — a
# plain-employee Developer with no admin flag can see this even though it
# lives under /admin/reports/*, same URL family as the other 3 reports for
# consistency. Org-wide, no department scoping (see
# require_developer_or_admin's docstring) — adoption of a logging method
# isn't a per-team question the way Attendance/Strikes are.
@router.get("/usage")
def reports_usage(
    request: Request,
    rng: str = Query(reports.DEFAULT_RANGE, alias="range"),
    start: str = "",
    end: str = "",
    user: m.Employee = Depends(require_developer_or_admin),
    db: Session = Depends(get_db),
):
    start_date, end_date = reports.resolve_date_range(rng, _parse_date(start), _parse_date(end))
    result = reports.feature_usage_report(db, start_date, end_date)
    return render(
        request, "admin/reports_usage.html",
        {
            "user": user, "result": result, "range": rng, "start": start or "", "end": end or "",
            "resolved_start": start_date, "resolved_end": end_date,
            "range_presets": reports.RANGE_PRESETS,
        },
        db=db,
    )


@router.get("/usage.xlsx")
def reports_usage_xlsx(
    rng: str = Query(reports.DEFAULT_RANGE, alias="range"),
    start: str = "",
    end: str = "",
    user: m.Employee = Depends(require_developer_or_admin),
    db: Session = Depends(get_db),
):
    start_date, end_date = reports.resolve_date_range(rng, _parse_date(start), _parse_date(end))
    result = reports.feature_usage_report(db, start_date, end_date)
    wb = Workbook()
    ws = wb.active
    ws.title = "Adoption summary"
    ws.append(["Date range", f"{fmt_date(start_date)} to {fmt_date(end_date)}"])
    ws.append(["Active tracked employees", result["total_employees"]])
    ws["A1"].font = Font(bold=True)
    ws["A2"].font = Font(bold=True)
    ws.append([])
    _header(ws, ["Method", "Employees who used it", "% adoption"])
    for row in result["methods"]:
        ws.append([row["label"], row["count"], row["pct"]])
    ws.append(["Punch In/Out", result["punch"]["count"], result["punch"]["pct"]])
    ws.append([])
    ws2 = wb.create_sheet("By employee")
    _header(
        ws2,
        ["Name", "Department"] + [m.ENTRY_METHOD_LABELS[k] for k in m.ENTRY_METHODS] + ["Punch In/Out"],
    )
    for r in result["employees"]:
        ws2.append(
            [r["employee"].name, r["employee"].department or "—"]
            + ["Yes" if r["used"][k] else "" for k in m.ENTRY_METHODS]
            + ["Yes" if r["punch"] else ""]
        )
    return xlsx_response(wb, f"feature_usage_{start_date}_{end_date}.xlsx")


# ---- Task Logs Report (Ganesh, 2026-08-29) ------------------------------------
@router.get("/tasklogs")
def reports_tasklogs(
    request: Request,
    dept: str = "",
    emp: int = 0,
    rng: str = Query(reports.DEFAULT_RANGE, alias="range"),
    start: str = "",
    end: str = "",
    admin: m.Employee = Depends(require_admin),
    db: Session = Depends(get_db),
):
    ctx = _filter_ctx(db, admin, dept, emp, rng, start, end)
    result = reports.daily_task_log_report(
        db, ctx["resolved_start"], ctx["resolved_end"], department=ctx["dept"] or None, employee_id=emp or None
    )
    return render(
        request, "admin/reports_tasklogs.html",
        {"user": admin, "result": result, **ctx},
        db=db,
    )


@router.get("/tasklogs_employee.xlsx")
def reports_tasklogs_employee_xlsx(
    dept: str = "",
    emp: int = 0,
    rng: str = Query(reports.DEFAULT_RANGE, alias="range"),
    start: str = "",
    end: str = "",
    admin: m.Employee = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Download #1 of 2 (Ganesh, 2026-08-29: "i want download report option
    as well like project wise and employee wise") — the same filtered rows
    reports.task_log_export_rows() returns, sorted by employee then date.
    Shares that one query with the project-wise export below so the two
    downloads can never disagree about which entries are in scope."""
    dept = _scoped_dept(admin, dept)
    start_date, end_date = reports.resolve_date_range(rng, _parse_date(start), _parse_date(end))
    rows = reports.task_log_export_rows(db, start_date, end_date, department=dept or None, employee_id=emp or None)
    rows.sort(key=lambda r: (r["employee"].name, r["date"], r["start_minute"]))
    wb = Workbook()
    ws = wb.active
    ws.title = "By employee"
    _header(ws, ["Name", "Department", "Date", "Project", "Task", "Client", "Details",
                 "Start", "End", "Duration", "Unplanned"])
    for r in rows:
        ws.append([
            r["employee"].name, r["employee"].department or "—", fmt_date(r["date"]),
            r["project"].name, r["task"].name, r["client"], r["details"],
            fmt_time(r["start_minute"]), fmt_time(r["end_minute"]), fmt_hm(r["duration_minutes"]),
            "Yes" if r["unplanned"] else "",
        ])
    return xlsx_response(wb, f"tasklogs_by_employee_{start_date}_{end_date}.xlsx")


@router.get("/tasklogs_project.xlsx")
def reports_tasklogs_project_xlsx(
    dept: str = "",
    emp: int = 0,
    rng: str = Query(reports.DEFAULT_RANGE, alias="range"),
    start: str = "",
    end: str = "",
    admin: m.Employee = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Download #2 of 2 — same underlying rows as the employee-wise export
    above, just sorted/grouped by project instead. Second sheet is a
    per-project time total, mirroring the "By Project" breakdown Time by
    Project/Task already has, so the two "by project" views in this app
    look and feel consistent."""
    dept = _scoped_dept(admin, dept)
    start_date, end_date = reports.resolve_date_range(rng, _parse_date(start), _parse_date(end))
    rows = reports.task_log_export_rows(db, start_date, end_date, department=dept or None, employee_id=emp or None)
    rows.sort(key=lambda r: (r["project"].name, r["employee"].name, r["date"], r["start_minute"]))
    wb = Workbook()
    ws = wb.active
    ws.title = "By project"
    _header(ws, ["Project", "Task", "Name", "Department", "Date", "Client", "Details",
                 "Start", "End", "Duration", "Unplanned"])
    for r in rows:
        ws.append([
            r["project"].name, r["task"].name, r["employee"].name, r["employee"].department or "—",
            fmt_date(r["date"]), r["client"], r["details"],
            fmt_time(r["start_minute"]), fmt_time(r["end_minute"]), fmt_hm(r["duration_minutes"]),
            "Yes" if r["unplanned"] else "",
        ])
    totals: Dict[str, int] = {}
    for r in rows:
        totals[r["project"].name] = totals.get(r["project"].name, 0) + r["duration_minutes"]
    ws2 = wb.create_sheet("Project totals")
    _header(ws2, ["Project", "Total"])
    for name, minutes in sorted(totals.items(), key=lambda kv: -kv[1]):
        ws2.append([name, fmt_hm(minutes)])
    return xlsx_response(wb, f"tasklogs_by_project_{start_date}_{end_date}.xlsx")
