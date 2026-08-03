"""Bulk employee onboarding + bulk update via Excel (Roster -> Bulk upload).

One upload handles two cases, told apart per-row by whether Employee ID is
filled in:

  * Employee ID blank  -> onboard a NEW employee. Full Name, Department,
    Designation, Target/day and Workdays are required; everything else is
    optional. A new Employee ID (LOMK001, LOMK002, ...) is generated and
    assigned automatically — it is never something the uploader supplies.
  * Employee ID filled -> UPDATE that existing employee. Nothing is
    required. Only the columns actually filled in are applied — a blank
    cell means "leave this field as it is", not "clear it". This is what
    makes "just add phone numbers for the people who have one" safe: a
    sheet with only Employee ID + Phone filled in touches nothing else.
    Setting Action to "Deactivate" bulk-offboards that employee — this is
    always a soft deactivate (Employee.active = False), never a hard
    delete, so all their time/leave/strike history stays intact. Same
    effect as Roster -> Edit -> untick Active, just for many people at once.

Row-parsing rules are plain functions you can read and test in isolation,
the same way app/validation.py keeps entry rules separate from the routes
that call them. Only process_upload() touches the database.
"""
import datetime as dt
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import select

from app import models as m
from app.util import flags_to_role, format_employee_code, highest_employee_code_number, role_to_flags

MAX_ROWS = 500
_PENDING = -1  # sentinel: "claimed by a new-employee row earlier in this same upload"

TEMPLATE_HEADERS = [
    "Employee ID", "Full Name", "Department", "Designation", "Target/day",
    "Workdays", "Joining Date", "DOB", "Email", "Country Code", "Phone", "Role",
    "Reports To", "Action",
]
REQUIRED_FOR_NEW = ["Full Name", "Department", "Designation", "Target/day", "Workdays"]
COL_WIDTHS = [12, 22, 16, 18, 11, 16, 14, 12, 26, 12, 16, 10, 14, 12]
COL_LETTERS = "ABCDEFGHIJKLMN"  # one per TEMPLATE_HEADERS column, in order

# Both the exact letter-codes from the spec (M/T/W/Th/F/S/Su) and a few
# common 3-letter aliases, so a sheet that says "Mon,Tue,Wed" instead still
# works. Monday=0 matches Employee.work_days elsewhere in the app.
WORKDAY_TOKENS = {
    "m": 0, "mon": 0,
    "t": 1, "tue": 1,
    "w": 2, "wed": 2,
    "th": 3, "thu": 3,
    "f": 4, "fri": 4,
    "s": 5, "sat": 5,
    "su": 6, "sun": 6,
}
WORKDAY_LETTERS = {0: "M", 1: "T", 2: "W", 3: "Th", 4: "F", 5: "S", 6: "Su"}

_DATE_FORMATS = ["%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y", "%d %b %Y", "%d %B %Y"]


def workdays_to_letters(work_days: str) -> str:
    """'0,1,2,3,4' -> 'M,T,W,Th,F' — used by the existing-employees export."""
    try:
        nums = sorted(int(x) for x in work_days.split(",") if x.strip() != "")
    except ValueError:
        return ""
    return ",".join(WORKDAY_LETTERS.get(n, "") for n in nums)


def parse_cell_date(raw) -> Optional[dt.date]:
    """Blank -> None (in both modes: "no date" for a new hire, or "no
    change" on an update). Excel date/datetime cells pass through as-is;
    text cells are tried against a handful of common formats."""
    if raw is None:
        return None
    if isinstance(raw, dt.datetime):
        return raw.date()
    if isinstance(raw, dt.date):
        return raw
    text = str(raw).strip()
    if not text:
        return None
    for fmt in _DATE_FORMATS:
        try:
            return dt.datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"'{text}' isn't a recognized date — use YYYY-MM-DD")


def parse_target_minutes(raw) -> Optional[int]:
    """Blank -> None ("required" is enforced by the caller for new rows
    only). Accepts a plain number of hours (8, 7.5) or H:MM text (8:00)."""
    if raw is None or str(raw).strip() == "":
        return None
    try:
        if isinstance(raw, (int, float)):
            hours = float(raw)
        else:
            text = str(raw).strip()
            if ":" in text:
                h, mm = text.split(":")
                hours = int(h) + int(mm) / 60
            else:
                hours = float(text)
    except (ValueError, TypeError):
        raise ValueError(
            "Target/day must be a number of hours (e.g. 8 or 7.5) or H:MM (e.g. 8:00)"
        )
    minutes = int(round(hours * 60))
    if minutes <= 0:
        raise ValueError("Target/day must be greater than zero")
    return minutes


WORKDAY_DEFAULT_TOKENS = {"default", "def"}


def parse_workdays(raw) -> Optional[str]:
    """Blank -> None (see parse_target_minutes). 'Default' (or 'Def') is a
    shortcut for the standard 5-day Monday-to-Friday week (Ganesh,
    2026-08-01) — most employees use this, so a sheet can just say
    "Default" instead of typing "M,T,W,Th,F" on every row; anyone on a
    different schedule still lists their own days explicitly, exactly as
    before."""
    if raw is None or str(raw).strip() == "":
        return None
    text = str(raw).strip().lower()
    if text in WORKDAY_DEFAULT_TOKENS:
        return ",".join(str(d) for d in range(5))  # Mon-Fri: 0,1,2,3,4
    tokens = [t.strip().lower() for t in text.split(",") if t.strip() != ""]
    days, bad = set(), []
    for t in tokens:
        if t in WORKDAY_TOKENS:
            days.add(WORKDAY_TOKENS[t])
        else:
            bad.append(t)
    if bad:
        raise ValueError(
            f"Workdays has unrecognized value(s): {', '.join(bad)} "
            "— use M,T,W,Th,F,S,Su, or 'Default' for Mon-Fri"
        )
    return ",".join(str(d) for d in sorted(days))


def parse_role(raw) -> Optional[Tuple[bool, bool]]:
    """Blank -> None (new rows default this to Employee; update rows leave
    the existing role alone — see parse_row). Returns (is_admin,
    is_super_admin) otherwise — 'Admin' is the department-scoped tier
    (Dashboard/Leave Requests/Reports for their own department only),
    'Super Admin' is org-wide (see Employee.is_super_admin docstring)."""
    if raw is None or str(raw).strip() == "":
        return None
    text = str(raw).strip().lower()
    if text == "employee":
        return role_to_flags("employee")
    if text == "admin":
        return role_to_flags("admin")
    if text in ("super admin", "super_admin", "superadmin"):
        return role_to_flags("super_admin")
    raise ValueError("Role must be 'Employee', 'Admin', or 'Super Admin'")


def parse_action(raw) -> Optional[bool]:
    """Blank -> None (no action). True -> the row asks to deactivate that
    employee (bulk offboarding). Deactivation only — never a hard delete —
    so all their time/leave/strike history stays intact, same as Roster ->
    Edit -> untick Active."""
    if raw is None or str(raw).strip() == "":
        return None
    text = str(raw).strip().lower()
    if text in ("deactivate", "offboard"):
        return True
    raise ValueError("Action must be blank or 'Deactivate'")


def parse_row(raw: dict, code_to_id: Dict[str, int]) -> dict:
    """Returns one of:
      {"mode": "new", "fields": {...}, "error": None}
      {"mode": "update", "employee_id": int, "fields": {...}, "error": None}
      {"mode": "error", "fields": {}, "error": "..."}

    "new" fields always include name/department/designation/
    daily_target_minutes/work_days (already validated non-blank) plus
    whatever optional columns were filled. "update" fields is a *patch* —
    only keys the sheet actually filled in are present; the caller must
    only touch those attributes on the existing employee.
    """
    errors: List[str] = []

    emp_id_raw = str(raw.get("Employee ID") or "").strip()
    is_update = bool(emp_id_raw)
    target_id = None
    if is_update:
        target_id = code_to_id.get(emp_id_raw.upper())
        if target_id is None:
            return {"mode": "error", "fields": {}, "error": f"Employee ID '{emp_id_raw}' not found"}

    fields: dict = {}

    def _text(key: str) -> str:
        v = raw.get(key)
        return str(v).strip() if v not in (None, "") else ""

    name = _text("Full Name")
    if name:
        fields["name"] = name
    elif not is_update:
        errors.append("Full Name is required")

    department = _text("Department")
    if department:
        fields["department"] = department
    elif not is_update:
        errors.append("Department is required")

    designation = _text("Designation")
    if designation:
        fields["designation"] = designation
    elif not is_update:
        errors.append("Designation is required")

    target_minutes = None
    try:
        target_minutes = parse_target_minutes(raw.get("Target/day"))
    except ValueError as e:
        errors.append(str(e))
    if target_minutes is not None:
        fields["daily_target_minutes"] = target_minutes
    elif not is_update:
        errors.append("Target/day is required")

    work_days = None
    try:
        work_days = parse_workdays(raw.get("Workdays"))
    except ValueError as e:
        errors.append(str(e))
    if work_days is not None:
        fields["work_days"] = work_days
    elif not is_update:
        errors.append("Workdays is required")

    start_date = None
    try:
        start_date = parse_cell_date(raw.get("Joining Date"))
    except ValueError as e:
        errors.append(str(e))
    if start_date is not None:
        fields["start_date"] = start_date

    dob = None
    try:
        dob = parse_cell_date(raw.get("DOB"))
    except ValueError as e:
        errors.append(str(e))
    if dob is not None:
        fields["date_of_birth"] = dob

    email_raw = raw.get("Email")
    email = str(email_raw).strip() if email_raw not in (None, "") else ""
    if email:
        if "@" not in email:
            errors.append("Email doesn't look valid")
        else:
            fields["email"] = email

    country_code_raw = raw.get("Country Code")
    country_code = str(country_code_raw).strip() if country_code_raw not in (None, "") else ""
    if country_code:
        fields["country_code"] = country_code

    phone_raw = raw.get("Phone")
    phone = str(phone_raw).strip() if phone_raw not in (None, "") else ""
    if phone:
        fields["phone"] = phone

    role_flags = None
    try:
        role_flags = parse_role(raw.get("Role"))
    except ValueError as e:
        errors.append(str(e))
    if role_flags is not None:
        fields["is_admin"], fields["is_super_admin"] = role_flags
    elif not is_update:
        fields["is_admin"], fields["is_super_admin"] = False, False  # default for new rows only

    # Must reference an employee who already exists BEFORE this upload —
    # two new-hire rows in the same sheet can't point at each other, since
    # neither has an Employee ID yet. Matches the same code_to_id lookup
    # Employee ID itself uses for update-mode rows.
    reports_to_raw = _text("Reports To")
    if reports_to_raw:
        reports_to_id = code_to_id.get(reports_to_raw.upper())
        if reports_to_id is None:
            errors.append(f"Reports To '{reports_to_raw}' isn't a known Employee ID")
        else:
            fields["reports_to_id"] = reports_to_id

    deactivate = None
    try:
        deactivate = parse_action(raw.get("Action"))
    except ValueError as e:
        errors.append(str(e))
    if deactivate:
        if is_update:
            fields["active"] = False
        else:
            errors.append("Action can only deactivate an existing employee — Employee ID is required")

    if errors:
        return {"mode": "error", "fields": {}, "error": "; ".join(errors)}
    if is_update:
        return {"mode": "update", "employee_id": target_id, "fields": fields, "error": None}
    return {"mode": "new", "fields": fields, "error": None}


def read_upload_rows(wb: Workbook) -> Tuple[List[dict], Optional[str]]:
    """Returns (rows, error). error is set (rows empty) only when the sheet
    itself is unusable (no header row at all) — every column is optional at
    the sheet level, since which ones matter depends on new-vs-update mode,
    decided per row (see parse_row)."""
    ws = wb.active
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    # A brand-new openpyxl sheet isn't truly empty even with nothing written
    # to it — it still has a default A1 cell, so iter_rows yields (None,)
    # rather than nothing at all. Treat an all-blank header row the same as
    # no header row, or a genuinely empty upload silently produces zero rows
    # and zero errors instead of the "sheet is empty" message it should get.
    if header_row is None or all(c is None or str(c).strip() == "" for c in header_row):
        return [], "The sheet is empty — no header row found."
    header_map = {}
    for idx, cell in enumerate(header_row):
        if cell is None:
            continue
        header_map[str(cell).strip().lower()] = idx
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue  # blank row — skip silently, not an error
        rows.append({
            field: (row[header_map[field.lower()]]
                     if field.lower() in header_map and header_map[field.lower()] < len(row)
                     else None)
            for field in TEMPLATE_HEADERS
        })
    return rows, None


def process_upload(db, wb: Workbook) -> dict:
    """Parses + applies an uploaded workbook. New-employee rows and
    update-existing rows can be mixed freely in the same sheet. Valid rows
    are applied and committed in one transaction; invalid, duplicate, or
    unresolvable rows are skipped and listed with a reason, never silently
    dropped. Returns {"added": int, "updated": int, "deactivated": int,
    "skipped": [{"row": int, "name": str, "reason": str}],
    "header_error": str | None}. "deactivated" is a subset of "updated" —
    called out separately since offboarding is a more consequential action
    than a routine field edit and the admin should see it clearly."""
    rows, header_error = read_upload_rows(wb)
    if header_error:
        return {"added": 0, "updated": 0, "deactivated": 0, "skipped": [], "header_error": header_error}
    if len(rows) > MAX_ROWS:
        return {
            "added": 0, "updated": 0, "deactivated": 0, "skipped": [],
            "header_error": f"Sheet has {len(rows)} data rows — max is {MAX_ROWS} per upload. Split it into batches.",
        }

    existing = list(
        db.execute(select(m.Employee.id, m.Employee.employee_code, m.Employee.email, m.Employee.name)).all()
    )
    code_to_id = {code.upper(): eid for (eid, code, _e, _n) in existing if code}
    email_to_id = {e.lower(): eid for (eid, _c, e, _n) in existing if e}
    name_to_id = {n.strip().lower(): eid for (eid, _c, _e, n) in existing if n}

    next_n = highest_employee_code_number(db) + 1

    added = updated = 0
    skipped = []
    to_add: List[dict] = []
    to_update: List[Tuple[int, dict]] = []

    for i, raw in enumerate(rows, start=2):  # row 1 is the header
        display = (str(raw.get("Full Name") or raw.get("Employee ID") or "").strip()) or "(blank)"
        result = parse_row(raw, code_to_id)
        if result["error"]:
            skipped.append({"row": i, "name": display, "reason": result["error"]})
            continue

        if result["mode"] == "new":
            fields = result["fields"]
            name_key = fields["name"].strip().lower()
            email_key = (fields.get("email") or "").lower()
            if name_key in name_to_id or (email_key and email_key in email_to_id):
                skipped.append({
                    "row": i, "name": fields["name"],
                    "reason": "Already exists in the roster (matched on name or email) — skipped, nothing overwritten",
                })
                continue
            name_to_id[name_key] = _PENDING
            if email_key:
                email_to_id[email_key] = _PENDING
            fields["employee_code"] = format_employee_code(next_n)
            next_n += 1
            to_add.append(fields)
        else:  # update
            emp_id = result["employee_id"]
            fields = result["fields"]
            email_key = (fields.get("email") or "").lower()
            if email_key:
                owner = email_to_id.get(email_key)
                if owner is not None and owner != emp_id:
                    skipped.append({
                        "row": i, "name": display,
                        "reason": f"Email '{fields['email']}' is already used by another employee — skipped",
                    })
                    continue
                email_to_id[email_key] = emp_id
            to_update.append((emp_id, fields))

    for fields in to_add:
        db.add(m.Employee(**fields))
        added += 1

    deactivated = 0
    if to_update:
        ids = [eid for eid, _ in to_update]
        emp_by_id = {e.id: e for e in db.execute(select(m.Employee).where(m.Employee.id.in_(ids))).scalars()}
        for emp_id, fields in to_update:
            emp = emp_by_id.get(emp_id)
            if emp is None:
                continue
            if fields.get("active") is False:
                deactivated += 1
            for key, value in fields.items():
                setattr(emp, key, value)
            updated += 1

    if added or updated:
        db.commit()
    return {"added": added, "updated": updated, "deactivated": deactivated, "skipped": skipped, "header_error": None}


def build_sample_workbook() -> Workbook:
    """The onboarding-oriented template — Employee ID column present but
    left blank in the example row, since new hires don't have one yet."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Employees"
    ws.append(TEMPLATE_HEADERS)
    for c in ws[1]:
        c.font = Font(bold=True)
    ws.append([
        "", "Jane Doe", "Accounts", "Associate", 8,
        "Default", "2026-08-03", "1990-05-14", "jane.doe@example.com",
        "+1", "555 0100", "Employee", "", "",
    ])
    for col, width in zip(COL_LETTERS, COL_WIDTHS):
        ws.column_dimensions[col].width = width

    info = wb.create_sheet("Instructions")
    info.append(["Column", "Required?", "Format / allowed values"])
    for c in info[1]:
        c.font = Font(bold=True)
    for row in [
        ("Employee ID", "Leave blank for new hires", "System-assigned (LOMK001, ...) — never type your own"),
        ("Full Name", "Yes, for new hires", "Any text"),
        ("Department", "Yes, for new hires", "Any text, e.g. Accounts"),
        ("Designation", "Yes, for new hires", "Any text, e.g. Associate"),
        ("Target/day", "Yes, for new hires", "Hours as a number, e.g. 8 or 7.5, or H:MM like 8:00"),
        ("Workdays", "Yes, for new hires", "'Default' = Mon-Fri (most employees). For anything else, comma-separated: "
         "M=Mon, T=Tue, W=Wed, Th=Thu, F=Fri, S=Sat, Su=Sun, e.g. M,T,W,Th,F,S"),
        ("Joining Date", "No", "YYYY-MM-DD, e.g. 2026-08-03"),
        ("DOB", "No", "YYYY-MM-DD"),
        ("Email", "No, but needed before the employee can sign in", "name@company.com"),
        ("Country Code", "No", "e.g. +91, +1 — kept separate from Phone so it always round-trips cleanly"),
        ("Phone", "No", "Just the number, without the country code, e.g. 9876543210"),
        ("Role", "No — defaults to Employee", "Employee, Admin (own department only), or Super Admin (all departments)"),
        ("Reports To", "No", "The Employee ID (e.g. LOMK003) of their team lead/manager — must already exist; "
         "can't point at another new hire in the same sheet"),
        ("Action", "No — only valid on an update row (Employee ID filled in)",
         "Blank, or 'Deactivate' to offboard — keeps all their history, same as Roster -> Edit -> untick Active"),
        ("", "", ""),
        ("To UPDATE an existing employee instead of onboarding a new one,", "", ""),
        ("use Roster -> Bulk upload -> \"Download existing employees\" and fill in", "", ""),
        ("just the columns you're changing. Blank cells are left untouched.", "", ""),
    ]:
        info.append(row)
    info.column_dimensions["A"].width = 40
    info.column_dimensions["B"].width = 26
    info.column_dimensions["C"].width = 60
    return wb


def build_existing_employees_workbook(db) -> Workbook:
    """The update-oriented download — every active employee, one row each,
    Employee ID first so it survives round-tripping through Excel. Every
    other column is pre-filled with the CURRENT value so the admin can see
    what they're changing; re-uploading unchanged is a safe no-op."""
    emps = list(
        db.execute(
            select(m.Employee)
            .where(m.Employee.active.is_(True))
            .order_by(m.Employee.department, m.Employee.name)
        ).scalars()
    )
    # id -> code for EVERY employee (not just active ones), so a lead who's
    # since been deactivated still round-trips correctly on export
    code_by_id = {
        row.id: row.employee_code
        for row in db.execute(select(m.Employee.id, m.Employee.employee_code)).all()
    }
    wb = Workbook()
    ws = wb.active
    ws.title = "Employees"
    ws.append(TEMPLATE_HEADERS)
    for c in ws[1]:
        c.font = Font(bold=True)
    for e in emps:
        ws.append([
            e.employee_code or "", e.name, e.department, e.designation,
            round(e.daily_target_minutes / 60, 2), workdays_to_letters(e.work_days),
            e.start_date.isoformat() if e.start_date else "",
            e.date_of_birth.isoformat() if e.date_of_birth else "",
            e.email or "", e.country_code or "", e.phone or "",
            {"employee": "Employee", "admin": "Admin", "super_admin": "Super Admin"}[
                flags_to_role(e.is_admin, e.is_super_admin)
            ],
            (code_by_id.get(e.reports_to_id) or "") if e.reports_to_id else "",
            "",
        ])
    for col, width in zip(COL_LETTERS, COL_WIDTHS):
        ws.column_dimensions[col].width = width
    return wb
