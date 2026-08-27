"""Bulk-add Project/Employer and Task Type dropdown values via Excel
(Projects & Tasks -> Bulk upload).

Smaller sibling of app/bulk_upload.py and app/leave_bulk_upload.py, same
shape: plain parsing functions the routes call, only process_upload()
touches the database. There's no "update" mode (nothing to update on a
dropdown value beyond the name itself, which this sheet never renames) and
no deactivate-via-sheet (that stays a single click on the Lists page, same
as it already is) — this only ever adds new values/links, exactly like
typing into the "Add project/task" box on that page, just many at once.

`kind` is "project" or "task" throughout, matching the same sentinel
already used by app/routes/admin.py's lists_add()/lists_toggle().

Project-scoped tasks (Ganesh, 2026-08-27 — see ProjectTask's docstring in
app/models.py): the Tasks sheet is no longer a single column. Each row is
now a (Task Name, Project Name) PAIR, not just a task name — one row per
task+project link, so a task meant for 3 projects needs 3 rows (same task
name, a different project each time), the simplest shape to parse and
consistent with this sheet's existing one-row-one-thing convention rather
than a comma-separated multi-project cell. The Project Name must already
exist (this sheet only ever adds Tasks/links, never Projects, even from
the Tasks sheet) — a "Projects" reference sheet is included in both the
sample template and the existing-values download specifically so an admin
can copy-paste the exact spelling before uploading.

Department-scoped projects (Ganesh, 2026-08-28 — see ProjectDepartment's
docstring in app/models.py): the Projects sheet now ALSO accepts an
optional second column, Department — same one-row-per-pair convention the
Tasks sheet already established above, not a comma-separated cell, so the
same project name can appear on more than one row to link it to more than
one department. Unlike the Tasks sheet's Project Name column, Department
is genuinely optional at the HEADER level too: a plain single-column
Projects sheet (every one uploaded before this feature existed, and
read_upload_names()'s own shape below) still uploads exactly as before —
every project ends up unrestricted, same as if the column were present but
every cell left blank. Also unlike Tasks (which only ever links an
EXISTING project, never creates one), this sheet can both create a brand
new project AND add a department link to it in the same pass, or add a
department link to a project that already existed before this upload —
whichever the row's Project Name resolves to.
"""
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import select

from app import models as m, reports

MAX_ROWS = 500

HEADERS = {
    "project": "Project / Employer Name",
    "task": "Task Name",
}
TASK_PROJECT_HEADER = "Project Name"
PROJECT_DEPARTMENT_HEADER = "Department"
SHEET_TITLES = {
    "project": "Projects",
    "task": "Tasks",
}
MODELS = {
    "project": m.Project,
    "task": m.TaskType,
}


def _model_for(kind: str):
    model = MODELS.get(kind)
    if model is None:
        raise ValueError(f"Unknown kind '{kind}' — must be 'project' or 'task'")
    return model


def read_upload_names(wb: Workbook, kind: str) -> Tuple[List[Tuple[int, str]], Optional[str]]:
    """Single-column reader — one name per row, no other columns read even
    if present. Kept as its own tested building block (department-scoped
    projects, Ganesh, 2026-08-28, didn't change this function at all) even
    though the real Projects upload path now goes through
    read_project_rows() below instead, which is a superset of this exact
    shape plus an optional Department column. Still what kind="task"
    would use if it ever needed a plain name-only read, though in practice
    it always uses read_task_rows()'s two-column shape today."""
    header = HEADERS[kind]
    ws = wb.active
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if header_row is None or all(c is None or str(c).strip() == "" for c in header_row):
        return [], "The sheet is empty — no header row found."
    col_idx = None
    for idx, cell in enumerate(header_row):
        if cell is not None and str(cell).strip().lower() == header.lower():
            col_idx = idx
            break
    if col_idx is None:
        return [], f"Expected a column called '{header}' — use the sample template."

    out: List[Tuple[int, str]] = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or col_idx >= len(row):
            continue
        value = row[col_idx]
        name = str(value).strip() if value is not None else ""
        if not name:
            continue  # blank row/cell — skip silently, not an error
        out.append((i, name))
    return out, None


def read_project_rows(wb: Workbook) -> Tuple[List[Tuple[int, str, str]], Optional[str]]:
    """Projects sheet — returns [(row_number, project_name, department), ...].
    Department-scoped projects (Ganesh, 2026-08-28) — see
    ProjectDepartment's docstring in app/models.py. Unlike read_task_rows()
    below, the Department column here is genuinely OPTIONAL at the header
    level: if it's missing from the sheet entirely, every row just gets
    department="" (unrestricted) — a plain single-column Projects sheet
    from before this feature, including read_upload_names()'s own shape
    above, still uploads exactly as before. A row's Department cell can
    also simply be left blank even when the column exists — same
    "unrestricted" meaning for that one row. Same one-row-per-pair
    convention Project-scoped Tasks already established for its own second
    column: the same project name can appear on more than one row, once
    per department it should be linked to, not a comma-separated cell."""
    header = HEADERS["project"]
    ws = wb.active
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if header_row is None or all(c is None or str(c).strip() == "" for c in header_row):
        return [], "The sheet is empty — no header row found."
    name_idx = dept_idx = None
    for idx, cell in enumerate(header_row):
        label = str(cell).strip().lower() if cell is not None else ""
        if label == header.lower():
            name_idx = idx
        elif label == PROJECT_DEPARTMENT_HEADER.lower():
            dept_idx = idx
    if name_idx is None:
        return [], f"Expected a column called '{header}' — use the sample template."

    out: List[Tuple[int, str, str]] = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or name_idx >= len(row):
            continue
        name_val = row[name_idx]
        name = str(name_val).strip() if name_val is not None else ""
        if not name:
            continue  # blank row/cell — skip silently, not an error
        dept = ""
        if dept_idx is not None and dept_idx < len(row) and row[dept_idx] is not None:
            dept = str(row[dept_idx]).strip()
        out.append((i, name, dept))
    return out, None


def read_task_rows(wb: Workbook) -> Tuple[List[Tuple[int, str, str]], Optional[str]]:
    """Tasks sheet — returns [(row_number, task_name, project_name), ...].
    Both columns must be present; a row missing either value is skipped
    silently (blank row), same "not an error" treatment
    read_upload_names() gives a blank single-column row."""
    ws = wb.active
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if header_row is None or all(c is None or str(c).strip() == "" for c in header_row):
        return [], "The sheet is empty — no header row found."
    task_idx = project_idx = None
    for idx, cell in enumerate(header_row):
        label = str(cell).strip().lower() if cell is not None else ""
        if label == HEADERS["task"].lower():
            task_idx = idx
        elif label == TASK_PROJECT_HEADER.lower():
            project_idx = idx
    if task_idx is None or project_idx is None:
        return [], f"Expected columns called '{HEADERS['task']}' and '{TASK_PROJECT_HEADER}' — use the sample template."

    out: List[Tuple[int, str, str]] = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row is None:
            continue
        task_val = row[task_idx] if task_idx < len(row) else None
        project_val = row[project_idx] if project_idx < len(row) else None
        task_name = str(task_val).strip() if task_val is not None else ""
        project_name = str(project_val).strip() if project_val is not None else ""
        if not task_name and not project_name:
            continue  # fully blank row — skip silently
        out.append((i, task_name, project_name))
    return out, None


def process_upload(db, wb: Workbook, kind: str) -> dict:
    """Parses + applies an uploaded workbook. New rows are added and
    committed in one transaction; anything skipped is never silently
    dropped or duplicated — always reported with a reason. Returns
    {"added": int, "skipped": [{"row": int, "name": str, "reason": str}],
    "header_error": str | None}, plus, for kind="project" only,
    "department_links_added": int (Ganesh, 2026-08-28 — see
    _process_project_upload()'s docstring).

    kind="project": a name already on the list (case-sensitive exact
    match, same rule the single "Add" box on the Lists page uses) or
    repeated within the file with no Department cell is skipped;
    "added" still means new PROJECT rows created, unchanged from before
    this feature — see _process_project_upload() for how a Department
    column (optional) is handled separately.

    kind="task" (Ganesh, 2026-08-27): each row is a (task, project) pair,
    not just a task name — see this module's docstring. "added" counts
    task+project LINKS created, not distinct task names (a task linked to
    3 projects via 3 rows counts as 3, matching "N project(s)" wording
    used elsewhere for this feature, e.g. admin/lists.html)."""
    if kind == "project":
        return _process_project_upload(db, wb)
    return _process_task_upload(db, wb)


def _process_project_upload(db, wb: Workbook) -> dict:
    """Department-scoped projects (Ganesh, 2026-08-28) — rewritten to use
    read_project_rows() (Project Name + optional Department) instead of
    read_upload_names(), so this now does two independent things per row:
    (1) create the project if its name isn't already known (from the DB or
    an earlier row in this same file) — "added" counts ONLY this, same
    meaning "N project(s) added" already had before this feature; (2) if
    the row also has a Department value, link that project to it — counted
    separately in "department_links_added" so the two numbers can't be
    confused with each other in the flash message. This can add a
    department link to a project that already existed before this upload,
    not just to brand-new ones — deliberately, so retroactively scoping
    the ~300 existing projects by department doesn't need a different
    tool than the one used to add new ones."""
    rows, header_error = read_project_rows(wb)
    if header_error:
        return {"added": 0, "department_links_added": 0, "skipped": [], "header_error": header_error}
    if len(rows) > MAX_ROWS:
        return {
            "added": 0, "department_links_added": 0, "skipped": [],
            "header_error": f"Sheet has {len(rows)} data rows — max is {MAX_ROWS} per upload. Split it into batches.",
        }

    project_id_by_name: Dict[str, int] = {
        name: pid for pid, name in db.execute(select(m.Project.id, m.Project.name)).all()
    }
    existing_dept_links = {
        (pid, dept) for pid, dept in db.execute(
            select(m.ProjectDepartment.project_id, m.ProjectDepartment.department)
        ).all()
    }

    new_names: List[str] = []  # ordered — projects to create this pass
    seen_new_names = set()
    seen_dept_pairs_in_file = set()
    to_link: List[Tuple[str, str]] = []  # (project_name, department)
    skipped = []

    for row_num, name, dept in rows:
        already_in_db = name in project_id_by_name
        already_in_file = name in seen_new_names
        is_new_this_row = not already_in_db and not already_in_file
        if is_new_this_row:
            seen_new_names.add(name)
            new_names.append(name)

        if not dept:
            if not is_new_this_row:
                reason = (
                    "Already on the list — skipped, nothing changed" if already_in_db
                    else "Duplicate row in this file — only added once"
                )
                skipped.append({"row": row_num, "name": name, "reason": reason})
            # else: brand-new project, no department on this row — nothing
            # more to report, it'll be created below
            continue

        pair = (name, dept)
        if pair in seen_dept_pairs_in_file:
            skipped.append({"row": row_num, "name": name, "reason": f"Duplicate row in this file — already linked to '{dept}'"})
            continue
        if already_in_db and (project_id_by_name[name], dept) in existing_dept_links:
            skipped.append({"row": row_num, "name": name, "reason": f"Already linked to '{dept}' — skipped, nothing changed"})
            continue
        seen_dept_pairs_in_file.add(pair)
        to_link.append(pair)

    for name in new_names:
        item = m.Project(name=name)
        db.add(item)
        db.flush()  # need item.id for the ProjectDepartment link below
        project_id_by_name[name] = item.id

    links_added = 0
    for name, dept in to_link:
        db.add(m.ProjectDepartment(project_id=project_id_by_name[name], department=dept, created_by="bulk upload"))
        links_added += 1

    if new_names or links_added:
        db.commit()
    return {
        "added": len(new_names), "department_links_added": links_added,
        "skipped": skipped, "header_error": None,
    }


def _process_task_upload(db, wb: Workbook) -> dict:
    rows, header_error = read_task_rows(wb)
    if header_error:
        return {"added": 0, "skipped": [], "header_error": header_error}
    if len(rows) > MAX_ROWS:
        return {
            "added": 0, "skipped": [],
            "header_error": f"Sheet has {len(rows)} data rows — max is {MAX_ROWS} per upload. Split it into batches.",
        }

    project_id_by_name: Dict[str, int] = {
        name: pid for pid, name in db.execute(select(m.Project.id, m.Project.name)).all()
    }
    task_id_by_name: Dict[str, int] = {
        name: tid for tid, name in db.execute(select(m.TaskType.id, m.TaskType.name)).all()
    }
    existing_links = {
        (tid, pid) for tid, pid in db.execute(select(m.ProjectTask.task_type_id, m.ProjectTask.project_id)).all()
    }

    seen_pairs_in_file = set()
    new_task_names: Dict[str, None] = {}  # ordered set: task names to create this pass
    to_link: List[Tuple[str, str]] = []  # (task_name, project_name) pairs to link
    skipped = []

    for row_num, task_name, project_name in rows:
        if not task_name:
            skipped.append({"row": row_num, "name": project_name or "(blank)", "reason": f"Missing {HEADERS['task']}"})
            continue
        if not project_name:
            skipped.append({"row": row_num, "name": task_name, "reason": f"Missing {TASK_PROJECT_HEADER}"})
            continue
        if project_name not in project_id_by_name:
            skipped.append({
                "row": row_num, "name": task_name,
                "reason": f"Project '{project_name}' not found — check spelling against the Projects sheet",
            })
            continue
        pair = (task_name, project_name)
        if pair in seen_pairs_in_file:
            skipped.append({"row": row_num, "name": task_name, "reason": f"Duplicate row in this file — only linked once to '{project_name}'"})
            continue
        task_id = task_id_by_name.get(task_name)
        if task_id is not None and (task_id, project_id_by_name[project_name]) in existing_links:
            skipped.append({"row": row_num, "name": task_name, "reason": f"Already linked to '{project_name}' — skipped, nothing changed"})
            continue
        seen_pairs_in_file.add(pair)
        if task_id is None and task_name not in new_task_names:
            new_task_names[task_name] = None
        to_link.append(pair)

    if not to_link:
        return {"added": 0, "skipped": skipped, "header_error": None}

    for name in new_task_names:
        item = m.TaskType(name=name)
        db.add(item)
        db.flush()  # need item.id below
        task_id_by_name[name] = item.id

    for task_name, project_name in to_link:
        db.add(m.ProjectTask(
            task_type_id=task_id_by_name[task_name],
            project_id=project_id_by_name[project_name],
            created_by="bulk upload",
        ))
    db.commit()
    return {"added": len(to_link), "skipped": skipped, "header_error": None}


def build_sample_workbook(kind: str, db=None) -> Workbook:
    """db (Ganesh, 2026-08-27, task kind; also project kind as of 2026-08-28
    for department-scoped projects) — real Project names (task kind) or
    real in-use department names (project kind) for each kind's own
    reference sheet, same reasoning as _add_projects_reference_sheet()'s
    docstring. Optional for both — kept optional rather than required so
    nothing upstream that already calls build_sample_workbook() without a
    session breaks; falls back to placeholder names when db is None."""
    if kind == "task":
        return _build_task_sample_workbook(db)
    header = HEADERS[kind]
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_TITLES[kind]
    ws.append([header, PROJECT_DEPARTMENT_HEADER])
    for c in ws[1]:
        c.font = Font(bold=True)
    # Department-scoped projects (Ganesh, 2026-08-28) — two example rows
    # for the SAME project, showing how to link one project to more than
    # one department (one row per department, not a comma-separated
    # cell), plus a third project with a blank Department cell showing
    # the "unrestricted, usable by everyone" default.
    ws.append(["Acme Corp.", "Sales"])
    ws.append(["Acme Corp.", "Support"])
    ws.append(["Bluepeak Consulting Inc.", ""])
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 24

    _add_departments_reference_sheet(wb, db)

    info = wb.create_sheet("Instructions")
    info.append(["Column", "Notes"])
    for c in info[1]:
        c.font = Font(bold=True)
    for row in [
        (header, "One project/employer name per row."),
        (PROJECT_DEPARTMENT_HEADER, "Optional. Leave blank for a project every department can use."),
        ("", ""),
        ("To restrict a project to more than one department, list it on", ""),
        ("more than one row — once per department (see the 'Acme Corp.' example).", ""),
        ("", ""),
        ("Add-only — this sheet never renames or removes a value.", ""),
        ("A project name already on the list is skipped (unless this file", ""),
        ("adds a NEW department link to it — that part still applies).", ""),
        ("To deactivate a value, use the Deactivate button on the", ""),
        ("Projects & Tasks page instead — it's not done via this sheet.", ""),
    ]:
        info.append(row)
    info.column_dimensions["A"].width = 46
    info.column_dimensions["B"].width = 50
    return wb


def _existing_project_names(db) -> List[str]:
    if db is None:
        return ["Acme Corp.", "Bluepeak Consulting Inc."]  # sample-only placeholders, no DB yet
    return [n for (n,) in db.execute(select(m.Project.name).where(m.Project.active.is_(True)).order_by(m.Project.name)).all()]


def _existing_department_names(db) -> List[str]:
    if db is None:
        return ["Sales", "Support"]  # sample-only placeholders, no DB yet
    return [d for d in reports.departments_list(db) if d != "—"]


def _add_departments_reference_sheet(wb: Workbook, db) -> None:
    """Same idea as _add_projects_reference_sheet() below, for the
    Projects sheet's own optional Department column (Ganesh, 2026-08-28)
    — department has no canonical table anywhere in this app (see
    reports.departments_list()), so this is just every department name
    currently in use by an active employee, for spelling reference only."""
    names = _existing_department_names(db)
    ref = wb.create_sheet("Departments")
    ref.append(["Departments currently in use — for reference only, not read by the upload"])
    ref[1][0].font = Font(bold=True)
    ref.append([PROJECT_DEPARTMENT_HEADER])
    for c in ref[2]:
        c.font = Font(bold=True)
    for n in names:
        ref.append([n])
    ref.column_dimensions["A"].width = 46


def _add_projects_reference_sheet(wb: Workbook, db) -> None:
    """Inserted BEFORE Instructions (Ganesh, 2026-08-27 — "before
    instructions sub sheet, we should have existed projects list sheet so
    that when we uploading in bulk we can cross check the exact name of
    project"). openpyxl always appends new sheets at the end, so this
    builds the workbook's sheet order by creating Instructions AFTER this
    one rather than moving sheets around post-hoc."""
    names = _existing_project_names(db)
    ref = wb.create_sheet("Projects")
    ref.append(["Existing Project / Employer names — for reference only, not read by the upload"])
    ref[1][0].font = Font(bold=True)
    ref.append([TASK_PROJECT_HEADER])
    for c in ref[2]:
        c.font = Font(bold=True)
    for n in names:
        ref.append([n])
    ref.column_dimensions["A"].width = 46


def _build_task_sample_workbook(db) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Tasks"
    ws.append([HEADERS["task"], TASK_PROJECT_HEADER])
    for c in ws[1]:
        c.font = Font(bold=True)
    ws.append(["File review", "Acme Corp."])
    ws.append(["File review", "Bluepeak Consulting Inc."])  # same task, second project — 2 rows, not 1
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 32

    _add_projects_reference_sheet(wb, db)

    info = wb.create_sheet("Instructions")
    info.append(["Column", "Notes"])
    for c in info[1]:
        c.font = Font(bold=True)
    for row in [
        (HEADERS["task"], "The task's name."),
        (TASK_PROJECT_HEADER, "Which existing Project/Employer this task is for — must match a name on the 'Projects' sheet exactly."),
        ("", ""),
        ("One row per Task + Project pair, not one row per task.", ""),
        ("To link the same task to more than one project, list it on", ""),
        ("more than one row — once per project (see the 'File review' example).", ""),
        ("", ""),
        ("Add-only — this sheet never renames or removes a value, and never", ""),
        ("creates a new Project (that's the Projects sheet's own upload).", ""),
        ("A task+project pair already linked is skipped, not duplicated.", ""),
        ("To deactivate a task, or to unlink it from a project, use the", ""),
        ("Projects & Tasks page instead — it's not done via this sheet.", ""),
    ]:
        info.append(row)
    info.column_dimensions["A"].width = 46
    info.column_dimensions["B"].width = 60
    return wb


def build_existing_workbook(db, kind: str) -> Workbook:
    """Every current value (active and inactive) for reference, so it's
    easy to see what's already on the list before uploading more.

    kind="project" (department-scoped projects, Ganesh, 2026-08-28): shows
    one row per existing project+department link, same shape the upload
    itself expects — a project with no department links at all
    (unrestricted, usable by every department — see ProjectDepartment's
    docstring in app/models.py) gets one row with a blank Department cell
    rather than being omitted, same "still visible even with nothing to
    show" convention kind="task" below already established.

    kind="task" (Ganesh, 2026-08-27): shows one row per existing
    task+project link, same shape the upload itself expects — a task with
    no links at all (unrestricted, usable under every project — see
    ProjectTask's docstring in app/models.py) gets one row with a blank
    Project Name rather than being omitted, so it's visible that the task
    exists even though it has nothing to cross-check against yet."""
    if kind == "project":
        header = HEADERS["project"]
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_TITLES["project"]
        ws.append([header, PROJECT_DEPARTMENT_HEADER])
        for c in ws[1]:
            c.font = Font(bold=True)
        project_rows = list(
            db.execute(select(m.Project.id, m.Project.name).where(m.Project.active.is_(True)).order_by(m.Project.name)).all()
        )
        dept_links: Dict[int, List[str]] = {}
        for pid, dept in db.execute(select(m.ProjectDepartment.project_id, m.ProjectDepartment.department)).all():
            dept_links.setdefault(pid, []).append(dept)
        for pid, pname in project_rows:
            depts = dept_links.get(pid)
            if not depts:
                ws.append([pname, ""])  # unrestricted — no links yet
            else:
                for dept in sorted(depts):
                    ws.append([pname, dept])
        ws.column_dimensions["A"].width = 42
        ws.column_dimensions["B"].width = 24
        _add_departments_reference_sheet(wb, db)
        return wb

    wb = Workbook()
    ws = wb.active
    ws.title = "Tasks"
    ws.append([HEADERS["task"], TASK_PROJECT_HEADER])
    for c in ws[1]:
        c.font = Font(bold=True)
    task_rows = list(
        db.execute(select(m.TaskType.id, m.TaskType.name).where(m.TaskType.active.is_(True)).order_by(m.TaskType.name)).all()
    )
    links: Dict[int, List[str]] = {}
    for tid, pname in db.execute(
        select(m.ProjectTask.task_type_id, m.Project.name)
        .join(m.Project, m.Project.id == m.ProjectTask.project_id)
        .where(m.Project.active.is_(True))
        .order_by(m.Project.name)
    ).all():
        links.setdefault(tid, []).append(pname)
    for tid, tname in task_rows:
        pnames = links.get(tid)
        if not pnames:
            ws.append([tname, ""])  # unrestricted — no links yet
        else:
            for pname in pnames:
                ws.append([tname, pname])
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 32

    _add_projects_reference_sheet(wb, db)
    return wb
